import os
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file, Response
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_mail import Mail, Message as MailMessage
from flask_babel import Babel, gettext, ngettext, get_locale
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import pytz
import qrcode
from io import BytesIO
import base64
import json
from dotenv import load_dotenv
import secrets
import cv2
import openpyxl
from openpyxl.styles import Font, Alignment
from pywebpush import webpush, WebPushException
import random
import string

load_dotenv()

# Türkiye saat dilimi (UTC+3)
TURKEY_TZ = pytz.timezone('Etc/GMT-3')

def get_turkey_time():
    """Türkiye saatini döndürür"""
    return datetime.now(TURKEY_TZ)

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///cafe_loyalty.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# UTF-8 encoding configuration for Turkish characters
app.config['JSON_AS_ASCII'] = False
app.config['JSON_SORT_KEYS'] = False

# CORS configuration for Flutter mobile app
CORS(app, 
     origins=["http://localhost:*", "http://127.0.0.1:*", "http://192.168.*.*:*"],
     methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
     allow_headers=["Content-Type", "Authorization", "Accept"],
     supports_credentials=True)

# Babel konfigürasyonu
app.config['LANGUAGES'] = {
    'tr': 'Türkçe',
    'en': 'English',
    'ru': 'Русский',
    'de': 'Deutsch'
}
app.config['BABEL_DEFAULT_LOCALE'] = 'tr'
app.config['BABEL_DEFAULT_TIMEZONE'] = 'Europe/Istanbul'

# Mail configuration
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'True').lower() == 'true'
app.config['MAIL_USE_SSL'] = os.environ.get('MAIL_USE_SSL', 'False').lower() == 'true'
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_USERNAME')
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
# Dış linkler için temel alan adı (örn. şifre sıfırlama linki)
app.config['BASE_URL'] = os.environ.get('BASE_URL', 'https://reevpoints.tr')

# Dosya yükleme için izin verilen uzantılar
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Upload klasörünü oluştur
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Flask-Mail init
mail = Mail(app)

# Configure UTF-8 encoding for email
app.config['MAIL_ASCII_ATTACHMENTS'] = False
app.config['MAIL_DEFAULT_CHARSET'] = 'utf-8'

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Bu sayfaya erişmek için giriş yapmalısınız.'
login_manager.login_message_category = 'info'

babel = Babel()

def get_locale():
    # 1. URL parametresinden dil kontrolü
    if request.args.get('lang'):
        session['language'] = request.args.get('lang')
    
    # 2. Session'dan dil kontrolü
    if 'language' in session:
        return session['language']
    
    # 3. Kullanıcı profilinden dil kontrolü
    if current_user.is_authenticated and hasattr(current_user, 'language') and current_user.language:
        return current_user.language
    
    # 4. Tarayıcı dilinden otomatik algılama
    return request.accept_languages.best_match(['tr', 'en', 'ru']) or 'tr'

babel.init_app(app, locale_selector=get_locale)

# Veritabanı Modelleri
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    phone = db.Column(db.String(20), nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    points = db.Column(db.Integer, default=0)
    is_admin = db.Column(db.Boolean, default=False)
    language = db.Column(db.String(5), default='tr')
    is_verified = db.Column(db.Boolean, default=False)
    verification_code = db.Column(db.String(6))
    auth_token = db.Column(db.String(100), unique=True, nullable=True)
    preferred_branch_id = db.Column(db.Integer, db.ForeignKey('branch.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=get_turkey_time)
    updated_at = db.Column(db.DateTime, default=get_turkey_time, onupdate=get_turkey_time)
    
    # İlişkiler
    transactions = db.relationship('Transaction', backref='user', lazy=True)
    customer_qrs = db.relationship('CustomerQR', backref='customer', lazy=True)
    preferred_branch = db.relationship('Branch', backref='customers')
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    def generate_verification_code(self):
        self.verification_code = ''.join(random.choices(string.digits, k=6))
        return self.verification_code
    
    def generate_auth_token(self):
        self.auth_token = secrets.token_urlsafe(32)
        return self.auth_token

class Branch(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    address = db.Column(db.String(200))
    phone = db.Column(db.String(20))
    email = db.Column(db.String(100), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    image = db.Column(db.String(200))  # Şube görseli dosya adı
    working_hours = db.Column(db.String(100))  # Çalışma saatleri
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=get_turkey_time)
    scanned_qrs = db.relationship('CustomerQR', backref='scanning_branch', lazy=True)

class Transaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    points_earned = db.Column(db.Integer, default=0)
    points_used = db.Column(db.Integer, default=0)
    transaction_type = db.Column(db.String(20), nullable=False)  # 'purchase', 'redeem'
    description = db.Column(db.String(200))
    timestamp = db.Column(db.DateTime, default=get_turkey_time)

# Eski CustomerQR modeli (mevcut kod uyumluluğu için)
class CustomerQR(db.Model):
    __tablename__ = 'customer_qr'
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(100), unique=True, nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    points_earned = db.Column(db.Integer, default=1)  # Her QR kod 1 puan
    is_used = db.Column(db.Boolean, default=False)
    used_by_branch_id = db.Column(db.Integer, db.ForeignKey('branch.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=get_turkey_time)
    used_at = db.Column(db.DateTime, nullable=True)

class Category(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False, unique=True)
    description = db.Column(db.Text)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # İlişki
    products = db.relationship('Product', backref='category_ref', lazy=True)

class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    points_required = db.Column(db.Integer, nullable=False)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=True)
    category = db.Column(db.String(50), default='Genel')  # Geçici uyumluluk için
    image_filename = db.Column(db.String(255))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # İlişkiler
    redemptions = db.relationship('ProductRedemption', backref='product', lazy=True)

class ProductRedemption(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    points_used = db.Column(db.Integer, nullable=False)
    created_at = db.Column(db.DateTime, default=get_turkey_time)
    redeemed_at = db.Column(db.DateTime, default=get_turkey_time)
    is_confirmed = db.Column(db.Boolean, default=False)
    confirmed_at = db.Column(db.DateTime, nullable=True)
    confirmed_by_branch_id = db.Column(db.Integer, db.ForeignKey('branch.id'), nullable=True)
    confirmation_code = db.Column(db.String(255), nullable=True)
    qr_code = db.Column(db.String(255), nullable=True)
    
    # İlişkiler
    user = db.relationship('User', backref=db.backref('product_redemptions',
                                                     lazy='dynamic'))
    confirmed_by_branch = db.relationship('Branch', foreign_keys=[confirmed_by_branch_id])

class ProductRating(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    rating = db.Column(db.Integer, nullable=False)  # 1-5 arası puan
    comment = db.Column(db.Text, nullable=True)  # İsteğe bağlı yorum
    created_at = db.Column(db.DateTime, default=get_turkey_time)
    updated_at = db.Column(db.DateTime, nullable=True)
    
    # İlişkiler
    user = db.relationship('User', backref=db.backref('product_ratings', lazy='dynamic'))
    product = db.relationship('Product', backref=db.backref('ratings', lazy='dynamic'))
    
    # Unique constraint - bir kullanıcı bir ürünü sadece bir kez puanlayabilir
    __table_args__ = (db.UniqueConstraint('user_id', 'product_id', name='unique_user_product_rating'),)

class Campaign(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=False)
    image_filename = db.Column(db.String(200))
    is_active = db.Column(db.Boolean, default=True)
    start_date = db.Column(db.DateTime, nullable=False)
    end_date = db.Column(db.DateTime, nullable=False)
    created_at = db.Column(db.DateTime, default=get_turkey_time)
    
    # Kampanya QR kod özellikleri
    max_usage_per_customer = db.Column(db.Integer, default=1)  # Müşteri başına maksimum kullanım
    total_usage_limit = db.Column(db.Integer, nullable=True)  # Toplam kullanım limiti (None = sınırsız)
    qr_enabled = db.Column(db.Boolean, default=True)  # QR kod kullanımı aktif mi
    
    # Şube ilişkisi (many-to-many)
    branches = db.relationship('Branch', secondary='campaign_branches', backref='campaigns')
    
    # Kampanya ürünleri ilişkisi
    products = db.relationship('CampaignProduct', backref='campaign', lazy='dynamic', cascade='all, delete-orphan')
    
    # Kampanya kullanımları ilişkisi
    usages = db.relationship('CampaignUsage', backref='campaign', lazy='dynamic', cascade='all, delete-orphan')
    
    def is_valid(self):
        now = get_turkey_time().replace(tzinfo=None)  # Timezone bilgisini kaldır
        return self.is_active and self.start_date <= now <= self.end_date
    
    def get_usage_count(self):
        """Toplam kullanım sayısını döndür"""
        return self.usages.filter_by(is_used=True).count()
    
    def get_customer_usage_count(self, customer_id):
        """Belirli bir müşterinin kullanım sayısını döndür"""
        return self.usages.filter_by(customer_id=customer_id, is_used=True).count()
    
    def can_be_used_by_customer(self, customer_id, branch_id=None):
        """Müşteri bu kampanyayı kullanabilir mi?"""
        if not self.is_valid() or not self.qr_enabled:
            return False
        
        # Şube kontrolü - kampanya bu şubede geçerli mi?
        if branch_id and self.branches:
            valid_branch_ids = [branch.id for branch in self.branches]
            if branch_id not in valid_branch_ids:
                return False
        
        # Müşteri başına limit kontrolü
        customer_usage = self.get_customer_usage_count(customer_id)
        if customer_usage >= self.max_usage_per_customer:
            return False
        
        # Toplam limit kontrolü
        if self.total_usage_limit:
            total_usage = self.get_usage_count()
            if total_usage >= self.total_usage_limit:
                return False
        
        return True

# Kampanya-Şube ilişki tablosu
campaign_branches = db.Table('campaign_branches',
    db.Column('campaign_id', db.Integer, db.ForeignKey('campaign.id'), primary_key=True),
    db.Column('branch_id', db.Integer, db.ForeignKey('branch.id'), primary_key=True)
)

# Kampanya Ürünleri Modeli
class CampaignProduct(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    campaign_id = db.Column(db.Integer, db.ForeignKey('campaign.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=True)  # Mevcut ürünlere bağlantı
    product_name = db.Column(db.String(200), nullable=True)
    product_description = db.Column(db.Text)
    discount_type = db.Column(db.String(20), default='percentage')  # 'percentage', 'fixed', 'free'
    discount_value = db.Column(db.Float, default=0)  # İndirim miktarı
    discount = db.Column(db.Float, default=0)  # Yeni indirim alanı
    original_price = db.Column(db.Float, nullable=True)  # Orijinal fiyat
    campaign_price = db.Column(db.Float, nullable=True)  # Kampanya fiyatı
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=get_turkey_time)
    
    # Ürün ilişkisi
    product = db.relationship('Product', backref='campaign_products')

# Kampanya Kullanım Takibi Modeli
class CampaignUsage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    campaign_id = db.Column(db.Integer, db.ForeignKey('campaign.id'), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    qr_code = db.Column(db.String(200), unique=True, nullable=False)  # Benzersiz QR kod
    is_used = db.Column(db.Boolean, default=False)  # Kullanıldı mı?
    used_at = db.Column(db.DateTime, nullable=True)  # Kullanım tarihi
    used_by_branch_id = db.Column(db.Integer, db.ForeignKey('branch.id'), nullable=True)  # Hangi şubede kullanıldı
    created_at = db.Column(db.DateTime, default=get_turkey_time)  # QR kod oluşturulma tarihi
    expires_at = db.Column(db.DateTime, nullable=True)  # QR kod son kullanma tarihi
    
    # Seçilen ürün bilgileri
    selected_campaign_product_id = db.Column(db.Integer, db.ForeignKey('campaign_product.id'), nullable=True)  # Seçilen kampanya ürünü
    selected_product_name = db.Column(db.String(200), nullable=True)  # Seçilen ürün adı (yedek olarak)
    selected_product_details = db.Column(db.Text, nullable=True)  # Seçilen ürün detayları (JSON formatında)
    
    # İlişkiler
    customer = db.relationship('User', backref='campaign_usages')
    used_by_branch = db.relationship('Branch', backref='campaign_usages')
    selected_campaign_product = db.relationship('CampaignProduct', backref='campaign_usages')
    
    def is_expired(self):
        """QR kod süresi dolmuş mu?"""
        if not self.expires_at:
            return False
        return get_turkey_time().replace(tzinfo=None) > self.expires_at
    
    def can_be_used(self, branch_id=None):
        """QR kod kullanılabilir mi?"""
        if self.is_used or self.is_expired():
            return False
        
        # Şube kontrolü - kampanya bu şubede geçerli mi?
        if branch_id and self.campaign:
            # Kampanyanın geçerli şubelerini kontrol et
            valid_branch_ids = [branch.id for branch in self.campaign.branches]
            if valid_branch_ids and branch_id not in valid_branch_ids:
                return False
        
        return True

# Mesaj Modeli
class Message(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    content = db.Column(db.Text, nullable=False)
    sender_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # Admin gönderirse null
    recipient_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    is_read = db.Column(db.Boolean, default=False)
    is_admin_message = db.Column(db.Boolean, default=False)  # Admin tarafından gönderilen mesaj mı?
    created_at = db.Column(db.DateTime, default=get_turkey_time)
    read_at = db.Column(db.DateTime, nullable=True)
    
    # İlişkiler
    sender = db.relationship('User', foreign_keys=[sender_id], backref='sent_messages')
    recipient = db.relationship('User', foreign_keys=[recipient_id], backref='received_messages')
    
    def mark_as_read(self):
        """Mesajı okundu olarak işaretle"""
        if not self.is_read:
            self.is_read = True
            self.read_at = get_turkey_time()
            db.session.commit()

# Şifre Sıfırlama İstek Modeli
class PasswordResetRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    email = db.Column(db.String(100), nullable=False)
    token = db.Column(db.String(64), unique=True, nullable=False)
    is_used = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=get_turkey_time)
    expires_at = db.Column(db.DateTime, nullable=False)
    used_at = db.Column(db.DateTime, nullable=True)
    
    user = db.relationship('User', backref='password_reset_requests')
    
    def is_expired(self):
        return get_turkey_time().replace(tzinfo=None) > self.expires_at

class CustomerQRCode(db.Model):
    __tablename__ = 'customer_qr'
    __table_args__ = {'extend_existing': True}
    
    # Sadece mevcut tablo alanları (yeni alanlar eklenmeyecek)
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(100), unique=True, nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    points_earned = db.Column(db.Integer, default=1)
    is_used = db.Column(db.Boolean, default=False)
    used_by_branch_id = db.Column(db.Integer, db.ForeignKey('branch.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=get_turkey_time)
    used_at = db.Column(db.DateTime, nullable=True)
    
    # Relationships (sadece mevcut alanlar için)
    user = db.relationship('User', backref='customer_qr_codes')
    used_by_branch = db.relationship('Branch', backref='scanned_qr_codes')
    
    def is_expired(self):
        # Mevcut tabloda expires_at alanı yok, bu yüzden her zaman False
        return False
    
    def is_valid(self):
        return not self.is_used and not self.is_expired()

class SiteSetting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True, nullable=False)
    value = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=get_turkey_time)
    updated_at = db.Column(db.DateTime, default=get_turkey_time, onupdate=get_turkey_time)

# Survey Models
class Survey(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=False)
    start_date = db.Column(db.DateTime, nullable=False)
    end_date = db.Column(db.DateTime, nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=get_turkey_time)
    updated_at = db.Column(db.DateTime, default=get_turkey_time, onupdate=get_turkey_time)
    
    # Relationships
    questions = db.relationship('SurveyQuestion', backref='survey', lazy=True, cascade='all, delete-orphan')
    responses = db.relationship('SurveyResponse', backref='survey', lazy=True, cascade='all, delete-orphan')
    
    def is_valid(self):
        now = get_turkey_time().replace(tzinfo=None)
        return (self.is_active and 
                self.start_date <= now <= self.end_date)
    
    def to_dict(self):
        return {
            'id': self.id,
            'title': self.title,
            'description': self.description,
            'start_date': self.start_date.isoformat() if self.start_date else None,
            'end_date': self.end_date.isoformat() if self.end_date else None,
            'is_active': self.is_active,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'response_count': len(self.responses) if self.responses else 0
        }

class SurveyQuestion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    survey_id = db.Column(db.Integer, db.ForeignKey('survey.id'), nullable=False)
    question_text = db.Column(db.Text, nullable=False)
    question_type = db.Column(db.String(50), nullable=False)  # rating, text, multiple_choice, yes_no
    question_order = db.Column(db.Integer, nullable=False)
    options = db.Column(db.Text)  # JSON string for multiple choice options
    is_required = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=get_turkey_time)
    
    # Relationships
    answers = db.relationship('SurveyAnswer', backref='question', lazy=True, cascade='all, delete-orphan')
    
    def get_options(self):
        if self.options:
            try:
                return json.loads(self.options)
            except:
                return []
        return []
    
    def set_options(self, options_list):
        if options_list:
            self.options = json.dumps(options_list)
        else:
            self.options = None
    
    def to_dict(self):
        return {
            'id': self.id,
            'survey_id': self.survey_id,
            'question_text': self.question_text,
            'question_type': self.question_type,
            'question_order': self.question_order,
            'options': self.get_options(),
            'is_required': self.is_required
        }

class SurveyResponse(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    survey_id = db.Column(db.Integer, db.ForeignKey('survey.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    completed_at = db.Column(db.DateTime, default=get_turkey_time)
    
    # Relationships
    answers = db.relationship('SurveyAnswer', backref='response', lazy=True, cascade='all, delete-orphan')
    user = db.relationship('User', backref='survey_responses')
    
    # Unique constraint - bir kullanıcı bir anketi sadece bir kez cevaplayabilir
    __table_args__ = (db.UniqueConstraint('survey_id', 'user_id', name='unique_survey_user_response'),)
    
    def to_dict(self):
        return {
            'id': self.id,
            'survey_id': self.survey_id,
            'user_id': self.user_id,
            'user_name': self.user.name if self.user else None,
            'completed_at': self.completed_at.isoformat() if self.completed_at else None,
            'answers': [answer.to_dict() for answer in self.answers]
        }

class SurveyAnswer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    response_id = db.Column(db.Integer, db.ForeignKey('survey_response.id'), nullable=False)
    question_id = db.Column(db.Integer, db.ForeignKey('survey_question.id'), nullable=False)
    answer_text = db.Column(db.Text)  # For text answers
    answer_rating = db.Column(db.Integer)  # For rating answers (1-5)
    answer_choice = db.Column(db.String(200))  # For multiple choice answers
    answer_boolean = db.Column(db.Boolean)  # For yes/no answers
    created_at = db.Column(db.DateTime, default=get_turkey_time)
    
    def to_dict(self):
        return {
            'id': self.id,
            'question_id': self.question_id,
            'question_text': self.question.question_text if self.question else None,
            'question_type': self.question.question_type if self.question else None,
            'answer_text': self.answer_text,
            'answer_rating': self.answer_rating,
            'answer_choice': self.answer_choice,
            'answer_boolean': self.answer_boolean
        }

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Şube giriş kontrolü
def is_branch_logged_in():
    return 'branch_id' in session

def get_current_branch():
    if is_branch_logged_in():
        return Branch.query.get(session['branch_id'])
    return None

# E-posta gönderme fonksiyonu
def send_verification_email(user):
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        import re
        
        # Clean user data completely - remove all non-ASCII characters
        def clean_text(text):
            if not text:
                return "User"
            # Replace Turkish characters with ASCII equivalents
            replacements = {
                'ş': 's', 'Ş': 'S', 'ğ': 'g', 'Ğ': 'G', 'ü': 'u', 'Ü': 'U',
                'ö': 'o', 'Ö': 'O', 'ç': 'c', 'Ç': 'C', 'ı': 'i', 'İ': 'I'
            }
            for tr_char, en_char in replacements.items():
                text = text.replace(tr_char, en_char)
            # Remove any remaining non-ASCII characters
            return re.sub(r'[^\x00-\x7F]+', '', text)
        
        clean_email = clean_text(user.email)
        clean_name = clean_text(user.name) if user.name else 'User'
        
        # Create message using standard email library
        msg = MIMEMultipart('alternative')
        msg['Subject'] = 'REEV Coffee - Account Verification'
        msg['From'] = app.config['MAIL_USERNAME']
        msg['To'] = clean_email
        
        # Create text content
        body_text = f'''Hello {clean_name},

Welcome to REEV Coffee! Please use the verification code below to activate your account:

Verification Code: {user.verification_code}

This code will expire in 15 minutes.

Best regards,
REEV Coffee Team'''
        
        # Create HTML content
        html_content = f'''
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background: linear-gradient(135deg, #8B4513, #D2691E); color: white; padding: 30px; text-align: center;">
                <h1>REEV COFFEE</h1>
                <h2>Account Verification</h2>
            </div>
            <div style="padding: 30px; background: #f8f9fa;">
                <p>Hello <strong>{clean_name}</strong>,</p>
                <p>Welcome to REEV COFFEE loyalty program! Please use the verification code below to activate your account:</p>
                
                <div style="text-align: center; margin: 30px 0;">
                    <div style="background: #fff; border: 2px dashed #8B4513; padding: 20px; display: inline-block; border-radius: 10px;">
                        <h1 style="color: #8B4513; margin: 0; font-size: 2.5rem; letter-spacing: 0.3em;">{user.verification_code}</h1>
                    </div>
                </div>
                
                <p>This code is valid for 15 minutes.</p>
                <p>If you did not create this account, you can safely ignore this email.</p>
                
                <hr style="margin: 30px 0;">
                <p style="color: #666; font-size: 0.9rem;">
                    REEV COFFEE Loyalty Program<br>
                    Earn points with every purchase, get free products with your points!
                </p>
            </div>
        </div>
        '''
        
        # Create MIME parts
        part1 = MIMEText(body_text, 'plain', 'ascii')
        part2 = MIMEText(html_content, 'html', 'ascii')
        
        msg.attach(part1)
        msg.attach(part2)
        
        # Send using SMTP
        server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
        server.starttls()
        server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
        
        print(f"Verification email sending to: {clean_email}")
        # Convert to string and encode as bytes
        text = msg.as_string().encode('ascii', 'ignore').decode('ascii')
        server.sendmail(app.config['MAIL_USERNAME'], [clean_email], text)
        server.quit()
        
        print("Verification email sent successfully!")
        return True
        
    except Exception as e:
        print(f"Verification email error: {e}")
        print(f"Error type: {type(e).__name__}")
        print(f"Error detail: {str(e)}")
        print(f"SMTP Server: {app.config['MAIL_SERVER']}")
        print(f"SMTP Port: {app.config['MAIL_PORT']}")
        print(f"SMTP Username: {app.config['MAIL_USERNAME']}")
        print(f"TLS: {app.config['MAIL_USE_TLS']}")
        print(f"SSL: {app.config['MAIL_USE_SSL']}")
        
        # Add more detailed error information for debugging
        import traceback
        print("Detailed error trace:")
        traceback.print_exc()   
        return False

# Şifre sıfırlama e-postası gönderme
def send_password_reset_email(user, reset_code):
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        import re
        
        # Extract reset code from the parameter (it comes as "Reset Code: 123456")
        if "Reset Code: " in reset_code:
            code = reset_code.split("Reset Code: ")[1]
        else:
            code = reset_code
        
        # Clean user data completely - remove all non-ASCII characters
        def clean_text(text):
            if not text:
                return "User"
            # Replace Turkish characters with ASCII equivalents
            replacements = {
                'ş': 's', 'Ş': 'S', 'ğ': 'g', 'Ğ': 'G', 'ü': 'u', 'Ü': 'U',
                'ö': 'o', 'Ö': 'O', 'ç': 'c', 'Ç': 'C', 'ı': 'i', 'İ': 'I'
            }
            for tr_char, en_char in replacements.items():
                text = text.replace(tr_char, en_char)
            # Remove any remaining non-ASCII characters
            return re.sub(r'[^\x00-\x7F]+', '', text)
        
        clean_email = clean_text(user.email)
        clean_name = clean_text(user.name) if user.name else 'User'
        
        # Create message using standard email library
        msg = MIMEMultipart('alternative')
        msg['Subject'] = 'REEV Coffee - Password Reset'
        msg['From'] = app.config['MAIL_USERNAME']
        msg['To'] = clean_email
        
        # Create text content
        body_text = f'''Hello {clean_name},

Your password reset request has been received. Please use the code below in the mobile app to set your new password:

Reset Code: {code}

This code is valid for 15 minutes. If you did not make this request, you can ignore this email.

Best regards,
REEV Coffee Team'''
        
        # Create HTML content
        html_content = f'''
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background: linear-gradient(135deg, #8B4513, #D2691E); color: white; padding: 30px; text-align: center;">
                <h1>REEV COFFEE</h1>
                <h2>Password Reset</h2>
            </div>
            <div style="padding: 30px; background: #f8f9fa;">
                <p>Hello <strong>{clean_name}</strong>,</p>
                <p>Your password reset request has been received. Please use the code below in the mobile app to set your new password:</p>
                
                <div style="text-align: center; margin: 30px 0;">
                    <div style="background: #fff; border: 2px dashed #8B4513; padding: 20px; display: inline-block; border-radius: 10px;">
                        <h1 style="color: #8B4513; margin: 0; font-size: 2.5rem; letter-spacing: 0.3em;">{code}</h1>
                    </div>
                </div>
                
                <p>This code is valid for 15 minutes.</p>
                <p>If you did not make this request, you can ignore this email.</p>
                
                <hr style="margin: 30px 0;">
                <p style="color: #666; font-size: 0.9rem;">
                    REEV COFFEE Loyalty Program<br>
                    Track your points using the mobile app!
                </p>
            </div>
        </div>
        '''
        
        # Create MIME parts
        part1 = MIMEText(body_text, 'plain', 'ascii')
        part2 = MIMEText(html_content, 'html', 'ascii')
        
        msg.attach(part1)
        msg.attach(part2)
        
        # Send using SMTP
        server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
        server.starttls()
        server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
        
        print(f"Password reset email sending to: {clean_email}")
        # Convert to string and encode as bytes
        text = msg.as_string().encode('ascii', 'ignore').decode('ascii')
        server.sendmail(app.config['MAIL_USERNAME'], [clean_email], text)
        server.quit()
        
        print("Password reset email sent successfully!")
        return True
        
    except Exception as e:
        print(f"Password reset email error: {e}")
        # Add more detailed error information for debugging
        import traceback
        print("Detailed error info:")
        traceback.print_exc()
        return False

# Yeni kampanya e-postası gönderme (tüm kullanıcılara)
def send_campaign_email(campaign):
    try:
        # Admin olmayan tüm kullanıcılar
        recipients = User.query.filter_by(is_admin=False).all()
        if not recipients:
            return 0

        sent = 0
        for u in recipients:
            try:
                msg = MailMessage(
                    subject=f"REEV Coffee - Yeni Kampanya: {campaign.title}",
                    recipients=[u.email],
                    sender=app.config['MAIL_USERNAME']
                )
                msg.charset = 'utf-8'
                msg.extra_headers = {'Content-Type': 'text/html; charset=utf-8'}
                # Kısa açıklama
                desc = campaign.description or ''
                if len(desc) > 180:
                    desc = desc[:180] + '...'

                # Görsel URL'si (varsa)
                base = app.config.get('BASE_URL', 'https://reevpoints.tr').rstrip('/')
                image_block = ''
                image_text = ''
                if getattr(campaign, 'image_filename', None):
                    image_url = f"{base}/static/uploads/{campaign.image_filename}"
                    image_block = f'<div style="text-align:center;margin:12px 0 18px;"><img src="{image_url}" alt="{campaign.title}" style="max-width:100%;border-radius:8px;" /></div>'
                    image_text = f"Görsel: {image_url}\n\n"

                # Body (plain text)
                msg.body = (
                    f"Merhaba {u.name},\n\n"
                    f"Yeni kampanyamız yayında: {campaign.title}\n\n"
                    f"{desc}\n\n"
                    f"{image_text}"
                    f"Detaylar için Kampanyalar sayfasını ziyaret edin: {base}/campaigns\n\n"
                    f"Saygılarımızla,\nREEV Coffee"
                )
                # HTML
                msg.html = f'''
                <div style="font-family: Arial, sans-serif; max-width: 640px; margin:0 auto; background:#f8f9fa;">
                  <div style="background: linear-gradient(135deg,#8B4513,#D2691E); color:#fff; padding:22px; text-align:center;">
                    <h2 style="margin:0;">REEV COFFEE</h2>
                    <h3 style="margin:6px 0 0;">Yeni Kampanya</h3>
                  </div>
                  <div style="padding:20px;">
                    <p>Merhaba <strong>{u.name}</strong>,</p>
                    <p><strong>{campaign.title}</strong> kampanyamız yayında!</p>
                    <p style="color:#333; white-space:pre-wrap;">{desc}</p>
                    {image_block}
                    <div style="text-align:center;margin:14px 0;">
                      <a href="{base}/campaigns" target="_blank" style="background:#8B4513;color:#fff;text-decoration:none;padding:10px 16px;border-radius:6px;display:inline-block;">Kampanyalara Git</a>
                    </div>
                    <hr>
                    <p style="font-size: 0.9rem; color:#666;">Detaylar için uygulamadaki Kampanyalar sayfasını ziyaret edin.</p>
                  </div>
                </div>
                '''
                mail.send(msg)
                sent += 1
            except Exception as inner:
                print(f"Kampanya e-postası gönderilemedi ({u.email}): {inner}")
                continue

        print(f"✉️ Campaign email sent to {sent} users: {campaign.title}")
        return sent
    except Exception as e:
        print(f"Error in send_campaign_email: {e}")
        return 0

# Ana Sayfa
@app.context_processor
def inject_user():
    # Site logosu için global değişken
    site_logo = SiteSetting.query.filter_by(key='site_logo').first()
    return dict(current_user=current_user, site_logo=site_logo)

@app.route('/')
def index():
    # Arka plan resmini al
    site_background = SiteSetting.query.filter_by(key='site_background').first()
    # Site logosunu al
    site_logo = SiteSetting.query.filter_by(key='site_logo').first()
    return render_template('index.html', site_background=site_background, site_logo=site_logo)

# Kullanıcı Kayıt
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        phone = request.form.get('phone', '')
        password = request.form['password']
        consent = request.form.get('consent')
        
        # Açık rıza kontrolü
        if not consent:
            flash('Kişisel verilerin işlenmesi için açık rızanızı vermeniz gerekmektedir!', 'error')
            return redirect(url_for('register'))
        
        # E-posta kontrolü
        if User.query.filter_by(email=email).first():
            flash('Bu e-posta adresi zaten kullanılıyor!', 'error')
            return redirect(url_for('register'))
        
        # Yeni kullanıcı oluştur (e-posta doğrulaması gerekecek)
        user = User(
            name=name,
            email=email,
            phone=phone,
            is_verified=False
        )
        user.set_password(password)
        # Doğrulama kodu üret
        user.generate_verification_code()
        
        db.session.add(user)
        db.session.commit()
        
        # Doğrulama e-postası gönder
        if send_verification_email(user):
            flash('Kayıt başarılı! Lütfen e-posta adresinize gönderilen doğrulama kodunu girin.', 'success')
        else:
            flash('Kayıt oluşturuldu fakat doğrulama e-postası gönderilirken bir hata oluştu. Lütfen daha sonra tekrar deneyiniz.', 'error')
        
        # Doğrulama adımına yönlendir
        session['pending_user_id'] = user.id
        return redirect(url_for('verify_email'))
    
    return render_template('register.html')

# Kullanıcı Giriş
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        user = User.query.filter_by(email=email).first()
        
        if user and user.check_password(password):
            login_user(user)
            flash('Giriş başarılı!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('E-posta veya şifre hatalı!', 'error')
    
    return render_template('login.html')

# E-posta Doğrulama
@app.route('/verify_email', methods=['GET', 'POST'])
def verify_email():
    if 'pending_user_id' not in session:
        flash('Geçersiz erişim!', 'error')
        return redirect(url_for('register'))
    
    user = User.query.get(session['pending_user_id'])
    if not user:
        flash('Kullanıcı bulunamadı!', 'error')
        return redirect(url_for('register'))
    
    if request.method == 'POST':
        verification_code = request.form['verification_code']
        
        if user.verification_code == verification_code:
            user.is_verified = True
            user.verification_code = None
            db.session.commit()
            session.pop('pending_user_id', None)
            
            login_user(user)
            flash('E-posta doğrulandı! Hesabınız aktifleştirildi.', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Doğrulama kodu hatalı!', 'error')
    
    return render_template('verify_email.html', email=user.email)

# Doğrulama kodu tekrar gönder
@app.route('/resend_verification')
def resend_verification():
    if 'pending_user_id' not in session:
        flash('Geçersiz erişim!', 'error')
        return redirect(url_for('register'))
    
    user = User.query.get(session['pending_user_id'])
    if not user:
        flash('Kullanıcı bulunamadı!', 'error')
        return redirect(url_for('register'))
    
    user.generate_verification_code()
    db.session.commit()
    
    if send_verification_email(user):
        flash('Doğrulama kodu tekrar gönderildi!', 'success')
    else:
        flash('E-posta gönderilirken hata oluştu!', 'error')
    
    return redirect(url_for('verify_email'))

# Çıkış
@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Başarıyla çıkış yaptınız!', 'info')
    return redirect(url_for('index'))

# Şube Giriş
@app.route('/branch/login', methods=['GET', 'POST'])
def branch_login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        branch = Branch.query.filter_by(email=email, is_active=True).first()
        
        if branch and check_password_hash(branch.password_hash, password):
            session['branch_id'] = branch.id
            flash('Şube girişi başarılı!', 'success')
            return redirect(url_for('branch_panel'))
        else:
            flash('Geçersiz e-posta veya şifre!', 'error')
    
    return render_template('branch_login.html')

# Şube Çıkış
@app.route('/branch/logout')
def branch_logout():
    session.pop('branch_id', None)
    flash('Şube çıkışı yapıldı!', 'info')
    return redirect(url_for('branch_login'))

# Müşteri Dashboard
@app.route('/dashboard')
@login_required
def dashboard():
    # Son işlemleri getir
    recent_transactions = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.timestamp.desc()).limit(5).all()
    
    # Kampanya kullanımlarını getir (kullanılmış olanlar)
    campaign_usages = CampaignUsage.query.filter_by(
        customer_id=current_user.id,
        is_used=True
    ).order_by(CampaignUsage.used_at.desc()).limit(10).all()
    
    # Kampanya kullanımlarını işle
    campaign_activities = []
    for usage in campaign_usages:
        # Seçilen ürün bilgisini al
        selected_product_info = None
        if usage.selected_product_details:
            try:
                selected_product_info = json.loads(usage.selected_product_details)
            except:
                selected_product_info = None
        
        campaign_activities.append({
            'campaign': usage.campaign,
            'used_at': usage.used_at,
            'used_by_branch': usage.used_by_branch,
            'selected_product_name': usage.selected_product_name,
            'selected_product_info': selected_product_info
        })
    
    # Özel metrikler
    try:
        # Timezone uyumluluğu için naive datetime kullan
        last_30 = (get_turkey_time() - timedelta(days=30)).replace(tzinfo=None)
        # Puan kazanımları CustomerQR kullanımıyla gerçekleşiyor; bu nedenle CustomerQR üzerinden hesapla
        points_last_30 = db.session.query(db.func.coalesce(db.func.sum(CustomerQR.points_earned), 0)) \
            .filter(
                CustomerQR.customer_id == current_user.id,
                CustomerQR.is_used == True,
                CustomerQR.used_at >= last_30
            ).scalar() or 0
    except Exception:
        points_last_30 = 0

    # En çok gidilen şube (QR kullanımına göre)
    most_branch_name = None
    most_branch_count = 0
    try:
        from sqlalchemy import func
        rows = db.session.query(CustomerQR.used_by_branch_id, func.count(CustomerQR.id)) \
            .filter(CustomerQR.customer_id == current_user.id, CustomerQR.is_used == True, CustomerQR.used_by_branch_id != None) \
            .group_by(CustomerQR.used_by_branch_id) \
            .order_by(func.count(CustomerQR.id).desc()) \
            .limit(1).all()
        if rows:
            branch_id, cnt = rows[0]
            branch_obj = Branch.query.get(branch_id)
            most_branch_name = branch_obj.name if branch_obj else None
            most_branch_count = int(cnt)
    except Exception:
        pass

    # Okunmamış mesaj sayısını getir
    unread_messages_count = Message.query.filter_by(
        recipient_id=current_user.id,
        is_read=False
    ).count()
    
    # Ödül ilerleme: hedef 6 puan (isteğe göre ayarlanabilir)
    reward_target = 6
    user_points = current_user.points or 0
    # bir sonraki ödüle kalan puan (döngüsel)
    points_to_reward = reward_target - (user_points % reward_target)
    if points_to_reward == reward_target:
        points_to_reward = 0
    reward_progress_percent = int(round(((user_points % reward_target) / reward_target) * 100))

    # Son 30 gün günlük puan dağılımı (etiketler ve değerler)
    try:
        from collections import defaultdict
        daily_map = defaultdict(int)
        # Son 30 gün için başlangıç gününü hesapla (sadece tarih)
        start_day = (get_turkey_time() - timedelta(days=29)).date()
        end_day = get_turkey_time().date()
        # İlgili aralıkta kullanılan QR'ları çek
        used_qrs = CustomerQR.query.filter(
            CustomerQR.customer_id == current_user.id,
            CustomerQR.is_used == True,
            CustomerQR.used_at >= datetime.combine(start_day, datetime.min.time())
        ).all()
        for qr in used_qrs:
            d = (qr.used_at.date() if isinstance(qr.used_at, datetime) else qr.used_at)
            daily_map[d] += qr.points_earned or 0
        # Etiket ve değer listeleri
        points_daily_labels = []
        points_daily_values = []
        day = start_day
        while day <= end_day:
            points_daily_labels.append(day.strftime('%d.%m'))
            points_daily_values.append(daily_map.get(day, 0))
            day += timedelta(days=1)
    except Exception:
        points_daily_labels = []
        points_daily_values = []
    
    return render_template('dashboard.html', 
                         transactions=recent_transactions,
                         campaign_activities=campaign_activities,
                         points_last_30=points_last_30,
                         most_branch_name=most_branch_name,
                         most_branch_count=most_branch_count,
                         reward_target=reward_target,
                         reward_progress_percent=reward_progress_percent,
                         points_to_reward=points_to_reward,
                         points_daily_labels=points_daily_labels,
                         points_daily_values=points_daily_values,
                         unread_messages_count=unread_messages_count)

# Kampanyalar Sayfası
@app.route('/campaigns')
@login_required
def campaigns():
    # Kullanıcının tercih ettiği şubeye göre kampanyaları filtrele
    if current_user.preferred_branch_id:
        # Kullanıcının şubesindeki aktif kampanyalar
        active_campaigns = Campaign.query.filter(
            Campaign.is_active == True,
            Campaign.branches.any(Branch.id == current_user.preferred_branch_id)
        ).order_by(Campaign.created_at.desc()).all()
    else:
        # Şube seçimi yapılmamışsa tüm kampanyalar
        active_campaigns = Campaign.query.filter_by(is_active=True).order_by(Campaign.created_at.desc()).all()
    
    # Geçerli kampanyaları filtrele
    valid_campaigns = [campaign for campaign in active_campaigns if campaign.is_valid()]
    
    # Her kampanya için ürün bilgilerini ekle
    campaigns_with_products = []
    for campaign in valid_campaigns:
        campaign_products = CampaignProduct.query.filter_by(
            campaign_id=campaign.id, 
            is_active=True
        ).all()
        
        products_data = []
        for cp in campaign_products:
            if cp.product_id and cp.product:
                # Mevcut ürün
                product_data = {
                    'id': cp.product.id,
                    'campaign_product_id': cp.id,  # CampaignProduct ID'si
                    'name': cp.product.name,
                    'description': cp.product.description,
                    'image_filename': cp.product.image_filename,
                    'category': cp.product.category if isinstance(cp.product.category, str) else (cp.product.category.name if cp.product.category else 'Kategori Yok'),
                    'original_points': int(cp.product.points_required),
                    'discount': cp.discount or 0,
                    'discount_type': cp.discount_type,
                    'source': 'existing_product'
                }
            else:
                # Manuel ürün
                product_data = {
                    'id': f'manual_{cp.id}',
                    'campaign_product_id': cp.id,  # CampaignProduct ID'si
                    'name': cp.product_name,
                    'description': cp.product_description or '',
                    'image_filename': None,
                    'category': 'Kampanya Ürünü',
                    'original_points': int(cp.original_price or 0),
                    'discount': cp.discount_value or 0,
                    'discount_type': cp.discount_type,
                    'source': 'manual_product'
                }
            
            # İndirimli fiyatı hesapla
            if product_data['discount_type'] == 'percentage':
                discounted_points = product_data['original_points'] * (1 - product_data['discount'] / 100)
            elif product_data['discount_type'] == 'fixed':
                discounted_points = max(0, product_data['original_points'] - product_data['discount'])
            else:  # free
                discounted_points = 0
            
            product_data['discounted_points'] = int(discounted_points)
            products_data.append(product_data)
        
        campaigns_with_products.append({
            'campaign': campaign,
            'products': products_data
        })
    
    return render_template('campaigns.html', campaigns_with_products=campaigns_with_products)

# Kampanya QR Kod Oluşturma
@app.route('/campaign/<int:campaign_id>/generate_qr', methods=['POST'])
@login_required
def generate_campaign_qr(campaign_id):
    try:
        campaign = Campaign.query.get_or_404(campaign_id)
        
        # Seçilen şube bilgisini al (opsiyonel - şube kontrolü için)
        data = request.get_json() if request.is_json else {}
        selected_branch_id = data.get('selected_branch_id')
        
        # Kampanya kullanılabilir mi kontrol et (şube kontrolü dahil)
        if not campaign.can_be_used_by_customer(current_user.id, branch_id=selected_branch_id):
            # Şube kontrolü başarısızsa detaylı hata mesajı ver
            if selected_branch_id and campaign.branches:
                valid_branch_names = [b.name for b in campaign.branches]
                return jsonify({
                    'success': False,
                    'error': f'Bu kampanya sadece şu şubelerde geçerlidir: {", ".join(valid_branch_names)}'
                }), 400
            else:
                return jsonify({
                    'success': False,
                    'error': 'Bu kampanyayı kullanma hakkınız bulunmuyor veya kampanya süresi dolmuş.'
                }), 400
        
        # Seçilen ürün bilgisini al
        data = request.get_json() if request.is_json else {}
        selected_product_id = data.get('selected_product_id')
        
        if not selected_product_id:
            return jsonify({
                'success': False,
                'error': 'Lütfen bir ürün seçiniz.'
            }), 400
        
        # Seçilen kampanya ürününü bul
        campaign_product = CampaignProduct.query.filter_by(
            id=selected_product_id,
            campaign_id=campaign_id,
            is_active=True
        ).first()
        
        if not campaign_product:
            return jsonify({
                'success': False,
                'error': 'Seçilen ürün bulunamadı veya bu kampanyaya ait değil.'
            }), 400
        
        # Ürün bilgilerini hazırla
        if campaign_product.product_id and campaign_product.product:
            # Mevcut ürün
            product_name = campaign_product.product.name
            product_details = {
                'type': 'existing_product',
                'product_id': campaign_product.product.id,
                'name': campaign_product.product.name,
                'description': campaign_product.product.description,
                'original_points': campaign_product.product.points_required,
                'discount_type': campaign_product.discount_type,
                'discount_value': campaign_product.discount,
                'category': campaign_product.product.category if isinstance(campaign_product.product.category, str) else (campaign_product.product.category.name if campaign_product.product.category else 'Kategori Yok')
            }
        else:
            # Manuel ürün
            product_name = campaign_product.product_name
            product_details = {
                'type': 'manual_product',
                'name': campaign_product.product_name,
                'description': campaign_product.product_description,
                'original_price': campaign_product.original_price,
                'campaign_price': campaign_product.campaign_price,
                'discount_type': campaign_product.discount_type,
                'discount_value': campaign_product.discount_value
            }
        
        # Benzersiz QR kod oluştur
        import uuid
        qr_code = f"CAMPAIGN{campaign_id}USER{current_user.id}{str(uuid.uuid4())[:8].upper()}"
        
        # QR kod son kullanma tarihi (kampanya bitiş tarihinden önce, maksimum 24 saat)
        from datetime import timedelta
        expires_at = min(
            campaign.end_date,
            get_turkey_time().replace(tzinfo=None) + timedelta(hours=24)
        )
        
        # Kampanya kullanımı kaydet
        campaign_usage = CampaignUsage(
            campaign_id=campaign_id,
            customer_id=current_user.id,
            qr_code=qr_code,
            expires_at=expires_at,
            selected_campaign_product_id=campaign_product.id,
            selected_product_name=product_name,
            selected_product_details=json.dumps(product_details, ensure_ascii=False)
        )
        
        # Aynı zamanda customer_qr tablosuna da kaydet
        customer_qr = CustomerQRCode(
            customer_id=current_user.id,
            code=qr_code,
            points_earned=1,  # Kampanya QR'ı için varsayılan puan
            campaign_id=campaign_id
        )
        
        db.session.add(campaign_usage)
        db.session.add(customer_qr)
        db.session.commit()
        
        # QR kod görselini oluştur
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(qr_code)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Base64'e çevir
        img_buffer = BytesIO()
        img.save(img_buffer, format='PNG')
        img_str = base64.b64encode(img_buffer.getvalue()).decode()
        
        return jsonify({
            'success': True,
            'qr_code': qr_code,
            'qr_image': f'data:image/png;base64,{img_str}',
            'expires_at': expires_at.strftime('%d.%m.%Y %H:%M'),
            'campaign_title': campaign.title,
            'selected_product': {
                'name': product_name,
                'details': product_details
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Kasiyer için Kampanya QR Kod Kullanımı
@app.route('/branch/use_campaign_qr', methods=['POST'])
def use_campaign_qr():
    # Şube giriş kontrolü
    if 'branch_id' not in session:
        return jsonify({'error': 'Şube girişi gereklidir'}), 403
    
    branch = Branch.query.get(session['branch_id'])
    if not branch:
        return jsonify({'error': 'Geçersiz şube'}), 403
    
    try:
        qr_code = request.json.get('qr_code')
        if not qr_code:
            return jsonify({'error': 'QR kod gereklidir'}), 400
        
        # QR kod kullanımını bul
        campaign_usage = CampaignUsage.query.filter_by(qr_code=qr_code).first()
        if not campaign_usage:
            return jsonify({'error': 'Geçersiz QR kod'}), 404
        
        # QR kod kullanılabilir mi kontrol et (şube kontrolü dahil)
        if not campaign_usage.can_be_used(branch_id=branch.id):
            if campaign_usage.is_used:
                return jsonify({'error': 'Bu QR kod daha önce kullanılmış'}), 400
            elif campaign_usage.is_expired():
                return jsonify({'error': 'QR kod süresi dolmuş'}), 400
            else:
                # Şube geçerliliği kontrolü
                campaign = campaign_usage.campaign
                if campaign and campaign.branches:
                    valid_branch_names = [b.name for b in campaign.branches]
                    return jsonify({
                        'error': f'Bu kampanya sadece şu şubelerde geçerlidir: {", ".join(valid_branch_names)}'
                    }), 400
                else:
                    return jsonify({'error': 'QR kod bu şubede kullanılamaz'}), 400
        
        # Kampanya bilgilerini al
        campaign = campaign_usage.campaign
        if not campaign.is_valid():
            return jsonify({'error': 'Kampanya artık geçerli değil'}), 400
        
        # QR kodu kullanıldı olarak işaretle
        campaign_usage.is_used = True
        campaign_usage.used_at = get_turkey_time().replace(tzinfo=None)
        campaign_usage.used_by_branch_id = branch.id
        
        db.session.commit()
        
        # Müşteriye kampanya kullanım bildirimi gönder
        send_push_notification(
            user_id=campaign_usage.customer_id,
            title=f"🎉 Kampanya Kullanıldı: {campaign.title}",
            body=f"{branch.name} şubesinde kampanyanızı kullandınız!",
            notification_type="campaign_usage",
            url="/campaigns"
        )
        
        # Kampanya ürünlerini al
        campaign_products = CampaignProduct.query.filter_by(
            campaign_id=campaign.id, 
            is_active=True
        ).all()
        
        products_data = []
        for cp in campaign_products:
            # Mevcut ürünlerden eklenen ürünler (product_id ile bağlantılı)
            if cp.product_id and cp.product:
                product = cp.product
                # İndirimli fiyatı hesapla
                if cp.discount_type == 'percentage':
                    discount_amount = (product.points_required * cp.discount) / 100
                    discounted_price = product.points_required - discount_amount
                    discount_text = f"%{int(cp.discount)} indirim"
                else:  # fixed
                    discounted_price = max(0, product.points_required - cp.discount)
                    discount_text = f"{int(cp.discount)} puan indirim"
                
                products_data.append({
                    'id': product.id,
                    'name': product.name,
                    'description': product.description or '',
                    'original_points': product.points_required,
                    'discounted_points': int(discounted_price),
                    'discount': cp.discount,
                    'discount_type': cp.discount_type,
                    'discount_text': discount_text,
                    'image_filename': product.image_filename,
                    'category': product.category if isinstance(product.category, str) else (product.category.name if product.category else 'Genel'),
                    'source': 'existing_product'
                })
            
            # Manuel eklenen kampanya ürünleri (sadece isim ve açıklama ile)
            elif cp.product_name:
                # Manuel ürünler için indirim hesaplama
                original_price = cp.original_price or 0
                if cp.discount_type == 'percentage' and cp.discount_value:
                    discount_amount = (original_price * cp.discount_value) / 100
                    discounted_price = original_price - discount_amount
                    discount_text = f"%{int(cp.discount_value)} indirim"
                elif cp.discount_type == 'fixed' and cp.discount_value:
                    discounted_price = max(0, original_price - cp.discount_value)
                    discount_text = f"{int(cp.discount_value)} TL indirim"
                else:
                    discounted_price = cp.campaign_price or original_price
                    discount_text = "Özel kampanya fiyatı"
                
                products_data.append({
                    'id': f'manual_{cp.id}',
                    'name': cp.product_name,
                    'description': cp.product_description or '',
                    'original_points': int(original_price),
                    'discounted_points': int(discounted_price),
                    'discount': cp.discount_value or 0,
                    'discount_type': cp.discount_type,
                    'discount_text': discount_text,
                    'image_filename': None,
                    'category': 'Kampanya Ürünü',
                    'source': 'manual_product'
                })
        
        # Müşteri bilgisini güvenli şekilde al
        customer = User.query.get(campaign_usage.customer_id)
        customer_name = customer.name if customer else 'Bilinmeyen Müşteri'
        
        # Seçilen ürün bilgisini hazırla
        selected_product_info = None
        if campaign_usage.selected_product_details:
            try:
                selected_product_info = json.loads(campaign_usage.selected_product_details)
                selected_product_info['selected_product_name'] = campaign_usage.selected_product_name
            except:
                selected_product_info = {
                    'name': campaign_usage.selected_product_name or 'Bilinmeyen Ürün',
                    'type': 'unknown'
                }
        
        return jsonify({
            'success': True,
            'message': 'Kampanya QR kodu başarıyla kullanıldı!',
            'campaign_title': campaign.title,
            'campaign_description': campaign.description,
            'customer_name': customer_name,
            'used_at': campaign_usage.used_at.strftime('%d.%m.%Y %H:%M'),
            'products': products_data,
            'selected_product': selected_product_info  # Seçilen ürün bilgisi
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Admin - Kampanya QR Kullanım Raporları
@app.route('/admin/campaign/<int:campaign_id>/qr_usage', methods=['GET'])
@login_required
def get_campaign_qr_usage(campaign_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Bu işlem için admin yetkisi gereklidir'}), 403
    
    try:
        campaign = Campaign.query.get_or_404(campaign_id)
        
        # QR kullanım verilerini al
        qr_usages = CampaignUsage.query.filter_by(campaign_id=campaign_id).all()
        
        # İstatistikleri hesapla
        total_qr_generated = len(qr_usages)
        total_qr_used = len([usage for usage in qr_usages if usage.is_used])
        total_qr_active = len([usage for usage in qr_usages if not usage.is_used and not usage.is_expired()])
        
        # QR kullanım listesi
        usage_list = []
        for usage in qr_usages:
            # Müşteri bilgisini güvenli şekilde al
            customer = User.query.get(usage.customer_id)
            customer_name = customer.name if customer else 'Bilinmeyen Müşteri'
            customer_email = customer.email if customer else 'Bilinmeyen Email'
            
            # Şube bilgisini güvenli şekilde al
            branch_name = None
            if usage.used_by_branch_id:
                branch = Branch.query.get(usage.used_by_branch_id)
                branch_name = branch.name if branch else 'Bilinmeyen Şube'
            
            usage_data = {
                'qr_code': usage.qr_code,
                'customer_name': customer_name,
                'customer_email': customer_email,
                'created_at': usage.created_at.strftime('%d.%m.%Y %H:%M'),
                'is_used': usage.is_used,
                'is_expired': usage.is_expired(),
                'used_at': usage.used_at.strftime('%d.%m.%Y %H:%M') if usage.used_at else None,
                'used_by_branch': branch_name
            }
            usage_list.append(usage_data)
        
        # En son kullanılanlar önce gelsin
        usage_list.sort(key=lambda x: x['created_at'], reverse=True)
        
        return jsonify({
            'success': True,
            'campaign': {
                'id': campaign.id,
                'title': campaign.title,
                'description': campaign.description
            },
            'total_qr_generated': total_qr_generated,
            'total_qr_used': total_qr_used,
            'total_qr_active': total_qr_active,
            'qr_usages': usage_list
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Müşteri QR Kod Oluşturma
@app.route('/create_my_qr', methods=['POST'])
@login_required
def create_my_qr():
    try:
        from datetime import datetime, timedelta
        
        # Son 10 dakikada oluşturulan QR kodlarını kontrol et
        ten_minutes_ago = get_turkey_time() - timedelta(minutes=10)
        recent_qrs = CustomerQR.query.filter(
            CustomerQR.customer_id == current_user.id,
            CustomerQR.created_at >= ten_minutes_ago
        ).count()
        
        # 10 dakikada maksimum 2 QR kod sınırı
        if recent_qrs >= 3:
            return jsonify({
                'success': False,
                'error': 'Son 10 dakika içerisinde maksimum 3 QR kod oluşturabilirsiniz. Lütfen biraz bekleyin.'
            }), 429
        
        # Benzersiz kod oluştur
        import uuid
        code = f"CUSTOMER{current_user.id}{str(uuid.uuid4())[:8].upper()}"
        
        customer_qr = CustomerQR(
            customer_id=current_user.id,
            code=code
        )
        
        db.session.add(customer_qr)
        db.session.commit()
        
        # QR kod görselini oluştur
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(code)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Base64'e çevir
        img_buffer = BytesIO()
        img.save(img_buffer, format='PNG')
        img_str = base64.b64encode(img_buffer.getvalue()).decode()
        
        return jsonify({
            'success': True,
            'code': code,
            'qr_image': f'data:image/png;base64,{img_str}',
            'customer_name': current_user.name,
            'remaining_qrs': 2 - recent_qrs - 1
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Puan ile Ürün Rezervasyonu
@app.route('/redeem_points', methods=['GET', 'POST'])
@login_required
def redeem_points():
    if request.method == 'POST':
        product_id = int(request.form['product_id'])
        product = Product.query.get_or_404(product_id)
        
        if not product.is_active:
            flash('Bu ürün şu anda mevcut değil!', 'error')
            return redirect(url_for('redeem_points'))
        
        if current_user.points < product.points_required:
            flash(f'Bu ürün için {product.points_required} puana ihtiyacınız var!', 'error')
            return redirect(url_for('redeem_points'))
        
        # Onay kodu oluştur
        import random
        import string
        confirmation_code = ''.join(random.choices(string.digits, k=6))
        
        # Ürün rezervasyonu (henüz puan düşmez)
        redemption = ProductRedemption(
            user_id=current_user.id,
            product_id=product.id,
            points_used=product.points_required,
            confirmation_code=confirmation_code,
            is_confirmed=False
        )
        
        db.session.add(redemption)
        db.session.commit()
        
        flash(f'{product.name} ürünü rezerve edildi! Onay Kodu: {confirmation_code}', 'success')
        return redirect(url_for('purchase_confirmation', redemption_id=redemption.id))
    
    # Aktif ürünleri ve kategorileri listele
    products = Product.query.filter_by(is_active=True).all()
    
    # Kategorileri Category tablosundan al
    db_categories = Category.query.filter_by(is_active=True).all()
    categories = {}
    
    # Her kategori için boş liste oluştur
    for cat in db_categories:
        categories[cat.name] = []
    
    # Ürünleri kategorilere dağıt
    for product in products:
        if product.category_id and product.category_ref:
            category_name = product.category_ref.name
        else:
            category_name = product.category or 'Genel'
        
        if category_name not in categories:
            categories[category_name] = []
        categories[category_name].append(product)
    
    # Boş kategorileri kaldır
    categories = {k: v for k, v in categories.items() if v}
    
    return render_template('redeem.html', products=products, categories=categories)

@app.route('/purchase_history')
@login_required
def purchase_history():
    """Müşteri alım ve puan kazanım geçmişi sayfası"""
    # Tarih filtreleme parametreleri
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    activity_type = request.args.get('type', 'all')  # all, purchases, points
    page = request.args.get('page', 1, type=int)
    per_page = 15
    
    # Ürün alımları
    redemptions_query = current_user.product_redemptions
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            redemptions_query = redemptions_query.filter(ProductRedemption.redeemed_at >= start_dt)
        except ValueError:
            pass
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            end_dt = end_dt.replace(hour=23, minute=59, second=59)
            redemptions_query = redemptions_query.filter(ProductRedemption.redeemed_at <= end_dt)
        except ValueError:
            pass
    
    # Puan kazanımları
    points_query = CustomerQR.query.filter_by(customer_id=current_user.id, is_used=True)
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            points_query = points_query.filter(CustomerQR.used_at >= start_dt)
        except ValueError:
            pass
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            end_dt = end_dt.replace(hour=23, minute=59, second=59)
            points_query = points_query.filter(CustomerQR.used_at <= end_dt)
        except ValueError:
            pass
    
    # Kampanya kullanımları
    campaign_query = CampaignUsage.query.filter_by(customer_id=current_user.id, is_used=True)
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            campaign_query = campaign_query.filter(CampaignUsage.used_at >= start_dt)
        except ValueError:
            pass
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            end_dt = end_dt.replace(hour=23, minute=59, second=59)
            campaign_query = campaign_query.filter(CampaignUsage.used_at <= end_dt)
        except ValueError:
            pass
    
    # Aktivite tipine göre filtreleme
    if activity_type == 'purchases':
        redemptions = redemptions_query.paginate(page=page, per_page=per_page, error_out=False)
        point_earnings = None
        campaign_usages = None
    elif activity_type == 'points':
        redemptions = None
        point_earnings = points_query.order_by(CustomerQR.used_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
        campaign_usages = None
    elif activity_type == 'campaigns':
        redemptions = None
        point_earnings = None
        campaign_usages = campaign_query.order_by(CampaignUsage.used_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    else:
        # Tüm aktiviteler - sayfalama olmadan
        redemptions = redemptions_query.all()
        point_earnings = points_query.order_by(CustomerQR.used_at.desc()).all()
        campaign_usages = campaign_query.order_by(CampaignUsage.used_at.desc()).all()
    
    # İstatistikler
    total_redemptions = current_user.product_redemptions.count()
    confirmed_redemptions = current_user.product_redemptions.filter_by(is_confirmed=True).count()
    total_points_earned = CustomerQR.query.filter_by(customer_id=current_user.id, is_used=True).count()
    total_points_spent = db.session.query(db.func.sum(ProductRedemption.points_used)).filter_by(user_id=current_user.id).scalar() or 0
    total_campaign_usages = CampaignUsage.query.filter_by(customer_id=current_user.id, is_used=True).count()
    
    # Kampanya kullanımlarını işle (seçilen ürün bilgileri ile)
    processed_campaign_usages = []
    if campaign_usages:
        campaign_list = campaign_usages.items if hasattr(campaign_usages, 'items') else campaign_usages
        for usage in campaign_list:
            # Seçilen ürün bilgisini al
            selected_product_info = None
            if usage.selected_product_details:
                try:
                    selected_product_info = json.loads(usage.selected_product_details)
                except:
                    selected_product_info = None
            
            processed_campaign_usages.append({
                'usage': usage,
                'campaign': usage.campaign,
                'selected_product_name': usage.selected_product_name,
                'selected_product_info': selected_product_info,
                'used_by_branch': usage.used_by_branch
            })
    
    return render_template('purchase_history.html', 
                         redemptions=redemptions,
                         point_earnings=point_earnings,
                         campaign_usages=campaign_usages,
                         processed_campaign_usages=processed_campaign_usages,
                         activity_type=activity_type,
                         start_date=start_date,
                         end_date=end_date,
                         total_redemptions=total_redemptions,
                         confirmed_redemptions=confirmed_redemptions,
                         total_points_earned=total_points_earned,
                         total_points_spent=total_points_spent,
                         total_campaign_usages=total_campaign_usages)

# Satın Alma Onay Sayfası
@app.route('/purchase_confirmation/<int:redemption_id>')
@login_required
def purchase_confirmation(redemption_id):
    redemption = ProductRedemption.query.filter_by(id=redemption_id, user_id=current_user.id).first_or_404()
    
    # Onay kodu için QR kod oluştur
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(redemption.confirmation_code)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Base64'e çevir
    img_buffer = BytesIO()
    img.save(img_buffer, format='PNG')
    qr_image_base64 = base64.b64encode(img_buffer.getvalue()).decode()
    
    return render_template('purchase_confirmation.html', 
                         redemption=redemption,
                         qr_image=f'data:image/png;base64,{qr_image_base64}')

# Şube Paneli
@app.route('/branch/panel')
def branch_panel():
    if not is_branch_logged_in():
        flash('Lütfen şube girişi yapın!', 'error')
        return redirect(url_for('branch_login'))
    
    branch = get_current_branch()
    if not branch:
        flash('Şube bilgisi bulunamadı. Lütfen tekrar giriş yapın!', 'error')
        session.pop('branch_id', None)
        return redirect(url_for('branch_login'))
    
    scanned_qrs = CustomerQRCode.query.filter_by(used_by_branch_id=branch.id)\
                                     .order_by(CustomerQRCode.used_at.desc())\
                                     .limit(10).all()
    
    # Onay bekleyen ürün alımları
    pending_redemptions = ProductRedemption.query.filter_by(is_confirmed=False)\
                                                 .order_by(ProductRedemption.redeemed_at.desc())\
                                                 .limit(10).all()
    
    # Bu şube tarafından onaylanan ürünler
    confirmed_redemptions = ProductRedemption.query.filter_by(is_confirmed=True, confirmed_by_branch_id=branch.id)\
                                                   .order_by(ProductRedemption.confirmed_at.desc())\
                                                   .limit(20).all()
    
    return render_template('branch_panel.html', branch=branch, scanned_qrs=scanned_qrs, 
                          pending_redemptions=pending_redemptions, confirmed_redemptions=confirmed_redemptions)

# Şube Müşteri QR Okutma
@app.route('/scan_qr', methods=['POST'])
def scan_qr():
    try:
        # Şube giriş kontrolü
        if 'branch_id' not in session:
            return jsonify({'success': False, 'error': 'Şube girişi gerekli'})
        
        branch = Branch.query.get(session['branch_id'])
        if not branch:
            return jsonify({'success': False, 'error': 'Geçersiz şube'})
        
        data = request.get_json()
        qr_code = data.get('qr_code', '').strip()
        
        if not qr_code:
            return jsonify({'success': False, 'error': 'QR kod boş olamaz'})
        
        # QR kodunu kontrol et
        customer_qr = CustomerQR.query.filter_by(code=qr_code, is_used=False).first()
        
        if not customer_qr:
            return jsonify({'success': False, 'error': 'Geçersiz veya kullanılmış QR kod'})
        
        # Müşteri bilgilerini al
        customer = User.query.get(customer_qr.customer_id)
        if not customer:
            return jsonify({'success': False, 'error': 'Müşteri bulunamadı'})
        
        # QR kodu kullanılmış olarak işaretle
        customer_qr.is_used = True
        customer_qr.used_at = get_turkey_time()
        customer_qr.used_by_branch_id = branch.id
        
        # Müşteriye 1 puan ekle
        customer.points += 1
        
        # İşlemi kaydet
        db.session.commit()
        
        # Müşteriye puan kazanım bildirimi gönder
        send_push_notification(
            user_id=customer.id,
            title="🎯 Puan Kazandınız!",
            body=f"Tebrikler! 1 puan kazandınız. Toplam puanınız: {customer.points}",
            notification_type="points",
            url="/dashboard"
        )
        
        return jsonify({
            'success': True,
            'customer_name': customer.name,
            'customer_email': customer.email,
            'points_earned': 1,
            'total_points': customer.points,
            'message': f'{customer.name} adlı müşteriye 1 puan eklendi!'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/scan_qr_local', methods=['POST'])
@login_required
def scan_qr_local():
    """Yerel OpenCV + PyZbar ile QR kod okuma endpoint'i"""
    try:
        data = request.get_json()
        base64_image = data.get('image', '')
        
        if not base64_image:
            return jsonify({'success': False, 'error': 'Görüntü verisi bulunamadı'})
        
        # QR kod okuma
        result = scan_qr_from_base64(base64_image)
        
        if not result['success']:
            return jsonify(result)
        
        qr_code = result['data'].strip()
        
        # QR kodunu kontrol et
        customer_qr = CustomerQR.query.filter_by(code=qr_code, is_used=False).first()
        
        if not customer_qr:
            return jsonify({'success': False, 'error': 'Geçersiz veya kullanılmış QR kod'})
        
        # Müşteri bilgilerini al
        customer = User.query.get(customer_qr.customer_id)
        if not customer:
            return jsonify({'success': False, 'error': 'Müşteri bulunamadı'})
        
        # QR kodu kullanılmış olarak işaretle
        customer_qr.is_used = True
        customer_qr.used_at = get_turkey_time()
        customer_qr.used_by_branch = current_user.id
        
        # Müşteriye 1 puan ekle
        customer.points += 1
        
        # İşlemi kaydet
        db.session.commit()
        
        # Müşteriye puan kazanım bildirimi gönder
        send_push_notification(
            user_id=customer.id,
            title="🎯 Puan Kazandınız!",
            body=f"Tebrikler! 1 puan kazandınız. Toplam puanınız: {customer.points}",
            notification_type="points",
            url="/dashboard"
        )
        
        return jsonify({
            'success': True,
            'data': qr_code,
            'method': result.get('method', 'OpenCV + PyZbar'),
            'customer_name': customer.name,
            'customer_email': customer.email,
            'points_earned': 1,
            'total_points': customer.points,
            'message': f'{customer.name} adlı müşteriye 1 puan eklendi!'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

# Şube Ürün Onaylama
@app.route('/branch/confirm_product', methods=['POST'])
def branch_confirm_product():
    if not is_branch_logged_in():
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    branch = get_current_branch()
    confirmation_code = request.form['confirmation_code']
    
    # Onay kodunu bul
    redemption = ProductRedemption.query.filter_by(confirmation_code=confirmation_code, is_confirmed=False).first()
    
    if not redemption:
        return jsonify({'error': 'Geçersiz veya zaten onaylanmış kod!'}), 400
    
    # Müşterinin puanını kontrol et
    customer = redemption.user
    if customer.points < redemption.points_used:
        return jsonify({'error': 'Müşterinin yeterli puanı yok!'}), 400
    
    # Puanı düş ve ürün alımını onayla
    customer.points -= redemption.points_used
    redemption.is_confirmed = True
    redemption.confirmed_at = get_turkey_time()
    redemption.confirmed_by_branch_id = branch.id
    
    # İşlem kaydı oluştur
    transaction = Transaction(
        user_id=customer.id,
        amount=0,
        points_used=redemption.points_used,
        transaction_type='redeem',
        description=f'{redemption.product.name} ürünü {redemption.points_used} puan ile alındı - Onay Kodu: {confirmation_code}'
    )
    
    db.session.add(transaction)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'customer_name': redemption.user.name,
        'product_name': redemption.product.name,
        'points_used': redemption.points_used,
        'message': f'{redemption.product.name} ürünü başarıyla onaylandı!'
    })

# Logo yükleme endpoint'i
@app.route('/admin/upload_logo', methods=['POST'])
@login_required
def upload_logo():
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    if 'logo' not in request.files:
        return jsonify({'error': 'Dosya seçilmedi'}), 400
    
    file = request.files['logo']
    if file.filename == '':
        return jsonify({'error': 'Dosya seçilmedi'}), 400
    
    if file and allowed_file(file.filename):
        # Eski logoyu sil
        old_logo = SiteSetting.query.filter_by(key='site_logo').first()
        if old_logo and old_logo.value:
            old_logo_path = os.path.join(app.config['UPLOAD_FOLDER'], old_logo.value)
            if os.path.exists(old_logo_path):
                os.remove(old_logo_path)
        
        # Yeni logo dosyasını kaydet
        filename = secure_filename(file.filename)
        timestamp = str(int(datetime.now().timestamp()))
        filename = f"logo_{timestamp}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Veritabanında güncelle
        if old_logo:
            old_logo.value = filename
            old_logo.updated_at = get_turkey_time()
        else:
            logo_setting = SiteSetting(key='site_logo', value=filename)
            db.session.add(logo_setting)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Logo başarıyla yüklendi!',
            'logo_url': url_for('static', filename=f'uploads/{filename}')
        })
    
    return jsonify({'error': 'Geçersiz dosya formatı'}), 400

# Arka plan resmi yükleme endpoint'i
@app.route('/admin/upload_background', methods=['POST'])
@login_required
def upload_background():
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    if 'background' not in request.files:
        return jsonify({'error': 'Dosya seçilmedi'}), 400
    
    file = request.files['background']
    if file.filename == '':
        return jsonify({'error': 'Dosya seçilmedi'}), 400
    
    if file and allowed_file(file.filename):
        # Eski arka plan resmini sil
        old_background = SiteSetting.query.filter_by(key='site_background').first()
        if old_background and old_background.value:
            old_bg_path = os.path.join(app.config['UPLOAD_FOLDER'], old_background.value)
            if os.path.exists(old_bg_path):
                os.remove(old_bg_path)
        
        # Yeni arka plan dosyasını kaydet
        filename = secure_filename(file.filename)
        timestamp = str(int(datetime.now().timestamp()))
        filename = f"background_{timestamp}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Veritabanında güncelle
        if old_background:
            old_background.value = filename
            old_background.updated_at = get_turkey_time()
        else:
            bg_setting = SiteSetting(key='site_background', value=filename)
            db.session.add(bg_setting)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Arka plan resmi başarıyla yüklendi!',
            'background_url': url_for('static', filename=f'uploads/{filename}')
        })
    
    return jsonify({'error': 'Geçersiz dosya formatı'}), 400

# Splash resmi yükleme endpoint'i
@app.route('/admin/upload_splash', methods=['POST'])
@login_required
def upload_splash():
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    if 'splash' not in request.files:
        return jsonify({'error': 'Dosya seçilmedi'}), 400
    
    file = request.files['splash']
    if file.filename == '':
        return jsonify({'error': 'Dosya seçilmedi'}), 400
    
    if file and allowed_file(file.filename):
        # Eski splash resmini sil
        old_splash = SiteSetting.query.filter_by(key='splash_image').first()
        if old_splash and old_splash.value:
            old_splash_path = os.path.join(app.config['UPLOAD_FOLDER'], old_splash.value)
            if os.path.exists(old_splash_path):
                os.remove(old_splash_path)
        
        # Yeni splash dosyasını kaydet
        filename = secure_filename(file.filename)
        timestamp = str(int(datetime.now().timestamp()))
        filename = f"splash_{timestamp}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Veritabanında güncelle
        if old_splash:
            old_splash.value = filename
            old_splash.updated_at = get_turkey_time()
        else:
            splash_setting = SiteSetting(key='splash_image', value=filename)
            db.session.add(splash_setting)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Splash resmi başarıyla yüklendi!',
            'splash_url': url_for('static', filename=f'uploads/{filename}')
        })
    
    return jsonify({'error': 'Geçersiz dosya formatı'}), 400

# App icon yükleme endpoint'i
@app.route('/admin/upload_app_icon', methods=['POST'])
@login_required
def upload_app_icon():
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    if 'app_icon' not in request.files:
        return jsonify({'error': 'Dosya seçilmedi'}), 400
    
    file = request.files['app_icon']
    if file.filename == '':
        return jsonify({'error': 'Dosya seçilmedi'}), 400
    
    if file and allowed_file(file.filename):
        # Eski app icon'u sil
        old_app_icon = SiteSetting.query.filter_by(key='app_icon').first()
        if old_app_icon and old_app_icon.value:
            old_app_icon_path = os.path.join(app.config['UPLOAD_FOLDER'], old_app_icon.value)
            if os.path.exists(old_app_icon_path):
                os.remove(old_app_icon_path)
        
        # Yeni app icon dosyasını kaydet
        filename = secure_filename(file.filename)
        timestamp = str(int(datetime.now().timestamp()))
        filename = f"app_icon_{timestamp}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Veritabanında güncelle
        if old_app_icon:
            old_app_icon.value = filename
            old_app_icon.updated_at = get_turkey_time()
        else:
            app_icon_setting = SiteSetting(key='app_icon', value=filename)
            db.session.add(app_icon_setting)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'App icon başarıyla yüklendi!',
            'app_icon_url': url_for('static', filename=f'uploads/{filename}')
        })
    
    return jsonify({'error': 'Geçersiz dosya formatı'}), 400

# Login logo yükleme endpoint'i
@app.route('/admin/upload_login_logo', methods=['POST'])
@login_required
def upload_login_logo():
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    if 'login_logo' not in request.files:
        return jsonify({'error': 'Dosya seçilmedi'}), 400
    
    file = request.files['login_logo']
    if file.filename == '':
        return jsonify({'error': 'Dosya seçilmedi'}), 400
    
    if file and allowed_file(file.filename):
        # Eski login logo'yu sil
        old_login_logo = SiteSetting.query.filter_by(key='login_logo').first()
        if old_login_logo and old_login_logo.value:
            old_login_logo_path = os.path.join(app.config['UPLOAD_FOLDER'], old_login_logo.value)
            if os.path.exists(old_login_logo_path):
                os.remove(old_login_logo_path)
        
        # Yeni login logo dosyasını kaydet
        filename = secure_filename(file.filename)
        timestamp = str(int(datetime.now().timestamp()))
        filename = f"login_logo_{timestamp}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Veritabanında güncelle
        if old_login_logo:
            old_login_logo.value = filename
            old_login_logo.updated_at = get_turkey_time()
        else:
            login_logo_setting = SiteSetting(key='login_logo', value=filename)
            db.session.add(login_logo_setting)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Login logo başarıyla yüklendi!',
            'login_logo_url': url_for('static', filename=f'uploads/{filename}')
        })
    
    return jsonify({'error': 'Geçersiz dosya formatı'}), 400

# Login logo URL'ini döndüren API endpoint'i
@app.route('/api/login-logo', methods=['GET'])
def get_login_logo():
    login_logo_setting = SiteSetting.query.filter_by(key='login_logo').first()
    if login_logo_setting and login_logo_setting.value:
        login_logo_url = url_for('static', filename=f'uploads/{login_logo_setting.value}', _external=True)
        return jsonify({
            'success': True,
            'login_logo_url': login_logo_url
        })
    else:
        # Varsayılan logo döndür
        default_logo_url = url_for('static', filename='icons/icon-192.png', _external=True)
        return jsonify({
            'success': True,
            'login_logo_url': default_logo_url,
            'is_default': True
        })

# Kampanya aktif/pasif yapma
@app.route('/admin/toggle_campaign/<int:campaign_id>', methods=['POST'])
@login_required
def toggle_campaign(campaign_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    campaign = Campaign.query.get_or_404(campaign_id)
    campaign.is_active = not campaign.is_active
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': f'Kampanya {"aktif" if campaign.is_active else "pasif"} yapıldı!',
        'is_active': campaign.is_active
    })

# Kampanya silme
@app.route('/admin/delete_campaign/<int:campaign_id>', methods=['DELETE'])
@login_required
def delete_campaign(campaign_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    campaign = Campaign.query.get_or_404(campaign_id)
    
    # Kampanya görselini sil
    if campaign.image_filename:
        image_path = os.path.join(app.config['UPLOAD_FOLDER'], campaign.image_filename)
        if os.path.exists(image_path):
            os.remove(image_path)
    
    db.session.delete(campaign)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Kampanya başarıyla silindi!'
    })

# Kampanya düzenleme bilgilerini getir
@app.route('/admin/get_campaign/<int:campaign_id>')
@login_required
def get_campaign(campaign_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    campaign = Campaign.query.get_or_404(campaign_id)
    
    return jsonify({
        'id': campaign.id,
        'title': campaign.title,
        'description': campaign.description,
        'start_date': campaign.start_date.strftime('%Y-%m-%dT%H:%M'),
        'end_date': campaign.end_date.strftime('%Y-%m-%dT%H:%M'),
        'is_active': campaign.is_active,
        'image_filename': campaign.image_filename,
        'branches': [branch.id for branch in campaign.branches]
    })

# Kampanya güncelleme
@app.route('/admin/update_campaign/<int:campaign_id>', methods=['POST'])
@login_required
def update_campaign(campaign_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    campaign = Campaign.query.get_or_404(campaign_id)
    
    try:
        # Form verilerini al
        title = request.form.get('title')
        description = request.form.get('description')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        branches_json = request.form.get('branches')
        
        # Tarih formatını kontrol et ve dönüştür
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%dT%H:%M')
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%dT%H:%M')
        
        # Kampanya bilgilerini güncelle
        campaign.title = title
        campaign.description = description
        campaign.start_date = start_date_obj
        campaign.end_date = end_date_obj
        campaign.updated_at = get_turkey_time()
        
        # Yeni görsel yüklendiyse işle
        if 'image' in request.files and request.files['image'].filename != '':
            file = request.files['image']
            if file and allowed_file(file.filename):
                # Eski görseli sil
                if campaign.image_filename:
                    old_image_path = os.path.join(app.config['UPLOAD_FOLDER'], campaign.image_filename)
                    if os.path.exists(old_image_path):
                        os.remove(old_image_path)
                
                # Yeni görseli kaydet
                filename = secure_filename(file.filename)
                timestamp = str(int(datetime.now().timestamp()))
                filename = f"campaign_{timestamp}_{filename}"
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)
                campaign.image_filename = filename
        
        # Şube ilişkilerini güncelle
        if branches_json:
            branch_ids = json.loads(branches_json)
            campaign.branches.clear()
            for branch_id in branch_ids:
                branch = Branch.query.get(int(branch_id))
                if branch:
                    campaign.branches.append(branch)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Kampanya başarıyla güncellendi!'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Kampanya güncellenirken hata oluştu: {str(e)}'}), 500

# Dil seçimi
@app.route('/set_language/<language>')
def set_language(language=None):
    if language and language in app.config['LANGUAGES']:
        session['language'] = language
        
        # Kullanıcı giriş yapmışsa dil tercihini kaydet
        if current_user.is_authenticated:
            current_user.language = language
            current_user.updated_at = get_turkey_time()
            db.session.commit()
    
    return redirect(request.referrer or url_for('index'))

# Basit çeviri sistemi
def _(text):
    """Basit çeviri fonksiyonu"""
    translations = {
        'tr': {
            'to reward':'İlerleme',
            'Rewards Progress':'Ödül İlerleme',
            'points':'puan',
            'Earn points by scanning QR codes to reach your next reward.':'QR kod okutarak kazanın devam et!',
            'Next reward target':'Bir sonraki ödül hedefi',
            'Points left':'Hediye ürün için kalan puan',
            'View Campaigns': 'Kampanyaları Görüntüle',
            'Single-use QR. Expires in 5 minutes':'Tek kullanımlık QR. 5 dakika sonra sona erer',
            'Last 5 listed':'En son listenlenen',
            'visits':'Ziyaret',
            'Most visited branch':'En çok ziyaret edilen şube',
            'Keep earning by scanning QR codes!': 'QR kod okutarak kazanın devam et!',
            'Points earned in last 30 days': 'Son 30 gün kazanan puanlar',
            'Here is a quick overview of your account': 'Hesabınız için hızlı bir genel bakış',
            'Points are valid for purchases of medium-sized products': 'Puanlar orta boy ürünler için geçerlidir',
            'Unread': 'Okunmamış',
            'Read': 'Okundu',
            'My Messages': 'Gelen Kutusu',
            'Messages': 'Mesajlarım',
            'Welcome back': 'Hoş Geldiniz',
            'Login': 'Giriş',
            'Register': 'Kayıt Ol',
            'Logout': 'Çıkış',
            'Dashboard': 'Panel',
            'Profile': 'Profil',
            'Admin': 'Yönetici',
            'Campaigns': 'Kampanyalar',
            'Join Now': 'Hemen Üye Ol',
            'Go to Dashboard': 'Panelime Git',
            'Earn Points with QR Code': 'QR Kod ile Puan Kazan',
            'Scan QR code at checkout and earn points': 'Alışveriş yaptığınızda QR kodu kasada okutturun ve puan kazanın',
            'Redeem Your Points': 'Puanlarınızı Kullanın',
            'Use your points for free products': 'Puanlarınızı ücretsiz ürünler için kullanın',
            'Special Campaigns': 'Özel Kampanyalar',
            'Take advantage of exclusive campaigns': 'Özel kampanyalardan yararlanın',
            'Earn points with every purchase, get free drinks with your points!': 'Her alışverişinizde puan kazanın, puanlarınızla ücretsiz içecek kazanın!',
            'Track Your Transactions': 'İşlemlerini Takip Et',
            'Track all your point earning and spending transactions in detail': 'Tüm puan kazanma ve kullanma işlemlerinizi detaylı olarak takip edin',
            'How It Works?': 'Nasıl Çalışır?',
            '1. Sign Up': '1. Üye Ol',
            'Quick and easy registration': 'Hızlı ve kolay kayıt',
            '2. Shop': '2. Alışveriş Yap',
            'Get your favorite drinks': 'Favori içeceklerini al',
            '3. Scan QR': '3. QR Okut',
            'Scan QR code to earn points': 'Puan kazanmak için QR kodu okut',
            '4. Earn Points': '4. Puan Kazan',
            'Use your points for discounts': 'İndirim için puanlarını kullan',
            'Email': 'E-posta',
            'Password': 'Şifre',
            'Don\'t have an account?': 'Hesabınız yok mu?',
            'Sign up': 'Kayıt olun',
            'Name': 'Ad Soyad',
            'Phone': 'Telefon',
            'Confirm Password': 'Şifre Tekrar',
            'Already have an account?': 'Zaten hesabınız var mı?',
            'My Points': 'Puanlarım',
            'QR Code': 'QR Kod',
            'Redeem Points': 'Puan Kullan',
            'My Profile': 'Profilim',
            'Edit Profile': 'Profili Düzenle',
            'Save Changes': 'Değişiklikleri Kaydet',
            'Current Password': 'Mevcut Şifre',
            'New Password': 'Yeni Şifre',
            'Change Password': 'Şifre Değiştir',
            'Phone (Optional)': 'Telefon (Opsiyonel)',
            'Personal Data Protection': 'Kişisel Verilerin Korunması',
            'CONSENT TEXT': 'AÇIK RIZA METNİ',
            'I have read and understood the consent text above': 'Yukarıdaki açık rıza metnini okudum, anladım ve kişisel verilerimin belirtilen amaçlarla işlenmesine açık rızamı veriyorum',
            'Account Summary': 'Hesap Özeti',
            'Current Points': 'Mevcut Puanınız',
            'Total Transactions': 'Toplam İşlem',
            'Membership Date': 'Üyelik Tarihi',
            'My QR Code': 'Benim QR Kodum',
            'Generate QR Code': 'QR Kodum Oluştur',
            'Quick Actions': 'Hızlı İşlemler',
            'Point System': 'Puan Sistemi',
            'Each QR code = 1 point': 'Her QR kod = 1 puan',
            'Minimum 5 points usage': 'Minimum 6 puan kullanılır',
            'Points never expire': 'Puanlar süresiz geçerli',
            'Show this code to cashier': 'Bu kodu kasiyere gösterin',
            'Time remaining': 'Kalan süre',
            'Your QR Code is Ready!': 'QR Kodunuz Hazır!',
            'Code': 'Kod',
            'Generate QR code to show to cashier': 'Kasiyere göstermek için QR kodunuzu oluşturun',
            'Profile Information': 'Profil Bilgileri',
            'Customer': 'Müşteri',
            'Full Name': 'Ad Soyad',
            'Phone': 'Telefon',
            'Registration Date': 'Kayıt Tarihi',
            'Not specified': 'Belirtilmemiş',
            'Unknown': 'Bilinmiyor',
            'My Points Status': 'Puan Durumum',
            'Point Value': 'Puan Değeri',
            'Total Purchases': 'Toplam Alım',
            'Confirm New Password': 'Yeni Şifre Tekrar',
            'At least 6 characters': 'En az 6 karakter olmalıdır',
            'Update Profile': 'Profil Güncelle',
            'Preferred Branch': 'Tercih Edilen Şube',
            'Select Branch': 'Şube Seç',
            'Campaigns will be filtered by this branch': 'Bu şube ile filtrelenen kampanyalar',
            'Generate QR': 'QR Kod Oluştur',
            'My Purchases': 'Alımlarım',
            'Passwords do not match': 'Şifreler uyuşmuyor',
            'Active Campaigns': 'Aktif Kampanyalar',
            'Don\'t miss current opportunities!': 'Anlık fırsatları kaçırmayın!',
            'Start': 'Başlangıç',
            'End': 'Bitiş',
            'Active Campaign': 'Aktif Kampanya',
            'Valid Branches': 'Geçerli Şubeler',
            'No active campaigns at the moment': 'Şu anda aktif kampanya yok',
            'Stay tuned for new campaigns!': 'Yeni kampanyalar için takip edin!',
            'Back to Home': 'Anasayfaya Dön',
            'Our Branches': 'Şubelerimiz',
            'You can find the same quality service and taste in all our branches': 'Tüm şubelerimizde aynı kaliteli servis ve tadı bulabilirsiniz',
            'Address': 'Address',
            'Phone': 'Telefon',
            'Working Hours': 'Çalışma Saatleri',
            'Active Branch': 'Aktif Şubeler',
            'Call': 'Ara',
            'Show on Map': 'Show on Map',
            'No Branches Yet': 'No Branches Yet',
            'We will be at your service with our new branches soon!': 'We will be at your service with our new branches soon!',
            'Information': 'Information',
            'Earn Points with QR Code': 'Earn Points with QR Code',
            'You can earn points by scanning your QR code at all our branches': 'You can earn points by scanning your QR code at all our branches',
            'Use Points': 'Use Points',
            'You can use your accumulated points at all our branches': 'You can use your accumulated points at all our branches',
            'Same Quality': 'Same Quality',
            'Whichever branch you go to, you get the same quality service': 'Whichever branch you go to, you get the same quality service',
            'Image not found': 'Image not found',
            'Active Campaigns': 'Aktif Kampanyalar',
            'Don\'t miss current opportunities!': 'Güncel fırsatları kaçırmayın!',
            'Start': 'Başlangıç',
            'End': 'Bitiş',
            'Active Campaign': 'Aktif Kampanya',
            'Valid Branches': 'Geçerli Şubeler',
            'No active campaigns at the moment': 'Şu anda aktif kampanya bulunmuyor',
            'Stay tuned for new campaigns!': 'Yeni kampanyalar için takipte kalın!',
            'Back to Home': 'Ana Sayfaya Dön',
            'Our Branches': 'Şubelerimiz',
            'You can find the same quality service and taste in all our branches': 'Tüm şubelerimizde aynı kaliteli hizmet ve lezzeti bulabilirsiniz',
            'Address': 'Adres',
            'Phone': 'Telefon',
            'Working Hours': 'Çalışma Saatleri',
            'Active Branch': 'Aktif Şube',
            'Call': 'Ara',
            'Show on Map': 'Haritada Göster',
            'No Branches Yet': 'Henüz Şube Bulunmuyor',
            'We will be at your service with our new branches soon!': 'Yakında yeni şubelerimizle hizmetinizdeyiz!',
            'Information': 'Bilgi',
            'Earn Points with QR Code': 'QR Kod ile Puan Kazan',
            'You can earn points by scanning your QR code at all our branches': 'Tüm şubelerimizde QR kodunuzu okutarak puan kazanabilirsiniz',
            'Use Points': 'Puan Kullan',
            'You can use your accumulated points at all our branches': 'Biriktirdiğiniz puanları tüm şubelerimizde kullanabilirsiniz',
            'Same Quality': 'Aynı Kalite',
            'Whichever branch you go to, you get the same quality service': 'Hangi şubeye giderseniz gidin, aynı kaliteli hizmeti alırsınız',
            'Image not found': 'Görsel bulunamadı',
            'Purchase History': 'Alımlarım & Puan Geçmişim',
            'View your product purchases and point earnings by date': 'Ürün alımları ve puan kazanımlarınızı tarih bazlı görüntüleyin',
            'Active Points': 'Aktif Puan',
            'Start Date': 'Başlangıç Tarihi',
            'End Date': 'Bitiş Tarihi',
            'Activity Type': 'Aktivite Tipi',
            'All Activities': 'Tüm Aktiviteler',
            'Only Product Purchases': 'Sadece Ürün Alımları',
            'Only Point Earnings': 'Sadece Puan Kazanımları',
            'Filter': 'Filtrele',
            'Clear': 'Temizle',
            'Total Purchases': 'Toplam Alım',
            'Product': 'Ürün',
            'Points Earned': 'Kazanılan Puan',
            'Points': 'Puan',
            'Points Spent': 'Harcanan Puan',
            'Confirmed': 'Onaylanan',
            'Purchase': 'Alım',
            'Product Purchases': 'Ürün Alımları',
            'Point Earnings': 'Puan Kazanımları',
            'Product Purchase': 'Ürün Alımı',
            'points spent': 'puan harcandı',
            'Confirmation Code': 'Onay Kodu',
            'Branch': 'Şube',
            'Unknown': 'Bilinmiyor',
            'Confirmed': 'Onaylandı',
            'Waiting': 'Bekliyor',
            'Point Earning': 'Puan Kazanımı',
            'You earned +1 point': '+1 puan kazandınız',
            'QR Code': 'QR Kod',
            'No Activity Yet': 'Henüz Aktivite Yok',
            'Earn points by scanning QR codes or buy products with your points.': 'QR kod okutarak puan kazanın veya puanlarınızla ürün alın.',
            'Generate QR': 'QR Oluştur',
            'Use Points': 'Puan Kullan',
            'Date': 'Tarih',
            'Used Points': 'Kullanılan Puan',
            'Status': 'Durum',
            'Not confirmed yet': 'Henüz onaylanmadı',
            'Earned Points': 'Kazanılan Puan',
            'No Product Purchases Yet': 'Henüz Ürün Alımınız Yok',
            'No Point Earnings Yet': 'Henüz Puan Kazanımınız Yok',
            'No Data Found': 'Veri Bulunamadı',
            'You can buy products using your points.': 'Puanlarınızı kullanarak ürün satın alabilirsiniz.',
            'You can earn points by scanning QR codes.': 'QR kod okutarak puan kazanabilirsiniz.',
            'No data found matching the selected criteria.': 'Seçilen kriterlere uygun veri bulunamadı.',
        
        # Redeem page translations
        'Points Available': 'Puan Mevcut',
        'Categories': 'Kategoriler',
        'All': 'Tümü',
        'Buy': 'Satın Al',
        'Insufficient Points': 'Yetersiz Puan',
        'more points needed': 'puan daha gerekli',
        'products displayed': 'ürün görüntüleniyor',
        'No Products Available Yet': 'Henüz Ürün Bulunmuyor',
        'Products will appear here when added by admin.': 'Admin tarafından ürünler eklendiğinde burada görünecek.',
        'Earn Points by Generating QR': 'QR Oluşturarak Puan Kazan',
        'Clear Filters': 'Filtreleri Temizle'
        },
        'en': {
            'Points are valid for purchases of medium-sized products.': 'Points are valid for purchases of medium-sized products.',
            'Welcome': 'Welcome',
            'Login': 'Login',
            'Register': 'Register',
            'Logout': 'Logout',
            'Dashboard': 'Dashboard',
            'Profile': 'Profile',
            'Admin': 'Admin',
            'Campaigns': 'Campaigns',
            'Branches': 'Branches',
            'Points': 'Points',
            'History': 'History',
            'Settings': 'Settings',
            'Language': 'Language',
            'Turkish': 'Turkish',
            'English': 'English',
            'Russian': 'Russian',
            'Welcome back': 'Welcome back',
            'Generate QR': 'Generate QR',
            'Use Points': 'Use Points',
            'Purchase History': 'Purchase History',
            'Our Branches': 'Our Branches',
            'Branch Login': 'Branch Login',
            'Join Now': 'Join Now',
            'Go to Dashboard': 'Go to Dashboard',
            'Earn Points with QR Code': 'Earn Points with QR Code',
            'Scan QR code at checkout and earn points': 'Scan QR code at checkout and earn points',
            'Redeem Your Points': 'Redeem Your Points',
            'Use your points for free products': 'Use your points for free products',
            'Special Campaigns': 'Special Campaigns',
            'Take advantage of exclusive campaigns': 'Take advantage of exclusive campaigns',
            'Earn points with every purchase, get free drinks with your points!': 'Earn points with every purchase, get free drinks with your points!',
            'Track Your Transactions': 'Track Your Transactions',
            'Track all your point earning and spending transactions in detail': 'Track all your point earning and spending transactions in detail',
            'How It Works?': 'How It Works?',
            '1. Sign Up': '1. Sign Up',
            'Quick and easy registration': 'Quick and easy registration',
            '2. Shop': '2. Shop',
            'Get your favorite drinks': 'Get your favorite drinks',
            '3. Scan QR': '3. Scan QR',
            'Scan QR code to earn points': 'Scan QR code to earn points',
            '4. Earn Points': '4. Earn Points',
            'Use your points for discounts': 'Use your points for discounts',
            'Email': 'Email',
            'Password': 'Password',
            'Don\'t have an account?': 'Don\'t have an account?',
            'Sign up': 'Sign up',
            'Name': 'Name',
            'Phone': 'Phone',
            'Confirm Password': 'Confirm Password',
            'Already have an account?': 'Already have an account?',
            'My Points': 'My Points',
            'QR Code': 'QR Code',
            'Redeem Points': 'Redeem Points',
            'My Profile': 'My Profile',
            'Edit Profile': 'Edit Profile',
            'Save Changes': 'Save Changes',
            'Current Password': 'Current Password',
            'New Password': 'New Password',
            'Change Password': 'Change Password',
            
            # Purchase history and redeem page translations
            'My Purchases & Point History': 'My Purchases & Point History',
            'View your product purchases and point earnings by date': 'View your product purchases and point earnings by date',
            'Active Points': 'Active Points',
            'Start Date': 'Start Date',
            'End Date': 'End Date',
            'Activity Type': 'Activity Type',
            'All Activities': 'All Activities',
            'Only Product Purchases': 'Only Product Purchases',
            'Only Point Earnings': 'Only Point Earnings',
            'Filter': 'Filter',
            'Clear': 'Clear',
            'Total Purchases': 'Total Purchases',
            'Product': 'Product',
            'Points Earned': 'Points Earned',
            'Points Spent': 'Points Spent',
            'Confirmed': 'Confirmed',
            'Purchase': 'Purchase',
            'Product Purchase': 'Product Purchase',
            'points spent': 'points spent',
            'Confirmation Code': 'Confirmation Code',
            'Branch': 'Branch',
            'Unknown': 'Unknown',
            'Waiting': 'Waiting',
            'Point Earning': 'Point Earning',
            'You earned +1 point': 'You earned +1 point',
            'No Activity Yet': 'No Activity Yet',
            'Generate QR': 'Generate QR',
            'Use Points': 'Use Points',
            'Date': 'Date',
            'Used Points': 'Used Points',
            'Status': 'Status',
            'Not confirmed yet': 'Not confirmed yet',
            'Earned Points': 'Earned Points',
            'No Product Purchases Yet': 'No Product Purchases Yet',
            'No Point Earnings Yet': 'No Point Earnings Yet',
            'No Data Found': 'No Data Found',
            'You can buy products using your points.': 'You can buy products using your points.',
            'You can earn points by scanning QR codes.': 'You can earn points by scanning QR codes.',
            'No data found matching the selected criteria.': 'No data found matching the selected criteria.',
            'Points Available': 'Points Available',
            'Categories': 'Categories',
            'All': 'All',
            'Buy': 'Buy',
            'Insufficient Points': 'Insufficient Points',
            'more points needed': 'more points needed',
            'products displayed': 'products displayed',
            'No Products Available Yet': 'No Products Available Yet',
            'Products will appear here when added by admin.': 'Products will appear here when added by admin.',
            'Earn Points by Generating QR': 'Earn Points by Generating QR',
            'Clear Filters': 'Clear Filters'
        },
        'ru': {
            'Welcome': 'Добро пожаловать',
            'Login': 'Войти',
            'Register': 'Регистрация',
            'Logout': 'Выйти',
            'Dashboard': 'Панель',
            'Profile': 'Профиль',
            'Admin': 'Админ',
            'Campaigns': 'Кампании',
            'Branches': 'Филиалы',
            'Points': 'Баллы',
            'History': 'История',
            'Settings': 'Настройки',
            'Language': 'Язык',
            'Turkish': 'Turkish',
            'English': 'English',
            'Russian': 'Russian',
            'German': 'German',
            'Welcome back': 'Добро пожаловать обратно',
            'Generate QR': 'Создать QR',
            'Use Points': 'Использовать баллы',
            'Purchase History': 'История покупок',
            'Our Branches': 'Наши филиалы',
            'Branch Login': 'Вход в филиал',
            'Join Now': 'Присоединиться сейчас',
            'Go to Dashboard': 'Перейти к панели',
            'Earn Points with QR Code': 'Зарабатывайте баллы с QR-кодом',
            'Scan QR code at checkout and earn points': 'Сканируйте QR-код при оплате и зарабатывайте баллы',
            'Redeem Your Points': 'Используйте свои баллы',
            'Use your points for free products': 'Используйте свои баллы для бесплатных продуктов',
            'Special Campaigns': 'Специальные кампании',
            'Take advantage of exclusive campaigns': 'Воспользуйтесь эксклюзивными кампаниями',
            'Earn points with every purchase, get free drinks with your points!': 'Зарабатывайте баллы с каждой покупкой, получайте бесплатные напитки за ваши баллы!',
            'Track Your Transactions': 'Отслеживайте ваши транзакции',
            'Track all your point earning and spending transactions in detail': 'Отслеживайте все ваши транзакции по заработку и трате баллов в деталях',
            'How It Works?': 'Как это работает?',
            '1. Sign Up': '1. Зарегистрируйтесь',
            'Quick and easy registration': 'Быстрая и простая регистрация',
            '2. Shop': '2. Покупайте',
            'Get your favorite drinks': 'Получите ваши любимые напитки',
            '3. Scan QR': '3. Сканируйте QR',
            'Scan QR code to earn points': 'Сканируйте QR-код для заработка баллов',
            '4. Earn Points': '4. Зарабатывайте баллы',
            'Use your points for discounts': 'Используйте ваши баллы для скидок',
            'Email': 'Электронная почта',
            'Password': 'Пароль',
            'Don\'t have an account?': 'Нет аккаунта?',
            'Sign up': 'Зарегистрироваться',
            'Name': 'Имя',
            'Phone': 'Телефон',
            'Confirm Password': 'Подтвердите пароль',
            'Already have an account?': 'Уже есть аккаунт?',
            'My Points': 'Мои баллы',
            'QR Code': 'QR-код',
            'Redeem Points': 'Использовать баллы',
            'My Profile': 'Мой профиль',
            'Edit Profile': 'Редактировать профиль',
            'Save Changes': 'Сохранить изменения',
            'Current Password': 'Текущий пароль',
            'New Password': 'Новый пароль',
            'Change Password': 'Изменить пароль',
            'Phone (Optional)': 'Телефон (Необязательно)',
            'Personal Data Protection': 'Защита персональных данных',
            'CONSENT TEXT': 'ТЕКСТ СОГЛАСИЯ',
            'I have read and understood the consent text above': 'Я прочитал и понял текст согласия выше и даю согласие на обработку моих персональных данных для указанных целей',
            'Account Summary': 'Сводка аккаунта',
            'Current Points': 'Текущие баллы',
            'Total Transactions': 'Всего транзакций',
            'Membership Date': 'Дата членства',
            'My QR Code': 'Мой QR-код',
            'Generate QR Code': 'Создать QR-код',
            'Quick Actions': 'Быстрые действия',
            'Point System': 'Система баллов',
            'Each QR code = 1 point': 'Каждый QR-код = 1 балл',
            'Minimum 10 points usage': 'Минимум 10 баллов для использования',
            'Points never expire': 'Баллы никогда не истекают',
            'Show this code to cashier': 'Покажите этот код кассиру',
            'Time remaining': 'Оставшееся время',
            'Your QR Code is Ready!': 'Ваш QR-код готов!',
            'Code': 'Код',
            'Generate QR code to show to cashier': 'Создайте QR-код для показа кассиру',
            'Profile Information': 'Информация профиля',
            'Customer': 'Клиент',
            'Full Name': 'Полное имя',
            'Phone': 'Телефон',
            'Registration Date': 'Дата регистрации',
            'Not specified': 'Не указано',
            'Unknown': 'Неизвестно',
            'My Points Status': 'Статус моих баллов',
            'Point Value': 'Стоимость баллов',
            'Total Purchases': 'Всего покупок',
            'Confirm New Password': 'Подтвердите новый пароль',
            'At least 6 characters': 'Минимум 6 символов',
            'Update Profile': 'Обновить профиль',
            'Preferred Branch': 'Предпочитаемый филиал',
            'Select Branch': 'Выберите филиал',
            'Campaigns will be filtered by this branch': 'Кампании будут отфильтрованы по этому филиалу',
            'Generate QR': 'Создать QR',
            'My Purchases': 'Мои покупки',
            'Passwords do not match': 'Пароли не совпадают',
            'Account Summary': 'Сводка аккаунта',
            'Current Points': 'Текущие баллы',
            'Total Transactions': 'Всего транзакций',
            'Membership Date': 'Дата членства',
            'My QR Code': 'Мой QR-код',
            'Generate QR Code': 'Создать QR-код',
            'Quick Actions': 'Быстрые действия',
            'Point System': 'Система баллов',
            'Each QR code = 1 point': 'Каждый QR-код = 1 балл',
            'Minimum 10 points usage': 'Минимум 10 баллов для использования',
            'Points never expire': 'Баллы никогда не истекают',
            'Show this code to cashier': 'Покажите этот код кассиру',
            'Time remaining': 'Оставшееся время',
            'Your QR Code is Ready!': 'Ваш QR-код готов!',
        },
        'de': {
            'Welcome': 'Willkommen',
            'Login': 'Anmelden',
            'Register': 'Registrieren',
            'Logout': 'Abmelden',
            'Dashboard': 'Dashboard',
            'Profile': 'Profil',
            'Admin': 'Admin',
            'Campaigns': 'Kampagnen',
            'Branches': 'Filialen',
            'Points': 'Punkte',
            'History': 'Verlauf',
            'Settings': 'Einstellungen',
            'Language': 'Sprache',
            'Turkish': 'Turkish',
            'English': 'English',
            'Russian': 'Russian',
            'German': 'German',
            'Home': 'Startseite',
            'About': 'Über uns',
            'Contact': 'Kontakt',
            'Services': 'Dienstleistungen',
            'Products': 'Produkte',
            'News': 'Nachrichten',
            'FAQ': 'FAQ',
            'Help': 'Hilfe',
            'Terms': 'Nutzungsbedingungen',
            'Privacy': 'Datenschutz',
            'Cookie Policy': 'Cookie-Richtlinie',
            'Copyright': 'Urheberrecht',
            'All Rights Reserved': 'Alle Rechte vorbehalten',
            'Loyalty Program': 'Treueprogramm',
            'Earn points with every purchase': 'Sammeln Sie Punkte bei jedem Einkauf',
            'Redeem points for rewards': 'Lösen Sie Punkte für Belohnungen ein',
            'Join now and start earning': 'Jetzt beitreten und sammeln beginnen',
            'How it works': 'So funktioniert es',
            '1. Sign Up': '1. Registrieren',
            'Quick and easy registration': 'Schnelle und einfache Registrierung',
            '2. Shop': '2. Einkaufen',
            'Get your favorite drinks': 'Holen Sie sich Ihre Lieblingsgetränke',
            '3. Scan QR': '3. QR scannen',
            'Scan QR code to earn points': 'QR-Code scannen, um Punkte zu sammeln',
            '4. Earn Points': '4. Punkte sammeln',
            'Use your points for discounts': 'Verwenden Sie Ihre Punkte für Rabatte',
            'Email': 'E-Mail',
            'Password': 'Passwort',
            'Don\'t have an account?': 'Haben Sie noch kein Konto?',
            'Sign up': 'Registrieren',
            'Name': 'Name',
            'Phone': 'Telefon',
            'Confirm Password': 'Passwort bestätigen',
            'Already have an account?': 'Haben Sie bereits ein Konto?',
            'My Points': 'Meine Punkte',
            'QR Code': 'QR-Code',
            'Redeem Points': 'Punkte einlösen',
            'My Profile': 'Mein Profil',
            'Edit Profile': 'Profil bearbeiten',
            'Save Changes': 'Änderungen speichern',
            'Current Password': 'Aktuelles Passwort',
            'New Password': 'Neues Passwort',
            'Change Password': 'Passwort ändern',
            
            # Purchase history and redeem page translations
            'My Purchases & Point History': 'Meine Einkäufe & Punkteverlauf',
            'View your product purchases and point earnings by date': 'Sehen Sie Ihre Produktkäufe und Punktesammlungen nach Datum',
            'Active Points': 'Aktive Punkte',
            'Start Date': 'Startdatum',
            'End Date': 'Enddatum',
            'Activity Type': 'Aktivitätstyp',
            'All Activities': 'Alle Aktivitäten',
            'Only Product Purchases': 'Nur Produktkäufe',
            'Only Point Earnings': 'Nur Punktesammlungen',
            'Filter': 'Filter',
            'Clear': 'Löschen',
            'Total Purchases': 'Gesamte Einkäufe',
            'Product': 'Produkt',
            'Points Earned': 'Gesammelte Punkte',
            'Points Spent': 'Ausgegebene Punkte',
            'Confirmed': 'Bestätigt',
            'Purchase': 'Einkauf',
            'Product Purchase': 'Produktkauf',
            'points spent': 'Punkte ausgegeben',
            'Confirmation Code': 'Bestätigungscode',
            'Branch': 'Filiale',
            'Unknown': 'Unbekannt',
            'Waiting': 'Wartend',
            'Point Earning': 'Punkte sammeln',
            'You earned +1 point': 'Sie haben +1 Punkt gesammelt',
            'No Activity Yet': 'Noch keine Aktivität',
            'Generate QR': 'QR generieren',
            'Use Points': 'Punkte verwenden',
            'Date': 'Datum',
            'Used Points': 'Verwendete Punkte',
            'Status': 'Status',
            'Not confirmed yet': 'Noch nicht bestätigt',
            'Earned Points': 'Gesammelte Punkte',
            'No Product Purchases Yet': 'Noch keine Produktkäufe',
            'No Point Earnings Yet': 'Noch keine Punktesammlungen',
            'No Data Found': 'Keine Daten gefunden',
            'You can buy products using your points.': 'Sie können Produkte mit Ihren Punkten kaufen.',
            'You can earn points by scanning QR codes.': 'Sie können Punkte durch Scannen von QR-Codes sammeln.',
            'No data found matching the selected criteria.': 'Keine Daten gefunden, die den ausgewählten Kriterien entsprechen.',
            'Points Available': 'Verfügbare Punkte',
            'Categories': 'Kategorien',
            'All': 'Alle',
            'Buy': 'Kaufen',
            'Insufficient Points': 'Unzureichende Punkte',
            'more points needed': 'weitere Punkte benötigt',
            'products displayed': 'Produkte angezeigt',
            'No Products Available Yet': 'Noch keine Produkte verfügbar',
            'Products will appear here when added by admin.': 'Produkte werden hier angezeigt, wenn sie vom Admin hinzugefügt werden.',
            'Earn Points by Generating QR': 'Punkte durch QR-Generierung sammeln',
            'Clear Filters': 'Filter löschen'
        }
    }
    
    current_lang = get_locale()
    return translations.get(current_lang, {}).get(text, text)

# Template context processor
@app.context_processor
def inject_conf_vars():
    return {
        'LANGUAGES': app.config['LANGUAGES'],
        'CURRENT_LANGUAGE': session.get('language', app.config['BABEL_DEFAULT_LOCALE']),
        'get_locale': get_locale,
        '_': _
    }

# Admin Paneli
@app.route('/admin')
@login_required
def admin():
    if not current_user.is_admin:
        flash('Bu sayfaya erişim yetkiniz yok!', 'error')
        return redirect(url_for('dashboard'))
    
    # Kullanıcılar için arama, sıralama, sayfalama parametreleri
    q = request.args.get('q', '').strip()
    sort = request.args.get('sort', 'created_at')  # name, email, points, created_at
    order = request.args.get('order', 'desc')      # asc, desc
    try:
        per_page = int(request.args.get('per_page', 20))
    except Exception:
        per_page = 20
    if per_page not in (10, 20, 50):
        per_page = 20
    try:
        page = int(request.args.get('page', 1))
    except Exception:
        page = 1

    # Sorguyu oluştur
    query = User.query
    if q:
        like = f"%{q}%"
        query = query.filter((User.name.ilike(like)) | (User.email.ilike(like)))

    # Sıralama
    sort_map = {
        'name': User.name,
        'email': User.email,
        'points': User.points,
        'created_at': User.created_at,
    }
    sort_col = sort_map.get(sort, User.created_at)
    if order == 'asc':
        query = query.order_by(sort_col.asc())
    else:
        query = query.order_by(sort_col.desc())

    users_page = query.paginate(page=page, per_page=per_page, error_out=False)
    users = users_page.items
    users_total = users_page.total
    branches = Branch.query.all()
    products = Product.query.all()
    campaigns = Campaign.query.order_by(Campaign.created_at.desc()).all()
    
    # Mevcut logoyu al
    current_logo = SiteSetting.query.filter_by(key='site_logo').first()
    current_background = SiteSetting.query.filter_by(key='background_image').first()
    current_splash = SiteSetting.query.filter_by(key='splash_image').first()
    current_app_icon = SiteSetting.query.filter_by(key='app_icon').first()
    current_login_logo = SiteSetting.query.filter_by(key='login_logo').first()
    
    return render_template('admin.html', 
        users=users, 
        users_page=users_page,
        users_total=users_total,
        branches=branches, 
        products=products, 
        campaigns=campaigns, 
        current_logo=current_logo, 
        current_background=current_background,
        current_splash=current_splash,
        current_app_icon=current_app_icon,
        current_login_logo=current_login_logo,
        q=q,
        sort=sort,
        order=order,
        per_page=per_page
    )

# Admin - Kampanya Ürün Yönetimi
@app.route('/admin/campaign/<int:campaign_id>/products', methods=['GET', 'POST'])
@login_required
def manage_campaign_products(campaign_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    campaign = Campaign.query.get_or_404(campaign_id)
    
    if request.method == 'POST':
        try:
            data = request.json
            
            # Yeni ürün ekle
            campaign_product = CampaignProduct(
                campaign_id=campaign_id,
                product_name=data['product_name'],
                product_description=data.get('product_description', ''),
                discount_type=data.get('discount_type', 'percentage'),
                discount_value=float(data.get('discount_value', 0)),
                original_price=float(data.get('original_price', 0)) if data.get('original_price') else None,
                campaign_price=float(data.get('campaign_price', 0)) if data.get('campaign_price') else None
            )
            
            db.session.add(campaign_product)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Kampanya ürünü başarıyla eklendi!',
                'product': {
                    'id': campaign_product.id,
                    'name': campaign_product.product_name,
                    'description': campaign_product.product_description,
                    'discount_type': campaign_product.discount_type,
                    'discount_value': campaign_product.discount_value,
                    'original_price': campaign_product.original_price,
                    'campaign_price': campaign_product.campaign_price
                }
            })
            
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # GET request - ürünleri listele
    products = campaign.products.filter_by(is_active=True).all()
    return jsonify({
        'campaign': {
            'id': campaign.id,
            'title': campaign.title,
            'description': campaign.description
        },
        'products': [{
            'id': p.id,
            'name': p.product_name,
            'description': p.product_description,
            'discount_type': p.discount_type,
            'discount_value': p.discount_value,
            'original_price': p.original_price,
            'campaign_price': p.campaign_price,
            'is_active': p.is_active
        } for p in products]
    })

# Admin - Kampanya Ürün Silme
@app.route('/admin/campaign_product/<int:product_id>/delete', methods=['DELETE'])
@login_required
def delete_campaign_product(product_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    try:
        product = CampaignProduct.query.get_or_404(product_id)
        product.is_active = False
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Kampanya ürünü başarıyla silindi!'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Admin - Kampanya Kullanım Limiti Güncelleme
@app.route('/admin/campaign/<int:campaign_id>/usage_limits', methods=['POST'])
@login_required
def update_campaign_usage_limits(campaign_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    try:
        campaign = Campaign.query.get_or_404(campaign_id)
        data = request.json
        
        # Kullanım limitlerini güncelle
        campaign.max_usage_per_customer = int(data.get('max_usage_per_customer', 1))
        campaign.total_usage_limit = int(data['total_usage_limit']) if data.get('total_usage_limit') else None
        campaign.qr_enabled = bool(data.get('qr_enabled', True))
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Kampanya kullanım limitleri başarıyla güncellendi!',
            'campaign': {
                'id': campaign.id,
                'title': campaign.title,
                'max_usage_per_customer': campaign.max_usage_per_customer,
                'total_usage_limit': campaign.total_usage_limit,
                'qr_enabled': campaign.qr_enabled,
                'current_usage': campaign.get_usage_count()
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Admin - Kampanya Kullanım Raporları
@app.route('/admin/campaign/<int:campaign_id>/usage_report')
@login_required
def campaign_usage_report(campaign_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    try:
        campaign = Campaign.query.get_or_404(campaign_id)
        
        # Kampanya kullanım istatistikleri
        total_generated = campaign.usages.count()
        total_used = campaign.usages.filter_by(is_used=True).count()
        total_expired = campaign.usages.filter(
            CampaignUsage.is_used == False,
            CampaignUsage.expires_at < get_turkey_time().replace(tzinfo=None)
        ).count()
        
        # Şube bazında kullanım
        branch_usage = db.session.query(
            Branch.name,
            db.func.count(CampaignUsage.id).label('usage_count')
        ).join(CampaignUsage, CampaignUsage.used_by_branch_id == Branch.id)\
        .filter(CampaignUsage.campaign_id == campaign_id, CampaignUsage.is_used == True)\
        .group_by(Branch.id).all()
        
        # Müşteri bazında kullanım
        customer_usage = db.session.query(
            User.name,
            User.email,
            db.func.count(CampaignUsage.id).label('usage_count'),
            db.func.max(CampaignUsage.used_at).label('last_used')
        ).join(CampaignUsage, CampaignUsage.customer_id == User.id)\
        .filter(CampaignUsage.campaign_id == campaign_id, CampaignUsage.is_used == True)\
        .group_by(User.id).all()
        
        return jsonify({
            'campaign': {
                'id': campaign.id,
                'title': campaign.title,
                'description': campaign.description,
                'max_usage_per_customer': campaign.max_usage_per_customer,
                'total_usage_limit': campaign.total_usage_limit
            },
            'statistics': {
                'total_generated': total_generated,
                'total_used': total_used,
                'total_expired': total_expired,
                'usage_rate': round((total_used / total_generated * 100) if total_generated > 0 else 0, 2)
            },
            'branch_usage': [{
                'branch_name': b.name,
                'usage_count': b.usage_count
            } for b in branch_usage],
            'customer_usage': [{
                'customer_name': c.name,
                'customer_email': c.email,
                'usage_count': c.usage_count,
                'last_used': c.last_used.strftime('%d.%m.%Y %H:%M') if c.last_used else None
            } for c in customer_usage]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Raporlama Sayfası
@app.route('/reports')
@login_required
def reports():
    if not current_user.is_admin:
        flash('Bu sayfaya erişim yetkiniz yok!', 'error')
        return redirect(url_for('dashboard'))
    
    branches = Branch.query.all()
    
    # Özet veriler
    total_points = db.session.query(db.func.sum(User.points)).scalar() or 0
    total_redemptions = ProductRedemption.query.count()
    active_customers = User.query.filter(User.points > 0).count()
    total_transactions = Transaction.query.count()
    
    # Kampanya verileri
    campaigns = Campaign.query.order_by(Campaign.created_at.desc()).all()
    campaign_stats = []
    
    for campaign in campaigns:
        # Kampanya QR kullanım istatistikleri
        total_qr_generated = CampaignUsage.query.filter_by(campaign_id=campaign.id).count()
        total_qr_used = CampaignUsage.query.filter_by(campaign_id=campaign.id, is_used=True).count()
        total_qr_active = CampaignUsage.query.filter(
            CampaignUsage.campaign_id == campaign.id,
            CampaignUsage.is_used == False
        ).count()
        
        # Kampanya ürün sayısı
        campaign_products_count = CampaignProduct.query.filter_by(
            campaign_id=campaign.id, 
            is_active=True
        ).count()
        
        # En çok kullanılan şube
        top_branch = db.session.query(
            Branch.name,
            db.func.count(CampaignUsage.id).label('usage_count')
        ).join(
            CampaignUsage, Branch.id == CampaignUsage.used_by_branch_id
        ).filter(
            CampaignUsage.campaign_id == campaign.id,
            CampaignUsage.is_used == True
        ).group_by(Branch.id, Branch.name).order_by(
            db.func.count(CampaignUsage.id).desc()
        ).first()
        
        campaign_stats.append({
            'campaign': campaign,
            'total_qr_generated': total_qr_generated,
            'total_qr_used': total_qr_used,
            'total_qr_active': total_qr_active,
            'usage_rate': round((total_qr_used / total_qr_generated * 100) if total_qr_generated > 0 else 0, 1),
            'products_count': campaign_products_count,
            'top_branch': top_branch.name if top_branch else 'Henüz kullanılmadı',
            'top_branch_count': top_branch.usage_count if top_branch else 0
        })
    
    return render_template('reports.html',
                         branches=branches,
                         total_points=total_points,
                         total_redemptions=total_redemptions,
                         active_customers=active_customers,
                         total_transactions=total_transactions,
                         campaign_stats=campaign_stats)

# Rapor Verisi API
@app.route('/admin/report_data')
@login_required
def report_data():
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    report_type = request.args.get('type', 'points')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    branch_id = request.args.get('branch_id')
    
    try:
        from datetime import datetime, timedelta
        
        # Tarih filtresi
        query_filter = []
        if start_date:
            query_filter.append(Transaction.timestamp >= datetime.strptime(start_date, '%Y-%m-%d'))
        if end_date:
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            query_filter.append(Transaction.timestamp < end_datetime)
        
        data = {}
        
        if report_type == 'points':
            # Puan kazanım raporu
            try:
                # Admin olmayan kullanıcıları getir
                query = db.session.query(
                    User.name,
                    User.email,
                    User.points,
                    db.func.max(CustomerQR.used_at).label('last_qr_scan'),
                    Branch.name.label('last_branch')
                ).filter(User.is_admin == False)\
                .outerjoin(CustomerQR, CustomerQR.customer_id == User.id)\
                .outerjoin(Branch, CustomerQR.used_by_branch_id == Branch.id)\
                .group_by(User.id, Branch.name)\
                .order_by(User.points.desc())
                
                # Tarih filtresi varsa uygula
                if start_date and end_date:
                    start_datetime = datetime.fromisoformat(start_date)
                    end_datetime = datetime.fromisoformat(end_date) + timedelta(days=1)
                    query = query.filter(CustomerQR.used_at.between(start_datetime, end_datetime))
                
                results = query.limit(50).all()
                data['points_data'] = [
                    [r.name, r.email, r.points, 
                     r.last_qr_scan.strftime('%d.%m.%Y') if r.last_qr_scan else 'Hiç QR okutmamış',
                     r.last_branch or 'Bilinmiyor']
                    for r in results
                ] if results else []
            except Exception as e:
                print(f"Points query error: {e}")
                data['points_data'] = []
            
        elif report_type == 'redemptions':
            # Ürün alım raporu
            try:
                query = db.session.query(
                    User.name.label('user_name'),
                    Product.name.label('product_name'),
                    ProductRedemption.points_used,
                    ProductRedemption.redeemed_at,
                    ProductRedemption.is_confirmed,
                    Branch.name.label('branch_name')
                ).select_from(ProductRedemption)\
                .join(User, ProductRedemption.user_id == User.id)\
                .join(Product, ProductRedemption.product_id == Product.id)\
                .outerjoin(Branch, ProductRedemption.confirmed_by_branch_id == Branch.id)
                
                results = query.order_by(ProductRedemption.redeemed_at.desc()).limit(50).all()
                data['redemptions_data'] = [
                    [r.user_name, r.product_name, r.points_used, 
                     r.redeemed_at.strftime('%d.%m.%Y %H:%M'),
                     'Onaylandı' if r.is_confirmed else 'Bekliyor',
                     r.branch_name if r.branch_name else 'Henüz Onaylanmadı']
                    for r in results
                ] if results else []
            except Exception as e:
                print(f"Redemptions query error: {e}")
                data['redemptions_data'] = []
            
        elif report_type == 'branches':
            # Şube performans raporu
            try:
                branches = Branch.query.all()
                data['branches_data'] = []
                for branch in branches:
                    # QR kod okutma sayısı
                    qr_scan_count = CustomerQR.query.filter_by(used_by_branch_id=branch.id, is_used=True).count()
                    
                    # Verilen toplam puan
                    total_points = db.session.query(db.func.sum(CustomerQR.points_earned)).filter_by(
                        used_by_branch_id=branch.id, is_used=True
                    ).scalar() or 0
                    
                    # Aktif müşteri sayısı (bu şubede QR okutan müşteriler)
                    active_customers = db.session.query(CustomerQR.customer_id).filter_by(
                        used_by_branch_id=branch.id, is_used=True
                    ).distinct().count()
                    
                    # Onaylanan ürün sayısı
                    confirmed_products = ProductRedemption.query.filter_by(
                        confirmed_by_branch_id=branch.id, is_confirmed=True
                    ).count()
                    
                    # Performans skoru
                    performance_score = qr_scan_count + confirmed_products
                    performance = 'Yüksek' if performance_score > 20 else 'Orta' if performance_score > 10 else 'Düşük'
                    
                    data['branches_data'].append([
                        branch.name,
                        qr_scan_count,
                        total_points,
                        active_customers,
                        f"{confirmed_products} Ürün Onayı",
                        performance
                    ])
            except Exception as e:
                print(f"Branches query error: {e}")
                data['branches_data'] = []
        
        elif report_type == 'customers':
            # Müşteri analiz raporu
            try:
                users = User.query.filter(~User.is_admin).all()
                data['customers_data'] = []
                for user in users:
                    used_points = db.session.query(db.func.sum(ProductRedemption.points_used)).filter_by(user_id=user.id).scalar() or 0
                    last_activity = db.session.query(db.func.max(Transaction.timestamp)).filter_by(user_id=user.id).scalar()
                    total_transactions = Transaction.query.filter_by(user_id=user.id).count()
                    total_spent = db.session.query(db.func.sum(Transaction.amount)).filter_by(user_id=user.id).scalar() or 0
                    
                    # Müşteri seviyesi belirleme
                    if user.points >= 100:
                        level = "🥇 VIP"
                    elif user.points >= 50:
                        level = "🥈 Premium"
                    elif user.points >= 20:
                        level = "🥉 Standart"
                    else:
                        level = "🆕 Yeni"
                    
                    # Son aktivite tarihi
                    last_activity_str = last_activity.strftime('%d.%m.%Y') if last_activity else 'Hiç yok'
                    
                    data['customers_data'].append([
                        user.name,
                        user.email,
                        user.phone or 'Yok',
                        user.created_at.strftime('%d.%m.%Y') if hasattr(user, 'created_at') else 'Bilinmiyor',
                        user.points,
                        used_points,
                        f"{total_spent:.2f} ₺",
                        total_transactions,
                        last_activity_str,
                        level,
                        'Aktif' if last_activity and (datetime.now() - last_activity).days < 30 else 'Pasif'
                    ])
            except Exception as e:
                print(f"Customers query error: {e}")
                data['customers_data'] = []
        
        elif report_type == 'transactions':
            # İşlem geçmiş raporu
            try:
                # Transaction tablosundan temel veriler
                transactions = db.session.query(
                    Transaction.timestamp,
                    User.name,
                    Transaction.transaction_type,
                    Transaction.amount,
                    Transaction.points_earned,
                    Transaction.points_used,
                    Transaction.description
                ).join(User).order_by(Transaction.timestamp.desc()).limit(100).all()
                
                data['transactions_data'] = []
                for t in transactions:
                    # İşlem tipine göre açıklama
                    if t.transaction_type == 'purchase':
                        type_desc = 'QR Puan Kazanımı'
                        amount_desc = f"+{t.points_earned} Puan"
                    elif t.transaction_type == 'redeem':
                        type_desc = 'Puan Kullanımı'
                        amount_desc = f"-{t.points_used} Puan"
                    else:
                        type_desc = t.transaction_type.title()
                        amount_desc = f"{t.amount:.2f} ₺"
                    
                    data['transactions_data'].append([
                        t.timestamp.strftime('%d.%m.%Y %H:%M'),
                        t.name,
                        type_desc,
                        amount_desc,
                        t.description or 'İşlem detayı'
                    ])
                    
            except Exception as e:
                print(f"Transactions query error: {e}")
                data['transactions_data'] = []
        
        elif report_type == 'campaigns':
            # Kampanya raporu
            try:
                campaigns = Campaign.query.order_by(Campaign.created_at.desc()).all()
                data['campaigns_data'] = []
                
                print(f"Found {len(campaigns)} campaigns")  # Debug log
                
                # Eğer hiç kampanya yoksa örnek veri ekle
                if not campaigns:
                    data['campaigns_data'] = [
                        ['Henüz kampanya oluşturulmamış', 0, 0, '%0', 0, 'Yok', 'Pasif']
                    ]
                else:
                    for campaign in campaigns:
                        # Kampanya QR kullanım istatistikleri
                        total_qr_generated = CampaignUsage.query.filter_by(campaign_id=campaign.id).count()
                        total_qr_used = CampaignUsage.query.filter_by(campaign_id=campaign.id, is_used=True).count()
                        
                        # Kampanya ürün sayısı
                        campaign_products_count = CampaignProduct.query.filter_by(
                            campaign_id=campaign.id, 
                            is_active=True
                        ).count()
                        
                        # En çok kullanılan şube
                        top_branch = db.session.query(
                            Branch.name,
                            db.func.count(CampaignUsage.id).label('usage_count')
                        ).join(
                            CampaignUsage, Branch.id == CampaignUsage.used_by_branch_id
                        ).filter(
                            CampaignUsage.campaign_id == campaign.id,
                            CampaignUsage.is_used == True
                        ).group_by(Branch.id, Branch.name).order_by(
                            db.func.count(CampaignUsage.id).desc()
                        ).first()
                        
                        # Kampanya ürünlerini al
                        campaign_products = CampaignProduct.query.filter_by(
                            campaign_id=campaign.id, 
                            is_active=True
                        ).all()
                        
                        product_names = []
                        for cp in campaign_products:
                            if cp.product_id and cp.product:
                                product_names.append(cp.product.name)
                            elif cp.product_name:
                                product_names.append(cp.product_name)
                        
                        products_str = ', '.join(product_names) if product_names else 'Ürün yok'
                        
                        # Kullanım oranı
                        usage_rate = round((total_qr_used / total_qr_generated * 100) if total_qr_generated > 0 else 0, 1)
                        
                        # Kampanya durumu
                        status = 'Aktif' if campaign.is_active else 'Pasif'
                        
                        data['campaigns_data'].append([
                            campaign.title,  # Campaign model'de 'name' değil 'title' var
                            total_qr_generated,
                            total_qr_used,
                            f"%{usage_rate}",
                            campaign_products_count,
                            products_str,
                            top_branch.name if top_branch else 'Henüz kullanılmadı',
                            status
                        ])
                    
            except Exception as e:
                print(f"Campaigns query error: {e}")
                data['campaigns_data'] = []
        
        # Grafik verileri
        data['chart_labels'] = []
        data['chart_points'] = []
        data['branch_labels'] = []
        data['branch_values'] = []
        
        # Son 7 günün verilerini al
        try:
            for i in range(7):
                date = datetime.now() - timedelta(days=i)
                day_points = db.session.query(db.func.sum(Transaction.points_earned)).filter(
                    db.func.date(Transaction.timestamp) == date.date()
                ).scalar() or 0
                
                data['chart_labels'].insert(0, date.strftime('%d.%m'))
                data['chart_points'].insert(0, day_points)
        except Exception as e:
            print(f"Chart data error: {e}")
            # Varsayılan veriler
            for i in range(7):
                date = datetime.now() - timedelta(days=i)
                data['chart_labels'].insert(0, date.strftime('%d.%m'))
                data['chart_points'].insert(0, 0)
        
        # Şube dağılımı
        try:
            branches = Branch.query.all()
            for branch in branches:
                points = db.session.query(db.func.sum(Transaction.points_earned)).filter_by(branch_id=branch.id).scalar() or 0
                data['branch_labels'].append(branch.name)
                data['branch_values'].append(points)
            
            # Eğer hiç veri yoksa örnek veri ekle
            if not data['branch_labels']:
                data['branch_labels'] = ['Merkez Şube', 'Kadıköy Şube', 'Beşiktaş Şube']
                data['branch_values'] = [0, 0, 0]
        except Exception as e:
            print(f"Branch data error: {e}")
            # Varsayılan şube verileri
            data['branch_labels'] = ['Merkez Şube', 'Kadıköy Şube', 'Beşiktaş Şube']
            data['branch_values'] = [0, 0, 0]
        
        return jsonify(data)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Excel Export
@app.route('/admin/export_report')
@login_required
def export_report():
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # 1. Puan Kazanımları Sayfası
        ws1 = wb.create_sheet("Puan Kazanımları")
        headers1 = ['Müşteri', 'E-posta', 'Toplam Puan', 'Son QR Okutma', 'Son Şube']
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Puan kazanımları verilerini al (limitsiz)
        try:
            points_data = db.session.query(
                User.name,
                User.email,
                User.points,
                db.func.max(CustomerQR.used_at).label('last_qr_scan'),
                Branch.name.label('last_branch')
            ).filter(User.is_admin == False)\
            .outerjoin(CustomerQR, CustomerQR.customer_id == User.id)\
            .outerjoin(Branch, CustomerQR.used_by_branch_id == Branch.id)\
            .group_by(User.id, Branch.name)\
            .all()
            
            for row, u in enumerate(points_data, 2):
                ws1.cell(row=row, column=1, value=u.name)
                ws1.cell(row=row, column=2, value=u.email)
                ws1.cell(row=row, column=3, value=u.points)
                ws1.cell(row=row, column=4, value=u.last_qr_scan.strftime('%d.%m.%Y') if u.last_qr_scan else 'Hiç QR okutmamış')
                ws1.cell(row=row, column=5, value=u.last_branch or 'Bilinmiyor')
        except Exception as e:
            print(f"Points export error: {e}")
        
        # 2. Ürün Alımları Sayfası
        ws2 = wb.create_sheet("Ürün Alımları")
        headers2 = ['Müşteri', 'Ürün', 'Kullanılan Puan', 'Tarih', 'Durum', 'Onaylayan Şube']
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Ürün alımları verilerini al (limitsiz)
        try:
            redemptions_data = db.session.query(
                User.name.label('user_name'),
                Product.name.label('product_name'),
                ProductRedemption.points_used,
                ProductRedemption.redeemed_at,
                ProductRedemption.is_confirmed,
                Branch.name.label('branch_name')
            ).select_from(ProductRedemption)\
            .join(User, ProductRedemption.user_id == User.id)\
            .join(Product, ProductRedemption.product_id == Product.id)\
            .outerjoin(Branch, ProductRedemption.confirmed_by_branch_id == Branch.id)\
            .order_by(ProductRedemption.redeemed_at.desc())\
            .all()
            
            for row, r in enumerate(redemptions_data, 2):
                ws2.cell(row=row, column=1, value=r.user_name)
                ws2.cell(row=row, column=2, value=r.product_name)
                ws2.cell(row=row, column=3, value=r.points_used)
                ws2.cell(row=row, column=4, value=r.redeemed_at.strftime('%d.%m.%Y %H:%M'))
                ws2.cell(row=row, column=5, value='Onaylandı' if r.is_confirmed else 'Bekliyor')
                ws2.cell(row=row, column=6, value=r.branch_name if r.branch_name else 'Henüz Onaylanmadı')
        except Exception as e:
            print(f"Redemptions export error: {e}")
        
        # 3. Şube Performansı Sayfası
        ws3 = wb.create_sheet("Şube Performansı")
        headers3 = ['Şube Adı', 'Adres', 'QR Okutma', 'Verilen Puan', 'Aktif Müşteri', 'Onaylanan Ürün', 'Performans']
        for col, header in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Şube performansı verilerini al (limitsiz)
        try:
            branches_data = db.session.query(
                Branch.name,
                Branch.address,
                db.func.count(CustomerQR.id).label('qr_scans'),
                db.func.sum(CustomerQR.points_earned).label('total_points_given'),
                db.func.count(db.distinct(CustomerQR.customer_id)).label('active_customers'),
                db.func.count(ProductRedemption.id).label('confirmed_products')
            ).outerjoin(CustomerQR, CustomerQR.used_by_branch_id == Branch.id)\
            .outerjoin(ProductRedemption, ProductRedemption.confirmed_by_branch_id == Branch.id)\
            .group_by(Branch.id)\
            .all()
            
            for row, b in enumerate(branches_data, 2):
                ws3.cell(row=row, column=1, value=b.name)
                ws3.cell(row=row, column=2, value=b.address)
                ws3.cell(row=row, column=3, value=b.qr_scans or 0)
                ws3.cell(row=row, column=4, value=b.total_points_given or 0)
                ws3.cell(row=row, column=5, value=b.active_customers or 0)
                ws3.cell(row=row, column=6, value=b.confirmed_products or 0)
                
                # Performans hesapla
                total_activity = (b.qr_scans or 0) + (b.confirmed_products or 0)
                if total_activity > 50:
                    performance = "Yüksek"
                elif total_activity > 20:
                    performance = "Orta"
                else:
                    performance = "Düşük"
                ws3.cell(row=row, column=7, value=performance)
        except Exception as e:
            print(f"Branches export error: {e}")
        
        # 4. Müşteri Analizi Sayfası
        ws4 = wb.create_sheet("Müşteri Analizi")
        headers4 = ['Müşteri', 'E-posta', 'Toplam Puan', 'QR Sayısı', 'Ürün Alımı', 'Son Aktivite']
        for col, header in enumerate(headers4, 1):
            cell = ws4.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="6f42c1", end_color="6f42c1", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Müşteri analizi verilerini al (limitsiz)
        try:
            customers_data = db.session.query(
                User.name,
                User.email,
                User.points,
                db.func.count(CustomerQR.id).label('qr_count'),
                db.func.count(ProductRedemption.id).label('redemption_count'),
                db.func.max(CustomerQR.used_at).label('last_activity')
            ).filter(User.is_admin == False)\
            .outerjoin(CustomerQR, CustomerQR.customer_id == User.id)\
            .outerjoin(ProductRedemption, ProductRedemption.user_id == User.id)\
            .group_by(User.id)\
            .all()
            
            for row, c in enumerate(customers_data, 2):
                ws4.cell(row=row, column=1, value=c.name)
                ws4.cell(row=row, column=2, value=c.email)
                ws4.cell(row=row, column=3, value=c.points)
                ws4.cell(row=row, column=4, value=c.qr_count or 0)
                ws4.cell(row=row, column=5, value=c.redemption_count or 0)
                ws4.cell(row=row, column=6, value=c.last_activity.strftime('%d.%m.%Y') if c.last_activity else 'Hiç aktivite yok')
        except Exception as e:
            print(f"Customers export error: {e}")
        
        # 5. İşlem Geçmişi Sayfası
        ws5 = wb.create_sheet("İşlem Geçmişi")
        headers5 = ['Müşteri', 'İşlem Türü', 'Tutar', 'Tarih', 'Açıklama', 'Puan Değişimi', 'Şube']
        for col, header in enumerate(headers5, 1):
            cell = ws5.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="17a2b8", end_color="17a2b8", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # İşlem geçmişi verilerini al (limitsiz)
        try:
            transactions = db.session.query(
                Transaction.user_id,
                Transaction.amount,
                Transaction.transaction_type,
                Transaction.description,
                Transaction.updated_at,
                Transaction.points_earned,
                User.name.label('user_name')
            ).join(User, Transaction.user_id == User.id)\
            .order_by(Transaction.updated_at.desc())\
            .all()
            
            for row, t in enumerate(transactions, 2):
                ws5.cell(row=row, column=1, value=t.user_name)
                ws5.cell(row=row, column=2, value=t.transaction_type)
                ws5.cell(row=row, column=3, value=f"{t.amount:.2f} ₺")
                ws5.cell(row=row, column=4, value=t.updated_at.strftime('%d.%m.%Y %H:%M'))
                ws5.cell(row=row, column=5, value=t.description or '')
                ws5.cell(row=row, column=6, value=t.points_earned or 0)
                ws5.cell(row=row, column=7, value='Bilinmiyor')
        except Exception as e:
            print(f"Transactions export error: {e}")
        
        # 6. Kampanya Raporu Sayfası
        ws6 = wb.create_sheet("Kampanya Raporu")
        headers6 = ['Kampanya Adı', 'QR Oluşturulan', 'QR Kullanılan', 'Kullanım Oranı', 'Ürün Sayısı', 'Ürünler', 'En Aktif Şube', 'Durum']
        for col, header in enumerate(headers6, 1):
            cell = ws6.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Kampanya raporu verilerini al (limitsiz)
        try:
            campaigns = Campaign.query.order_by(Campaign.created_at.desc()).all()
            
            if not campaigns:
                ws6.cell(row=2, column=1, value='Henüz kampanya oluşturulmamış')
                ws6.cell(row=2, column=2, value=0)
                ws6.cell(row=2, column=3, value=0)
                ws6.cell(row=2, column=4, value='%0')
                ws6.cell(row=2, column=5, value=0)
                ws6.cell(row=2, column=6, value='Ürün yok')
                ws6.cell(row=2, column=7, value='Yok')
                ws6.cell(row=2, column=8, value='Pasif')
            else:
                for row, campaign in enumerate(campaigns, 2):
                    # Kampanya QR kullanım istatistikleri
                    total_qr_generated = CampaignUsage.query.filter_by(campaign_id=campaign.id).count()
                    total_qr_used = CampaignUsage.query.filter_by(campaign_id=campaign.id, is_used=True).count()
                    
                    # Kampanya ürün sayısı
                    campaign_products_count = CampaignProduct.query.filter_by(
                        campaign_id=campaign.id, 
                        is_active=True
                    ).count()
                    
                    # Kampanya ürünlerini al
                    campaign_products = CampaignProduct.query.filter_by(
                        campaign_id=campaign.id, 
                        is_active=True
                    ).all()
                    
                    product_names = []
                    for cp in campaign_products:
                        if cp.product_id and cp.product:
                            product_names.append(cp.product.name)
                        elif cp.product_name:
                            product_names.append(cp.product_name)
                    
                    products_str = ', '.join(product_names) if product_names else 'Ürün yok'
                    
                    # En çok kullanılan şube
                    top_branch = db.session.query(
                        Branch.name,
                        db.func.count(CampaignUsage.id).label('usage_count')
                    ).join(
                        CampaignUsage, Branch.id == CampaignUsage.used_by_branch_id
                    ).filter(
                        CampaignUsage.campaign_id == campaign.id,
                        CampaignUsage.is_used == True
                    ).group_by(Branch.id, Branch.name).order_by(
                        db.func.count(CampaignUsage.id).desc()
                    ).first()
                    
                    # Kullanım oranı
                    usage_rate = round((total_qr_used / total_qr_generated * 100) if total_qr_generated > 0 else 0, 1)
                    
                    # Kampanya durumu
                    status = 'Aktif' if campaign.is_active else 'Pasif'
                    
                    ws6.cell(row=row, column=1, value=campaign.title)
                    ws6.cell(row=row, column=2, value=total_qr_generated)
                    ws6.cell(row=row, column=3, value=total_qr_used)
                    ws6.cell(row=row, column=4, value=f"%{usage_rate}")
                    ws6.cell(row=row, column=5, value=campaign_products_count)
                    ws6.cell(row=row, column=6, value=products_str)
                    ws6.cell(row=row, column=7, value=top_branch.name if top_branch else 'Henüz kullanılmadı')
                    ws6.cell(row=row, column=8, value=status)
                    
        except Exception as e:
            print(f"Campaigns export error: {e}")
        
        # Kolon genişliklerini ayarla
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Excel dosyasını memory'ye kaydet
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Response oluştur
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Dosya adını tarih ile oluştur
        today = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'cafe_sadakat_raporu_{today}.xlsx'
        
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response
        
    except Exception as e:
        print(f"Excel export hatası: {e}")
        return jsonify({'error': f'Export hatası: {str(e)}'}), 500

# Admin artık QR oluşturmaz, sadece sistem yönetimi yapar

# Şube Oluşturma (Admin)
@app.route('/admin/create_branch', methods=['POST'])
@login_required
def create_branch():
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    try:
        # JSON veya form data kontrolü
        if request.is_json:
            data = request.get_json()
            name = data.get('name')
            address = data.get('address')
            phone = data.get('phone')
            email = data.get('email')
            password = data.get('password')
            working_hours = data.get('working_hours', '')
            image_file = None
        else:
            name = request.form['name']
            address = request.form['address']
            phone = request.form['phone']
            email = request.form['email']
            password = request.form['password']
            working_hours = request.form.get('working_hours', '')
            image_file = request.files.get('image')
        
        # E-posta kontrolü
        existing_branch = Branch.query.filter_by(email=email).first()
        if existing_branch:
            return jsonify({'success': False, 'error': f'Bu e-posta adresi zaten {existing_branch.name} şubesi tarafından kullanılıyor'})
        
        # Görsel yükleme
        image_filename = None
        if image_file and image_file.filename:
            print(f"DEBUG: Görsel dosyası alındı: {image_file.filename}")
            filename = secure_filename(image_file.filename)
            # Benzersiz dosya adı oluştur
            import uuid
            unique_filename = f"{uuid.uuid4().hex}_{filename}"
            image_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            print(f"DEBUG: Görsel kaydediliyor: {image_path}")
            image_file.save(image_path)
            image_filename = unique_filename
            print(f"DEBUG: Görsel başarıyla kaydedildi: {unique_filename}")
        else:
            print("DEBUG: Görsel dosyası bulunamadı veya boş")
        
        branch = Branch(
            name=name,
            address=address,
            phone=phone,
            email=email,
            password_hash=generate_password_hash(password),
            working_hours=working_hours,
            image=image_filename
        )
        
        db.session.add(branch)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{name} şubesi başarıyla oluşturuldu',
            'branch_id': branch.id
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Hata: {str(e)}'})


# Ürün Oluşturma (Admin)
@app.route('/admin/create_product', methods=['POST'])
@login_required
def create_product():
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    try:
        name = request.form['name']
        description = request.form.get('description', '')
        points_required = int(request.form['points_required'])
        category_id = request.form.get('category_id')
        
        # Kategori ID'sinden kategori adını al
        category_name = 'Genel'
        if category_id and category_id.isdigit():
            category = Category.query.get(int(category_id))
            if category:
                category_name = category.name
        
        image_filename = None
        if 'product_image' in request.files:
            file = request.files['product_image']
            if file and file.filename != '' and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
                filename = timestamp + filename
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)
                image_filename = filename
        
        product = Product(
            name=name,
            description=description,
            points_required=points_required,
            category_id=int(category_id) if category_id and category_id.isdigit() else None,
            category=category_name,
            image_filename=image_filename
        )
        
        db.session.add(product)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{name} ürünü başarıyla oluşturuldu!'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# Ürün Kalıcı Silme (Admin) - İlişkileri güvenli biçimde temizler
@app.route('/admin/hard_delete_product/<int:product_id>', methods=['DELETE'])
@login_required
def hard_delete_product(product_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    try:
        product = Product.query.get_or_404(product_id)

        # 1) Bu ürüne bağlı ProductRedemption kayıtlarını sil
        redemptions = ProductRedemption.query.filter_by(product_id=product.id).all()
        for r in redemptions:
            db.session.delete(r)

        # 2) Kampanya ürünlerinde bu ürüne olan FK'yi kaldır (NULL yap)
        linked_campaign_products = CampaignProduct.query.filter_by(product_id=product.id).all()
        for cp in linked_campaign_products:
            cp.product_id = None

        # 3) Ürün görsel dosyasını sil
        if product.image_filename:
            image_path = os.path.join(app.config['UPLOAD_FOLDER'], product.image_filename)
            if os.path.exists(image_path):
                try:
                    os.remove(image_path)
                except Exception:
                    pass

        # 4) Ürünü sil
        db.session.delete(product)
        db.session.commit()

        return jsonify({'success': True, 'message': 'Ürün kalıcı olarak silindi.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

# Ürün Bilgisi Getirme (Admin)
@app.route('/admin/get_product/<int:product_id>', methods=['GET'])
@login_required
def get_product(product_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    product = Product.query.get_or_404(product_id)
    return jsonify({
        'success': True,
        'product': {
            'id': product.id,
            'name': product.name,
            'description': product.description or '',
            'points_required': product.points_required,
            'category_id': product.category_id,
            'category': product.category,
            'image_filename': product.image_filename,
            'is_active': product.is_active
        }
    })

# Ürün Güncelleme (Admin)
@app.route('/admin/update_product/<int:product_id>', methods=['POST'])
@login_required
def update_product(product_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    product = Product.query.get_or_404(product_id)
    try:
        # JSON veya form verisi gelebilir
        if request.is_json:
            data = request.get_json()
            name = data.get('name', product.name)
            description = data.get('description', product.description or '')
            points_required = int(data.get('points_required', product.points_required))
            category_id = data.get('category_id')
            image_file = None
        else:
            name = request.form.get('name', product.name)
            description = request.form.get('description', product.description or '')
            points_required = int(request.form.get('points_required', product.points_required))
            category_id = request.form.get('category_id')
            image_file = request.files.get('image') if 'image' in request.files else None

        # Kategori adı
        category_name = product.category or 'Genel'
        category_fk = product.category_id
        if category_id is not None:
            if str(category_id).isdigit():
                cat_obj = Category.query.get(int(category_id))
                if cat_obj:
                    category_name = cat_obj.name
                    category_fk = cat_obj.id
            elif category_id == '' or category_id is None:
                category_name = 'Genel'
                category_fk = None

        # Görsel güncelle
        if image_file and allowed_file(image_file.filename):
            # Eski görseli sil
            if product.image_filename:
                old_path = os.path.join(app.config['UPLOAD_FOLDER'], product.image_filename)
                if os.path.exists(old_path):
                    try:
                        os.remove(old_path)
                    except Exception:
                        pass
            filename = secure_filename(image_file.filename)
            unique_suffix = datetime.utcnow().strftime('%Y%m%d%H%M%S%f')
            filename = f"{unique_suffix}_{filename}"
            image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            image_file.save(image_path)
            product.image_filename = filename

        product.name = name
        product.description = description
        product.points_required = points_required
        product.category_id = category_fk
        product.category = category_name

        db.session.commit()
        return jsonify({'success': True, 'message': 'Ürün başarıyla güncellendi!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

# Ürün Silme (Admin) - yumuşak silme: is_active False
@app.route('/admin/delete_product/<int:product_id>', methods=['DELETE'])
@login_required
def delete_product(product_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    try:
        product = Product.query.get_or_404(product_id)
        # Ürünü pasif yap
        product.is_active = False
        # Görseli varsa silelim (opsiyonel)
        if product.image_filename:
            image_path = os.path.join(app.config['UPLOAD_FOLDER'], product.image_filename)
            if os.path.exists(image_path):
                try:
                    os.remove(image_path)
                except Exception:
                    pass
            product.image_filename = None
        db.session.commit()
        return jsonify({'success': True, 'message': 'Ürün başarıyla silindi (pasif hale getirildi)!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

# Kategori Yönetimi (Admin)
@app.route('/admin/categories', methods=['GET'])
@login_required
def get_categories():
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    # Category tablosundan aktif kategorileri al
    categories = Category.query.filter_by(is_active=True).order_by(Category.name).all()
    category_list = [{'id': cat.id, 'name': cat.name, 'description': cat.description} for cat in categories]
    
    return jsonify({'categories': category_list})

@app.route('/admin/add_category', methods=['POST'])
@login_required
def add_category():
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    try:
        category_name = request.form['category_name'].strip()
        category_description = request.form.get('category_description', '').strip()
        
        if not category_name:
            return jsonify({'error': 'Kategori adı boş olamaz'}), 400
        
        # Kategori zaten var mı kontrol et
        existing = Category.query.filter_by(name=category_name).first()
        if existing:
            return jsonify({'error': 'Bu kategori zaten mevcut'}), 400
        
        # Yeni kategori oluştur
        new_category = Category(
            name=category_name,
            description=category_description if category_description else f'{category_name} kategorisi'
        )
        
        db.session.add(new_category)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{category_name} kategorisi başarıyla eklendi!',
            'category': {
                'id': new_category.id,
                'name': new_category.name,
                'description': new_category.description
            }
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/admin/delete_category', methods=['POST'])
@login_required
def delete_category():
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    try:
        category_id = int(request.form['category_id'])
        
        # Kategoriyi bul
        category = Category.query.get_or_404(category_id)
        
        # Bu kategorideki ürün sayısını kontrol et
        product_count = Product.query.filter_by(category_id=category_id, is_active=True).count()
        
        if product_count > 0:
            return jsonify({'error': f'Bu kategoride {product_count} aktif ürün var. Önce ürünleri başka kategoriye taşıyın.'}), 400
        
        # Genel kategorisini bul
        general_category = Category.query.filter_by(name='Genel').first()
        if not general_category:
            # Genel kategori yoksa oluştur
            general_category = Category(name='Genel', description='Genel kategorideki ürünler')
            db.session.add(general_category)
            db.session.commit()
        
        # Kategorideki tüm ürünleri "Genel" kategorisine taşı
        Product.query.filter_by(category_id=category_id).update({
            'category_id': general_category.id,
            'category': 'Genel'
        })
        
        # Kategoriyi sil
        db.session.delete(category)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{category.name} kategorisi başarıyla silindi!'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/admin/edit_category', methods=['POST'])
@login_required
def edit_category():
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    try:
        category_id = int(request.form['category_id'])
        new_name = request.form['new_name'].strip()
        new_description = request.form.get('new_description', '').strip()
        
        if not new_name:
            return jsonify({'error': 'Kategori adı boş olamaz'}), 400
        
        # Kategoriyi bul
        category = Category.query.get_or_404(category_id)
        old_name = category.name
        
        if old_name == new_name:
            return jsonify({'error': 'Yeni kategori adı eskisiyle aynı olamaz'}), 400
        
        # Yeni kategori adı zaten var mı kontrol et
        existing = Category.query.filter_by(name=new_name).first()
        if existing and existing.id != category_id:
            return jsonify({'error': 'Bu kategori adı zaten mevcut'}), 400
        
        # Kategoriyi güncelle
        category.name = new_name
        if new_description:
            category.description = new_description
        
        # Ürünlerdeki eski kategori string'ini de güncelle (uyumluluk için)
        updated_count = Product.query.filter_by(category_id=category_id).update({'category': new_name})
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{old_name} kategorisi {new_name} olarak güncellendi! ({updated_count} ürün güncellendi)',
            'category': {
                'id': category.id,
                'name': category.name,
                'description': category.description
            }
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# Ürün Durumu Değiştirme (Admin)
@app.route('/admin/toggle_product/<int:product_id>', methods=['POST'])
@login_required
def toggle_product(product_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Yetkisiz erişim'}), 403
    
    product = Product.query.get_or_404(product_id)
    product.is_active = not product.is_active
    db.session.commit()
    
    status = "aktif" if product.is_active else "pasif"
    return jsonify({
        'success': True,
        'message': f'{product.name} ürünü {status} duruma getirildi.'
    })

# Profil Sayfası
@app.route('/profile')
@login_required
def profile():
    # Kullanıcının toplam alım sayısını hesapla (ProductRedemption tablosundan)
    purchase_count = ProductRedemption.query.filter_by(user_id=current_user.id).count()
    # Tüm şubeleri getir
    branches = Branch.query.all()
    return render_template('profile.html', purchase_count=purchase_count, branches=branches)

# QR Kodu Kontrol Et
@app.route('/check_qr_usage', methods=['POST'])
@login_required
def check_qr_usage():
    try:
        data = request.get_json()
        last_check = int(data.get('last_check', 0))
        last_check_datetime = datetime.fromtimestamp(last_check / 1000) if last_check > 0 else datetime.min
        
        # Son kontrol zamanından sonra kullanılan QR kodları bul
        recent_transactions = Transaction.query.filter(
            Transaction.user_id == current_user.id,
            Transaction.timestamp > last_check_datetime,
            Transaction.transaction_type == 'purchase'
        ).all()
        
        new_points = sum(t.points_earned for t in recent_transactions)
        qr_used = len(recent_transactions) > 0
        
        return jsonify({
            'success': True,
            'new_points': new_points,
            'qr_used': qr_used
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Şifre Değiştirme
@app.route('/change_password', methods=['POST'])
@login_required
def change_password():
    current_password = request.form['current_password']
    new_password = request.form['new_password']
    confirm_password = request.form['confirm_password']
    
    # Mevcut şifre kontrolü
    if not check_password_hash(current_user.password_hash, current_password):
        flash('Mevcut şifre yanlış!', 'password_error')
        return redirect(url_for('profile'))
    
    # Yeni şifre kontrolü
    if new_password != confirm_password:
        flash('Yeni şifreler eşleşmiyor!', 'password_error')
        return redirect(url_for('profile'))
    
    if len(new_password) < 6:
        flash('Şifre en az 6 karakter olmalıdır!', 'password_error')
        return redirect(url_for('profile'))
    
    # Şifreyi güncelle
    current_user.password_hash = generate_password_hash(new_password)
    db.session.commit()
    
    flash('Şifreniz başarıyla değiştirildi!', 'password')
    return redirect(url_for('profile'))

# Profil Güncelleme
@app.route('/update_profile', methods=['POST'])
@login_required
def update_profile():
    name = request.form['name'].strip()
    phone = request.form['phone'].strip()
    preferred_branch = request.form.get('preferred_branch')
    
    if not name:
        flash('Ad soyad boş olamaz!', 'profile_error')
        return redirect(url_for('profile'))
    
    # Profil bilgilerini güncelle
    current_user.name = name
    if phone:
        current_user.phone = phone
    
    # Tercih edilen şubeyi güncelle
    if preferred_branch and preferred_branch.isdigit():
        current_user.preferred_branch_id = int(preferred_branch)
    else:
        current_user.preferred_branch_id = None
    
    db.session.commit()
    
    flash('Profil bilgileriniz başarıyla güncellendi!', 'profile')
    return redirect(url_for('profile'))

# Kampanya Oluşturma
@app.route('/admin/create_campaign', methods=['POST'])
@login_required
def create_campaign():
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': 'Yetkiniz yok!'}), 403
    
    try:
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        
        # Validasyon
        if not title or not description:
            return jsonify({'success': False, 'error': 'Başlık ve açıklama zorunludur!'})
        
        if not start_date or not end_date:
            return jsonify({'success': False, 'error': 'Başlangıç ve bitiş tarihleri zorunludur!'})
        
        # Tarih formatını düzelt
        from datetime import datetime
        start_datetime = datetime.fromisoformat(start_date.replace('T', ' '))
        end_datetime = datetime.fromisoformat(end_date.replace('T', ' '))
        
        if start_datetime >= end_datetime:
            return jsonify({'success': False, 'error': 'Bitiş tarihi başlangıç tarihinden sonra olmalıdır!'})
        
        # Görsel yükleme
        image_filename = None
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename:
                import os
                from werkzeug.utils import secure_filename
                
                # Dosya uzantısı kontrolü
                allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
                if '.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions:
                    filename = secure_filename(file.filename)
                    # Benzersiz dosya adı oluştur
                    import uuid
                    unique_filename = f"{uuid.uuid4().hex}_{filename}"
                    
                    upload_folder = os.path.join('static', 'uploads')
                    os.makedirs(upload_folder, exist_ok=True)
                    
                    file_path = os.path.join(upload_folder, unique_filename)
                    file.save(file_path)
                    image_filename = unique_filename
        
        # Seçili şubeleri al
        selected_branches = []
        branches_data = request.form.get('branches')
        if branches_data:
            try:
                import json
                branch_ids = json.loads(branches_data)
                selected_branches = Branch.query.filter(Branch.id.in_(branch_ids)).all()
            except:
                pass
        
        # QR ayarlarını al
        qr_enabled = request.form.get('qr_enabled') == 'true'
        max_usage_per_customer = int(request.form.get('max_usage_per_customer', 1))
        total_usage_limit = int(request.form.get('total_usage_limit', 100))
        
        # Kampanya oluştur
        campaign = Campaign(
            title=title,
            description=description,
            image_filename=image_filename,
            start_date=start_datetime,
            end_date=end_datetime,
            is_active=True,
            qr_enabled=qr_enabled,
            max_usage_per_customer=max_usage_per_customer,
            total_usage_limit=total_usage_limit,
            created_at=get_turkey_time()
        )
        
        # Şubeleri kampanyaya ekle
        campaign.branches = selected_branches
        
        db.session.add(campaign)
        db.session.flush()  # Campaign ID'sini almak için
        
        # Seçili ürünleri ve indirim bilgilerini al
        products_data = request.form.get('products')
        if products_data:
            try:
                import json
                products_info = json.loads(products_data)
                
                for product_info in products_info:
                    product_id = product_info.get('product_id')
                    discount = product_info.get('discount')
                    discount_type = product_info.get('discount_type', 'percentage')
                    
                    # Ürünün var olduğunu kontrol et
                    product = db.session.get(Product, product_id)
                    if product and discount and discount > 0:
                        campaign_product = CampaignProduct(
                            campaign_id=campaign.id,
                            product_id=product_id,
                            product_name=product.name,  # Ürün adını ekle
                            product_description=product.description,  # Ürün açıklamasını ekle
                            discount=discount,
                            discount_type=discount_type,
                            created_at=get_turkey_time()
                        )
                        db.session.add(campaign_product)
            except Exception as e:
                print(f"Ürün ekleme hatası: {e}")
        
        db.session.commit()
        
        # Kampanya bildirimi gönder (push)
        send_campaign_notification(campaign)
        # Kampanya e-postasını tüm kullanıcılara gönder
        send_campaign_email(campaign)
        
        branch_count = len(selected_branches)
        product_count = len(json.loads(products_data)) if products_data else 0
        
        return jsonify({
            'success': True, 
            'message': f'Kampanya başarıyla oluşturuldu! ({branch_count} şube, {product_count} ürün seçildi)'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Hata: {str(e)}'})

# Şubelerimiz Sayfası
@app.route('/branches')
def branches():
    branches = Branch.query.filter_by(is_active=True).order_by(Branch.name).all()
    return render_template('branches.html', branches=branches)

# Admin - Şube Aktif/Pasif Yapma
@app.route('/admin/toggle_branch/<int:branch_id>', methods=['POST'])
@login_required
def toggle_branch(branch_id):
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': 'Yetki yok'})
    
    branch = Branch.query.get_or_404(branch_id)
    branch.is_active = not branch.is_active
    db.session.commit()
    
    status = 'aktif' if branch.is_active else 'pasif'
    return jsonify({
        'success': True, 
        'message': f'{branch.name} şubesi {status} yapıldı',
        'is_active': branch.is_active
    })

# Admin - Şube Düzenleme
@app.route('/admin/edit_branch/<int:branch_id>', methods=['GET', 'POST'])
@login_required
def edit_branch(branch_id):
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': 'Yetki yok'})
    
    branch = Branch.query.get_or_404(branch_id)
    
    if request.method == 'POST':
        try:
            # JSON veya form data kontrolü
            if request.is_json:
                data = request.get_json()
                branch.name = data.get('name', branch.name)
                branch.address = data.get('address', branch.address)
                branch.phone = data.get('phone', branch.phone)
                branch.working_hours = data.get('working_hours', branch.working_hours)
                
                if data.get('email') and data.get('email') != branch.email:
                    branch.email = data.get('email')
                
                if data.get('password'):
                    branch.password_hash = generate_password_hash(data.get('password'))
            else:
                # Form data ile güncelleme
                branch.name = request.form.get('name', branch.name)
                branch.address = request.form.get('address', branch.address)
                branch.phone = request.form.get('phone', branch.phone)
                branch.working_hours = request.form.get('working_hours', branch.working_hours)
                
                if request.form.get('email') and request.form.get('email') != branch.email:
                    branch.email = request.form.get('email')
                
                if request.form.get('password'):
                    branch.password_hash = generate_password_hash(request.form.get('password'))
                
                # Görsel güncelleme
                image_file = request.files.get('image')
                if image_file and image_file.filename:
                    # Eski görseli sil
                    if branch.image:
                        old_image_path = os.path.join(app.config['UPLOAD_FOLDER'], branch.image)
                        if os.path.exists(old_image_path):
                            os.remove(old_image_path)
                    
                    # Yeni görseli kaydet
                    filename = secure_filename(image_file.filename)
                    import uuid
                    unique_filename = f"{uuid.uuid4().hex}_{filename}"
                    image_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
                    image_file.save(image_path)
                    branch.image = unique_filename
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': f'{branch.name} şubesi güncellendi'
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': f'Hata: {str(e)}'})
    
    # GET request - şube bilgilerini döndür
    return jsonify({
        'success': True,
        'branch': {
            'id': branch.id,
            'name': branch.name,
            'address': branch.address,
            'phone': branch.phone,
            'email': branch.email,
            'working_hours': branch.working_hours,
            'is_active': branch.is_active
        }
    })

# Şifremi Unuttum
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()

        # Kullanıcıyı bul (varsa)
        user = User.query.filter_by(email=email).first()

        try:
            # Güvenli token üret ve istek kaydını oluştur (1 saat geçerli)
            token = secrets.token_hex(32)
            expires_at = (get_turkey_time().replace(tzinfo=None) + timedelta(hours=1))
            prr = PasswordResetRequest(
                user_id=user.id if user else None,
                email=email,
                token=token,
                expires_at=expires_at
            )
            db.session.add(prr)
            db.session.commit()

            # Adminlere bilgi mesajı gönder
            admins = User.query.filter_by(is_admin=True).all()
            if admins:
                requester = user.name if user else 'Bilinmeyen Kullanıcı'
                title = 'Şifre Sıfırlama Talebi'
                base = app.config.get('BASE_URL', 'https://reevpoints.tr').rstrip('/')
                reset_link = f"{base}/reset_password/{token}"

                # Kullanıcı mevcutsa e-posta ile sıfırlama bağlantısını gönder
                if user:
                    send_password_reset_email(user, reset_link)
                content = (
                    f"E-posta: {email}\n"
                    f"Talep Sahibi: {requester}\n"
                    f"Tarih: {get_turkey_time().strftime('%d.%m.%Y %H:%M')}\n"
                    f"Geçerlilik: 1 saat\n\n"
                    f"Şifreyi Sıfırla Bağlantısı:<br><a href=\"{reset_link}\" target=\"_blank\">{reset_link}</a>\n\n"
                    "Bu kullanıcı şifresini sıfırlamak istiyor. Lütfen linke tıklayarak yeni şifre belirleyin."
                )

                for admin in admins:
                    msg = Message(
                        title=title,
                        content=content,
                        recipient_id=admin.id,
                        is_admin_message=False
                    )
                    db.session.add(msg)

                db.session.commit()

            # Güvenlik: E-postanın var olup olmadığını belirtmeyelim
            flash('Talebiniz alınmıştır. En kısa sürede geçici şifreniz mail adresinize gönderilecektir.', 'info')
        except Exception as e:
            db.session.rollback()
            flash(f'Talep kaydedilirken hata oluştu: {str(e)}', 'error')

        return redirect(url_for('forgot_password'))

    return render_template('forgot_password.html')

# Herkese açık - Token ile Şifre Sıfırlama
@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def public_reset_password(token):
    prr = PasswordResetRequest.query.filter_by(token=token).first()
    if not prr:
        flash('Geçersiz şifre sıfırlama bağlantısı.', 'error')
        return redirect(url_for('forgot_password'))

    if prr.is_used:
        flash('Bu şifre sıfırlama bağlantısı zaten kullanılmış.', 'error')
        return redirect(url_for('login'))

    if prr.is_expired():
        flash('Şifre sıfırlama bağlantısının süresi dolmuş.', 'error')
        return redirect(url_for('forgot_password'))

    user = prr.user or User.query.filter_by(email=prr.email).first()
    if request.method == 'POST':
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')

        # Basit doğrulamalar
        if not new_password or not confirm_password:
            flash('Lütfen tüm alanları doldurun.', 'error')
            return redirect(url_for('public_reset_password', token=token))

        if new_password != confirm_password:
            flash('Parolalar eşleşmiyor.', 'error')
            return redirect(url_for('public_reset_password', token=token))

        if len(new_password) < 8:
            flash('Parola en az 8 karakter olmalıdır.', 'error')
            return redirect(url_for('public_reset_password', token=token))

        if not user:
            flash('Bu e-posta ile ilişkili aktif bir kullanıcı bulunamadı.', 'error')
            return redirect(url_for('forgot_password'))

        # Parolayı güncelle
        user.set_password(new_password)
        prr.is_used = True
        prr.used_at = get_turkey_time()
        db.session.commit()

        flash('Parola başarıyla güncellendi. Giriş yapabilirsiniz.', 'success')
        return redirect(url_for('login'))

    # GET: Formu göster
    return render_template('reset_password.html', token=token, email=prr.email)

# Admin - Token ile Şifre Sıfırlama
@app.route('/admin/reset_password/<token>', methods=['GET', 'POST'])
@login_required
def admin_reset_password(token):
    if not current_user.is_admin:
        flash('Bu sayfaya erişim yetkiniz yok!', 'error')
        return redirect(url_for('dashboard'))

    prr = PasswordResetRequest.query.filter_by(token=token).first()
    if not prr:
        flash('Geçersiz şifre sıfırlama bağlantısı.', 'error')
        return redirect(url_for('messages'))

    if prr.is_used:
        flash('Bu şifre sıfırlama bağlantısı zaten kullanılmış.', 'error')
        return redirect(url_for('messages'))

    if prr.is_expired():
        flash('Şifre sıfırlama bağlantısının süresi dolmuş.', 'error')
        return redirect(url_for('messages'))

    user = prr.user or User.query.filter_by(email=prr.email).first()
    if request.method == 'POST':
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')

        # Basit doğrulamalar
        if not new_password or not confirm_password:
            flash('Lütfen tüm alanları doldurun.', 'error')
            return redirect(url_for('admin_reset_password', token=token))

        if new_password != confirm_password:
            flash('Parolalar eşleşmiyor.', 'error')
            return redirect(url_for('admin_reset_password', token=token))

        if len(new_password) < 8:
            flash('Parola en az 8 karakter olmalıdır.', 'error')
            return redirect(url_for('admin_reset_password', token=token))

        if not user:
            flash('Bu e-posta ile ilişkili aktif bir kullanıcı bulunamadı.', 'error')
            return redirect(url_for('messages'))

        # Parolayı güncelle
        user.set_password(new_password)
        prr.is_used = True
        prr.used_at = get_turkey_time()
        db.session.commit()

        flash('Parola başarıyla güncellendi.', 'success')
        return redirect(url_for('messages'))

    # GET: Formu göster
    return render_template('admin_reset_password.html', token=token, user=user, email=prr.email)

# Mesajlar Sayfası
@app.route('/messages')
@login_required
def messages():
    # Kullanıcının mesajlarını getir (en yeni önce)
    user_messages = Message.query.filter_by(
        recipient_id=current_user.id
    ).order_by(Message.created_at.desc()).all()
    
    return render_template('messages.html', messages=user_messages)

# Mesaj Detayı
@app.route('/message/<int:message_id>')
@login_required
def message_detail(message_id):
    message = Message.query.filter_by(
        id=message_id,
        recipient_id=current_user.id
    ).first_or_404()
    
    # Mesajı okundu olarak işaretle
    message.mark_as_read()
    
    # Gönderen bilgisi
    sender_name = None
    if message.sender_id:
        sender_name = message.sender.name
    
    return jsonify({
        'success': True,
        'message': {
            'id': message.id,
            'title': message.title,
            'content': message.content,
            'created_at': message.created_at.strftime('%d.%m.%Y %H:%M'),
            'is_read': message.is_read,
            'is_admin_message': message.is_admin_message,
            'sender_name': sender_name
        }
    })

# Admin - Mesaj Gönderme Sayfası
@app.route('/admin/send_message', methods=['GET', 'POST'])
@login_required
def admin_send_message():
    if not current_user.is_admin:
        flash('Bu sayfaya erişim yetkiniz yok!', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            data = request.get_json()
            
            title = data.get('title', '').strip()
            content = data.get('content', '').strip()
            recipient_ids = data.get('recipient_ids', [])
            
            if not title or not content:
                return jsonify({
                    'success': False,
                    'error': 'Başlık ve mesaj içeriği gereklidir.'
                })
            
            if not recipient_ids:
                return jsonify({
                    'success': False,
                    'error': 'En az bir alıcı seçmelisiniz.'
                })
            
            # Mesajları kaydet ve push notification gönder
            for customer_id in recipient_ids:
                message = Message(
                    title=title,
                    content=content,
                    recipient_id=customer_id,
                    is_admin_message=True
                )
                db.session.add(message)
                
                # Push notification gönder
                send_push_notification(
                    user_id=customer_id,
                    title=f"📩 Yeni Mesaj: {title}",
                    body=content[:100] + "..." if len(content) > 100 else content,
                    notification_type="message",
                    url="/messages"
                )
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': f'{len(recipient_ids)} müşteriye mesaj başarıyla gönderildi!'
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({
                'success': False,
                'error': f'Mesaj gönderilirken hata oluştu: {str(e)}'
            })
    
    # GET request - müşteri listesini getir
    customers = User.query.filter_by(is_admin=False).order_by(User.name).all()
    return render_template('admin_send_message.html', customers=customers)

# Survey API Endpoints

# Create Survey (Admin)
@app.route('/admin/create_survey', methods=['POST'])
@login_required
def create_survey():
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': 'Yetkisiz erişim'}), 403
    
    try:
        data = request.get_json()
        
        title = data.get('title', '').strip()
        description = data.get('description', '').strip()
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date')
        is_active = data.get('is_active', True)
        questions = data.get('questions', [])
        
        if not title or not description:
            return jsonify({'success': False, 'error': 'Başlık ve açıklama gereklidir'})
        
        if not start_date_str or not end_date_str:
            return jsonify({'success': False, 'error': 'Başlangıç ve bitiş tarihleri gereklidir'})
        
        if not questions:
            return jsonify({'success': False, 'error': 'En az bir soru eklemelisiniz'})
        
        # Parse dates
        start_date = datetime.fromisoformat(start_date_str.replace('Z', '+00:00')).replace(tzinfo=None)
        end_date = datetime.fromisoformat(end_date_str.replace('Z', '+00:00')).replace(tzinfo=None)
        
        if start_date >= end_date:
            return jsonify({'success': False, 'error': 'Bitiş tarihi başlangıç tarihinden sonra olmalıdır'})
        
        # Create survey
        survey = Survey(
            title=title,
            description=description,
            start_date=start_date,
            end_date=end_date,
            is_active=is_active
        )
        db.session.add(survey)
        db.session.flush()  # Get survey ID
        
        # Create questions
        for question_data in questions:
            question = SurveyQuestion(
                survey_id=survey.id,
                question_text=question_data.get('text', ''),
                question_type=question_data.get('type', 'text'),
                question_order=question_data.get('order', 1),
                is_required=True
            )
            
            # Set options for multiple choice questions
            if question_data.get('type') == 'multiple_choice' and question_data.get('options'):
                question.set_options(question_data['options'])
            
            db.session.add(question)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Anket başarıyla oluşturuldu',
            'survey_id': survey.id
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# Get Surveys (Admin)
@app.route('/admin/get_surveys', methods=['GET'])
@login_required
def get_surveys():
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': 'Yetkisiz erişim'}), 403
    
    try:
        surveys = Survey.query.order_by(Survey.created_at.desc()).all()
        surveys_data = [survey.to_dict() for survey in surveys]
        
        return jsonify({
            'success': True,
            'surveys': surveys_data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Toggle Survey Status (Admin)
@app.route('/admin/toggle_survey/<int:survey_id>', methods=['POST'])
@login_required
def toggle_survey(survey_id):
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': 'Yetkisiz erişim'}), 403
    
    try:
        survey = Survey.query.get_or_404(survey_id)
        survey.is_active = not survey.is_active
        db.session.commit()
        
        status = 'aktif' if survey.is_active else 'pasif'
        return jsonify({
            'success': True,
            'message': f'Anket {status} yapıldı'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# Delete Survey (Admin)
@app.route('/admin/delete_survey/<int:survey_id>', methods=['DELETE'])
@login_required
def delete_survey(survey_id):
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': 'Yetkisiz erişim'}), 403
    
    try:
        survey = Survey.query.get_or_404(survey_id)
        db.session.delete(survey)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Anket başarıyla silindi'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# Get Active Surveys (Mobile API)
@app.route('/api/surveys/active', methods=['GET'])
def get_active_surveys():
    try:
        # Get authorization header
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({'success': False, 'error': 'Token gerekli'}), 401
        
        token = auth_header.split(' ')[1]
        
        # Find user by token (simplified - in production use proper JWT)
        user = User.query.filter_by(auth_token=token).first()
        if not user:
            return jsonify({'success': False, 'error': 'Token süresi dolmuş. Lütfen tekrar giriş yapın.'}), 401
        
        now = get_turkey_time().replace(tzinfo=None)
        surveys = Survey.query.filter(
            Survey.is_active == True,
            Survey.start_date <= now,
            Survey.end_date >= now
        ).order_by(Survey.created_at.desc()).all()
        
        # Filter out surveys user has already completed
        available_surveys = []
        for survey in surveys:
            existing_response = SurveyResponse.query.filter_by(
                survey_id=survey.id,
                user_id=user.id
            ).first()
            
            if not existing_response:
                survey_data = survey.to_dict()
                survey_data['questions'] = [q.to_dict() for q in survey.questions]
                available_surveys.append(survey_data)
        
        return jsonify({
            'success': True,
            'surveys': available_surveys
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Submit Survey Response (Mobile API)
@app.route('/api/surveys/<int:survey_id>/submit', methods=['POST'])
def submit_survey_response(survey_id):
    try:
        # Get authorization header
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({'success': False, 'error': 'Token gerekli'}), 401
        
        token = auth_header.split(' ')[1]
        
        # Find user by token
        user = User.query.filter_by(auth_token=token).first()
        if not user:
            return jsonify({'success': False, 'error': 'Geçersiz token'}), 401
        
        survey = Survey.query.get_or_404(survey_id)
        
        # Check if survey is active and valid
        if not survey.is_valid():
            return jsonify({'success': False, 'error': 'Anket aktif değil veya süresi dolmuş'}), 400
        
        # Check if user already responded
        existing_response = SurveyResponse.query.filter_by(
            survey_id=survey_id,
            user_id=user.id
        ).first()
        
        if existing_response:
            return jsonify({'success': False, 'error': 'Bu anketi zaten cevapladınız'}), 400
        
        data = request.get_json()
        answers = data.get('answers', [])
        
        if not answers:
            return jsonify({'success': False, 'error': 'Cevaplar gereklidir'}), 400
        
        # Create response
        response = SurveyResponse(
            survey_id=survey_id,
            user_id=user.id
        )
        db.session.add(response)
        db.session.flush()  # Get response ID
        
        # Create answers
        for answer_data in answers:
            question_id = answer_data.get('question_id')
            question = SurveyQuestion.query.get(question_id)
            
            if not question or question.survey_id != survey_id:
                continue
            
            answer = SurveyAnswer(
                response_id=response.id,
                question_id=question_id
            )
            
            # Set answer based on question type
            if question.question_type == 'rating':
                answer.answer_rating = answer_data.get('rating')
            elif question.question_type == 'text':
                answer.answer_text = answer_data.get('text')
            elif question.question_type == 'multiple_choice':
                answer.answer_choice = answer_data.get('choice')
            elif question.question_type == 'yes_no':
                answer.answer_boolean = answer_data.get('boolean')
            
            db.session.add(answer)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Anket cevabınız kaydedildi. Teşekkür ederiz!'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# View Survey Responses (Admin)
@app.route('/admin/survey_responses/<int:survey_id>')
@login_required
def view_survey_responses(survey_id):
    if not current_user.is_admin:
        flash('Bu sayfaya erişim yetkiniz yok!', 'error')
        return redirect(url_for('dashboard'))
    
    survey = Survey.query.get_or_404(survey_id)
    responses = SurveyResponse.query.filter_by(survey_id=survey_id).all()
    
    return render_template('survey_responses.html', survey=survey, responses=responses)

# Push Notification Subscription Model
class PushSubscription(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    endpoint = db.Column(db.Text, nullable=False)
    p256dh_key = db.Column(db.Text, nullable=False)
    auth_key = db.Column(db.Text, nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=get_turkey_time)
    
    # İlişki
    user = db.relationship('User', backref='push_subscriptions')

# VAPID Key API
@app.route('/api/vapid-key')
def get_vapid_key():
    public_key = os.getenv('VAPID_PUBLIC_KEY')
    if not public_key:
        return jsonify({'error': 'VAPID public key not configured'}), 500
    return jsonify({'publicKey': public_key})

# Push Notification Subscription
@app.route('/api/subscribe-notifications', methods=['POST'])
@login_required
def subscribe_notifications():
    try:
        data = request.get_json()
        subscription = data.get('subscription')
        
        if not subscription:
            return jsonify({'success': False, 'error': 'Subscription data required'})
        
        # Mevcut subscription'ı kontrol et
        existing = PushSubscription.query.filter_by(
            user_id=current_user.id,
            endpoint=subscription['endpoint']
        ).first()
        
        if existing:
            existing.is_active = True
            existing.p256dh_key = subscription['keys']['p256dh']
            existing.auth_key = subscription['keys']['auth']
        else:
            new_subscription = PushSubscription(
                user_id=current_user.id,
                endpoint=subscription['endpoint'],
                p256dh_key=subscription['keys']['p256dh'],
                auth_key=subscription['keys']['auth']
            )
            db.session.add(new_subscription)
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Subscription saved successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

# Push Notification Unsubscription
@app.route('/api/unsubscribe-notifications', methods=['POST'])
@login_required
def unsubscribe_notifications():
    try:
        # Kullanıcının tüm subscription'larını deaktif et
        PushSubscription.query.filter_by(user_id=current_user.id).update({'is_active': False})
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Unsubscribed successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

# Test Notification
@app.route('/api/test-notification', methods=['POST'])
@login_required
def send_test_notification():
    try:
        # Test bildirimi gönder
        send_push_notification(
            user_id=current_user.id,
            title="Test Bildirimi 🧪",
            body=f"Merhaba {current_user.name}! Bildirimler çalışıyor.",
            notification_type="test",
            url="/dashboard"
        )
        
        return jsonify({'success': True, 'message': 'Test notification sent'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

def send_push_notification(user_id, title, body, notification_type="general", url="/dashboard", data=None):
    """Push notification gönderme fonksiyonu"""
    try:
        # Kullanıcının aktif subscription'larını al
        subscriptions = PushSubscription.query.filter_by(
            user_id=user_id,
            is_active=True
        ).all()
        
        if not subscriptions:
            print(f"User {user_id} has no active push subscriptions")
            return False
        
        # Notification payload'ı hazırla
        notification_data = {
            'title': title,
            'body': body,
            'icon': '/static/icons/icon-192.png',
            'badge': '/static/icons/icon-192.png',
            'tag': f'cafe-{notification_type}',
            'requireInteraction': True,
            'data': {
                'type': notification_type,
                'url': url,
                'timestamp': get_turkey_time().isoformat(),
                **(data or {})
            }
        }
        
        # VAPID anahtarları environment variable'dan al
        vapid_private_key = os.getenv('VAPID_PRIVATE_KEY')
        vapid_claims = {
            "sub": f"mailto:{os.getenv('VAPID_CLAIMS_EMAIL', 'admin@reevpoints.com')}"
        }
        
        if not vapid_private_key:
            print("VAPID_PRIVATE_KEY environment variable not set")
            return False
        
        # Her subscription için bildirim gönder
        sent_count = 0
        for subscription in subscriptions:
            try:
                # Subscription bilgilerini parse et
                subscription_info = json.loads(subscription.subscription_json)
                
                # Push notification gönder
                webpush(
                    subscription_info=subscription_info,
                    data=json.dumps(notification_data),
                    vapid_private_key=vapid_private_key,
                    vapid_claims=vapid_claims
                )
                
                print(f"📱 Push notification sent to user {user_id}: {title}")
                sent_count += 1
                
            except WebPushException as e:
                print(f"WebPush failed for subscription {subscription.id}: {e}")
                # HTTP 410 (Gone) durumunda subscription'ı deaktif et
                if e.response and e.response.status_code == 410:
                    subscription.is_active = False
                    print(f"Subscription {subscription.id} marked as inactive (410 Gone)")
                    
            except Exception as e:
                print(f"Failed to send push notification to subscription {subscription.id}: {e}")
                # Genel hata durumunda subscription'ı deaktif et
                subscription.is_active = False
        
        if sent_count > 0:
            db.session.commit()
            return True
        else:
            return False
            
    except Exception as e:
        print(f"Error sending push notification: {e}")
        return False

def send_campaign_notification(campaign):
    """Yeni kampanya bildirimi gönder"""
    try:
        # Tüm müşterileri al (admin olmayanlar)
        customers = User.query.filter_by(is_admin=False).all()
        
        # Kampanya açıklamasını kısalt
        description = campaign.description
        if len(description) > 100:
            description = description[:100] + "..."
        
        # Her müşteriye bildirim gönder
        sent_count = 0
        for customer in customers:
            success = send_push_notification(
                user_id=customer.id,
                title=f"🎉 Yeni Kampanya: {campaign.title}",
                body=description,
                notification_type="campaign",
                url="/campaigns",
                data={
                    'campaign_id': campaign.id,
                    'campaign_title': campaign.title
                }
            )
            if success:
                sent_count += 1
        
        print(f"📢 Campaign notification sent to {sent_count} customers: {campaign.title}")
        return sent_count
        
    except Exception as e:
        print(f"Error sending campaign notifications: {e}")
        return 0

# Mobile API Endpoints for Flutter App
@app.route('/api/login', methods=['POST'])
def api_login():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
            
        email = data.get('email')
        password = data.get('password')
        
        if not email or not password:
            return jsonify({'success': False, 'error': 'Email and password required'}), 400
        
        user = User.query.filter_by(email=email).first()
        
        if user and user.check_password(password):
            if not user.is_verified:
                return jsonify({'success': False, 'error': 'Account not verified'}), 401
                
            # Generate and save auth token
            token = user.generate_auth_token()
            db.session.commit()
            
            return jsonify({
                'success': True,
                'token': token,
                'user': {
                    'id': user.id,
                    'name': user.name,
                    'email': user.email,
                    'phone': user.phone,
                    'points': user.points,
                    'language': user.language
                }
            })
        else:
            return jsonify({'success': False, 'error': 'Invalid credentials'}), 401
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/register', methods=['POST'])
def api_register():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
            
        name = data.get('name')
        email = data.get('email')
        phone = data.get('phone', '')
        password = data.get('password')
        
        if not all([name, email, password]):
            return jsonify({'success': False, 'error': 'Name, email and password required'}), 400
        
        # Check if user already exists
        if User.query.filter_by(email=email).first():
            return jsonify({'success': False, 'error': 'Email already registered'}), 400
        
        # Create new user
        user = User(
            name=name,
            email=email,
            phone=phone,
            is_verified=False
        )
        user.set_password(password)
        user.generate_verification_code()
        
        db.session.add(user)
        db.session.commit()
        
        # Send verification email
        email_sent = send_verification_email(user)
        
        return jsonify({
            'success': True,
            'message': 'Registration successful. Please check your email for verification code.',
            'user_id': user.id,
            'email_sent': email_sent
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/verify-email', methods=['POST'])
def api_verify_email():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
            
        user_id = data.get('user_id')
        verification_code = data.get('verification_code')
        
        if not user_id or not verification_code:
            return jsonify({'success': False, 'error': 'User ID and verification code required'}), 400
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 404
        
        if user.verification_code == verification_code:
            user.is_verified = True
            user.verification_code = None
            db.session.commit()
            
            # Generate token for auto-login
            token = secrets.token_urlsafe(32)
            
            return jsonify({
                'success': True,
                'message': 'Email verified successfully',
                'token': token,
                'user': {
                    'id': user.id,
                    'name': user.name,
                    'email': user.email,
                    'phone': user.phone,
                    'points': user.points,
                    'language': user.language
                }
            })
        else:
            return jsonify({'success': False, 'error': 'Invalid verification code'}), 400
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/dashboard', methods=['GET'])
def api_dashboard():
    try:
        # In a real app, you'd validate the token here
        # For now, we'll use a simple approach
        user_id = request.args.get('user_id')
        print(f"Dashboard request for user_id: {user_id}")
        
        if not user_id:
            return jsonify({'success': False, 'error': 'User ID required'}), 400
            
        user = User.query.get(user_id)
        if not user:
            print(f"User not found: {user_id}")
            return jsonify({'success': False, 'error': 'User not found'}), 404
        
        print(f"User found: {user.name} ({user.email})")
        
        # Bu ay kazanılan puanları customerQR tablosundan hesapla
        from datetime import datetime
        current_month = get_turkey_time().month
        current_year = get_turkey_time().year
        
        print(f"Calculating monthly points for month: {current_month}, year: {current_year}")
        
        # Önce hangi tabloda veri var kontrol edelim
        new_table_count = CustomerQRCode.query.filter_by(customer_id=user_id).count()
        old_table_count = CustomerQR.query.filter_by(customer_id=user_id).count()
        
        print(f"CustomerQRCode table records: {new_table_count}")
        print(f"CustomerQR table records: {old_table_count}")
        
        total_monthly_points = 0
        
        # Eğer yeni tabloda veri varsa, sadece yeni tablodan al
        if new_table_count > 0:
            monthly_qr_points = db.session.query(
                db.func.sum(CustomerQRCode.points_earned)
            ).filter(
                CustomerQRCode.customer_id == user_id,
                CustomerQRCode.is_used == True,
                db.func.extract('month', CustomerQRCode.used_at) == current_month,
                db.func.extract('year', CustomerQRCode.used_at) == current_year
            ).scalar() or 0
            
            total_monthly_points = int(monthly_qr_points)
            print(f"Using new table - Monthly points: {total_monthly_points}")
        else:
            # Eğer yeni tabloda veri yoksa, eski tablodan al
            monthly_old_qr_count = db.session.query(
                db.func.count(CustomerQR.id)
            ).filter(
                CustomerQR.customer_id == user_id,
                CustomerQR.used_at.isnot(None),
                db.func.extract('month', CustomerQR.used_at) == current_month,
                db.func.extract('year', CustomerQR.used_at) == current_year
            ).scalar() or 0
            
            # Her eski QR kod 1 puan
            total_monthly_points = int(monthly_old_qr_count)
            print(f"Using old table - Monthly QR count: {monthly_old_qr_count}, points: {total_monthly_points}")
        
        print(f"Final monthly points: {total_monthly_points}")
        
        # Toplam işlem sayısını sadece ProductRedemption tablosundan al
        total_transactions = ProductRedemption.query.filter_by(user_id=user_id).count()
        
        print(f"Total transactions from ProductRedemption table: {total_transactions}")
        
        # Get recent transactions
        recent_transactions = Transaction.query.filter_by(user_id=user.id).order_by(Transaction.timestamp.desc()).limit(5).all()
        
        # Get campaign usages
        campaign_usages = CampaignUsage.query.filter_by(
            customer_id=user.id,
            is_used=True
        ).order_by(CampaignUsage.used_at.desc()).limit(10).all()
        
        transactions_data = []
        for transaction in recent_transactions:
            transactions_data.append({
                'id': transaction.id,
                'amount': transaction.amount,
                'points_earned': transaction.points_earned,
                'points_used': transaction.points_used,
                'transaction_type': transaction.transaction_type,
                'description': transaction.description,
                'timestamp': transaction.timestamp.isoformat() if transaction.timestamp else None
            })
        
        campaign_activities = []
        for usage in campaign_usages:
            campaign_activities.append({
                'campaign_title': usage.campaign.title if usage.campaign else 'Unknown Campaign',
                'used_at': usage.used_at.isoformat() if usage.used_at else None,
                'branch_name': usage.used_by_branch.name if usage.used_by_branch else 'Unknown Branch',
                'selected_product_name': usage.selected_product_name
            })
        
        # En çok ziyaret edilen şube bilgisi
        most_visited_branch = None
        try:
            from sqlalchemy import func
            # Hem yeni hem eski tablodan şube ziyaretlerini al
            branch_visits = []
            
            # Yeni tablodan (CustomerQRCode)
            if new_table_count > 0:
                new_visits = db.session.query(
                    CustomerQRCode.used_by_branch_id, 
                    func.count(CustomerQRCode.id).label('visit_count')
                ).filter(
                    CustomerQRCode.customer_id == user_id,
                    CustomerQRCode.is_used == True,
                    CustomerQRCode.used_by_branch_id.isnot(None)
                ).group_by(CustomerQRCode.used_by_branch_id).all()
                branch_visits.extend(new_visits)
            
            # Eski tablodan (CustomerQR)
            if old_table_count > 0:
                old_visits = db.session.query(
                    CustomerQR.used_by_branch_id, 
                    func.count(CustomerQR.id).label('visit_count')
                ).filter(
                    CustomerQR.customer_id == user_id,
                    CustomerQR.is_used == True,
                    CustomerQR.used_by_branch_id.isnot(None)
                ).group_by(CustomerQR.used_by_branch_id).all()
                branch_visits.extend(old_visits)
            
            # Şube ziyaretlerini birleştir ve en çok ziyaret edileni bul
            if branch_visits:
                branch_totals = {}
                for branch_id, count in branch_visits:
                    if branch_id in branch_totals:
                        branch_totals[branch_id] += count
                    else:
                        branch_totals[branch_id] = count
                
                # En çok ziyaret edilen şubeyi bul
                most_visited_branch_id = max(branch_totals, key=branch_totals.get)
                most_visit_count = branch_totals[most_visited_branch_id]
                
                # Şube bilgilerini al
                branch = Branch.query.get(most_visited_branch_id)
                if branch:
                    most_visited_branch = {
                        'id': branch.id,
                        'name': branch.name,
                        'address': branch.address,
                        'phone': branch.phone,
                        'visit_count': most_visit_count,
                        'working_hours': branch.working_hours
                    }
                    
        except Exception as e:
            print(f"Error calculating most visited branch: {e}")
            most_visited_branch = None
        
        # En sevilen ürün bilgisi (gerçek satın alma geçmişinden)
        most_favorite_product = None
        try:
            # Debug: ProductRedemption tablosundaki veri sayısını kontrol et
            total_redemptions = ProductRedemption.query.filter_by(user_id=user_id).count()
            confirmed_redemptions = ProductRedemption.query.filter_by(user_id=user_id, is_confirmed=True).count()
            print(f"Total redemptions for user {user_id}: {total_redemptions}")
            print(f"Confirmed redemptions for user {user_id}: {confirmed_redemptions}")
            
            # Tüm redemption'ları listele (debug için)
            all_redemptions = ProductRedemption.query.filter_by(user_id=user_id).all()
            for redemption in all_redemptions:
                product_name = redemption.product.name if redemption.product else "Unknown Product"
                print(f"Redemption: {product_name}, confirmed: {redemption.is_confirmed}")
            
            # ProductRedemption tablosundan en çok satın alınan ürünü bul (Product tablosu ile join)
            product_redemption = db.session.query(
                Product.name,
                Product.points_required,
                func.count(ProductRedemption.id).label('purchase_count')
            ).join(
                Product, ProductRedemption.product_id == Product.id
            ).filter(
                ProductRedemption.user_id == user_id,
                ProductRedemption.is_confirmed == True
            ).group_by(Product.name, Product.points_required).order_by(
                func.count(ProductRedemption.id).desc()
            ).first()
            
            if product_redemption:
                most_favorite_product = {
                    'name': product_redemption.name,
                    'redemption_count': product_redemption.purchase_count,
                    'points_required': product_redemption.points_required,
                    'is_suggested': False
                }
                print(f"Most favorite product from purchase history: {most_favorite_product}")
            else:
                print("No favorite product found - user has no confirmed purchases yet")
                # Eğer confirmed purchase yoksa, tüm purchase'lara bak
                any_product_redemption = db.session.query(
                    Product.name,
                    Product.points_required,
                    func.count(ProductRedemption.id).label('purchase_count')
                ).join(
                    Product, ProductRedemption.product_id == Product.id
                ).filter(
                    ProductRedemption.user_id == user_id
                ).group_by(Product.name, Product.points_required).order_by(
                    func.count(ProductRedemption.id).desc()
                ).first()
                
                if any_product_redemption:
                    most_favorite_product = {
                        'name': any_product_redemption.name,
                        'redemption_count': any_product_redemption.purchase_count,
                        'points_required': any_product_redemption.points_required,
                        'is_suggested': False
                    }
                    print(f"Most favorite product from any purchase history: {most_favorite_product}")
                
        except Exception as e:
            print(f"Error calculating most favorite product: {e}")
            most_favorite_product = None
        
        return jsonify({
            'success': True,
            'user': {
                'id': user.id,
                'name': user.name,
                'email': user.email,
                'points': user.points
            },
            'monthly_points': total_monthly_points,
            'total_transactions': total_transactions,
            'recent_transactions': transactions_data,
            'campaign_activities': campaign_activities,
            'most_visited_branch': most_visited_branch,
            'most_favorite_product': most_favorite_product
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/campaigns', methods=['GET'])
def api_campaigns():
    try:
        print("Campaigns API called")
        # Get active campaigns
        now = get_turkey_time().replace(tzinfo=None)
        print(f"Current time: {now}")
        
        campaigns = Campaign.query.filter(
            Campaign.is_active == True,
            Campaign.start_date <= now,
            Campaign.end_date >= now
        ).all()
        
        print(f"Found {len(campaigns)} active campaigns")
        
        campaigns_data = []
        for campaign in campaigns:
            # Kampanya resmi URL'ini oluştur
            image_url = None
            if campaign.image_filename:
                image_url = url_for('static', filename=f'uploads/{campaign.image_filename}', _external=True)
            
            # Kampanya ürünlerini al - geçici olarak basit yaklaşım
            products_data = []
            try:
                # Basit test ürünü ekle
                products_data.append({
                    'id': 1,
                    'name': 'Test Ürün',
                    'points': 10,
                    'discount': 20,
                    'description': 'Test açıklaması'
                })
            except Exception as products_error:
                print(f"Error processing campaign products for campaign {campaign.id}: {products_error}")
                products_data = []
            
            campaigns_data.append({
                'id': campaign.id,
                'title': campaign.title,
                'description': campaign.description,
                'image_filename': campaign.image_filename,
                'image_url': image_url,
                'start_date': campaign.start_date.isoformat() if campaign.start_date else None,
                'end_date': campaign.end_date.isoformat() if campaign.end_date else None,
                'qr_enabled': campaign.qr_enabled,
                'products': products_data
            })
        
        return jsonify({
            'success': True,
            'campaigns': campaigns_data
        })
        
    except Exception as e:
        print(f"Campaigns API error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/resend-verification', methods=['POST'])
def api_resend_verification():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
            
        user_id = data.get('user_id')
        if not user_id:
            return jsonify({'success': False, 'error': 'User ID required'}), 400
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 404
        
        # Generate new verification code
        user.generate_verification_code()
        db.session.commit()
        
        # Send verification email
        email_sent = send_verification_email(user)
        
        return jsonify({
            'success': True,
            'message': 'Verification code resent successfully',
            'email_sent': email_sent
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/redeem', methods=['GET', 'POST'])
def api_redeem():
    try:
        if request.method == 'GET':
            # Get user points
            user_id = request.args.get('user_id')
            user_points = 0
            if user_id:
                user = User.query.get(user_id)
                if user:
                    user_points = user.points
            
            # Get available products for redemption
            products = Product.query.filter_by(is_active=True).all()
            
            # Get categories
            categories = Category.query.filter_by(is_active=True).all()
            categories_data = []
            for category in categories:
                categories_data.append({
                    'id': category.id,
                    'name': category.name,
                    'description': category.description
                })
            
            products_data = []
            for product in products:
                # Check if image file actually exists before creating URL
                image_url = None
                if product.image_filename:
                    import os
                    image_path = os.path.join('static', 'uploads', product.image_filename)
                    if os.path.exists(image_path):
                        image_url = f"{request.host_url}static/uploads/{product.image_filename}"
                
                products_data.append({
                    'id': product.id,
                    'name': product.name,
                    'description': product.description,
                    'points': product.points_required,  # Mobile app expects 'points'
                    'category_id': product.category_id,
                    'category_name': product.category_ref.name if product.category_ref else 'Diğer',
                    'image_url': image_url,  # Will be None if file doesn't exist
                    'image_filename': product.image_filename
                })
            
            return jsonify({
                'success': True,
                'products': products_data,
                'categories': categories_data,
                'user_points': user_points
            })
            
        elif request.method == 'POST':
            # Redeem a product
            data = request.get_json()
            if not data:
                return jsonify({'success': False, 'error': 'No data provided'}), 400
                
            user_id = data.get('user_id')
            product_id = data.get('product_id')
            
            if not user_id or not product_id:
                return jsonify({'success': False, 'error': 'User ID and Product ID required'}), 400
            
            user = User.query.get(user_id)
            if not user:
                return jsonify({'success': False, 'error': 'User not found'}), 404
                
            product = Product.query.get(product_id)
            if not product or not product.is_active:
                return jsonify({'success': False, 'error': 'Product not found or inactive'}), 404
            
            # Check if user has enough points
            if user.points < product.points_required:
                return jsonify({'success': False, 'error': 'Insufficient points'}), 400
            
            # Generate confirmation code
            confirmation_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
            
            # Create redemption record
            redemption = ProductRedemption(
                user_id=user.id,
                product_id=product.id,
                points_used=product.points_required,
                confirmation_code=confirmation_code
            )
            
            # Deduct points from user
            user.points -= product.points_required
            
            # Create transaction record
            transaction = Transaction(
                user_id=user.id,
                amount=0,
                points_used=product.points_required,
                transaction_type='redeem',
                description=f'Redeemed: {product.name}'
            )
            
            db.session.add(redemption)
            db.session.add(transaction)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Product redeemed successfully',
                'confirmation_code': confirmation_code,
                'remaining_points': user.points
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/generate-qr', methods=['POST'])
def api_generate_qr():
    try:
        data = request.get_json() or {}
        
        # Get user_id from request data or from query params
        user_id = data.get('user_id') or request.args.get('user_id')
        
        # If no user_id provided, try to get from authorization header
        if not user_id:
            # For now, we'll create a simple QR code for the user
            # In a real app, you'd decode the JWT token to get user_id
            return jsonify({'success': False, 'error': 'User ID required'}), 400
        
        campaign_id = data.get('campaign_id')
        selected_product_id = data.get('selected_product_id')
        
        # If no campaign_id provided, generate a simple user QR code
        if not campaign_id:
            user = User.query.get(user_id)
            if not user:
                return jsonify({'success': False, 'error': 'User not found'}), 404
            
            # Generate simple QR code for user identification
            qr_code = f"REEVUSER{user_id}{secrets.token_urlsafe(8).replace('_', '').replace('-', '')}"
            
            # Set expiry time (5 minutes from now)
            expires_at = get_turkey_time() + timedelta(minutes=5)
            
            # Save QR code to database (sadece mevcut alanları kullan)
            customer_qr = CustomerQRCode(
                customer_id=user_id,
                code=qr_code,
                points_earned=1
            )
            
            db.session.add(customer_qr)
            db.session.commit()
            
            print(f"QR code saved to database: {qr_code} for user {user_id}")
            
            return jsonify({
                'success': True,
                'qr_data': qr_code,
                'message': 'QR code generated successfully',
                'expires_at': expires_at.isoformat()
            })
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 404
            
        campaign = Campaign.query.get(campaign_id)
        if not campaign or not campaign.is_valid():
            return jsonify({'success': False, 'error': 'Campaign not found or inactive'}), 404
        
        # Check if user can use this campaign
        if not campaign.can_be_used_by_customer(user_id):
            return jsonify({'success': False, 'error': 'Campaign usage limit exceeded'}), 400
        
        # Generate unique QR code
        qr_code = f"REEV{campaign_id}{user_id}{secrets.token_urlsafe(8).replace('_', '').replace('-', '')}"
        
        # Set expiry time (5 minutes from now)
        expires_at = get_turkey_time() + timedelta(minutes=5)
        
        # Get selected product details
        selected_product_details = None
        selected_product_name = None
        if selected_product_id:
            campaign_product = CampaignProduct.query.get(selected_product_id)
            if campaign_product:
                selected_product_name = campaign_product.product_name
                selected_product_details = json.dumps({
                    'product_id': campaign_product.id,
                    'product_name': campaign_product.product_name,
                    'discount_type': campaign_product.discount_type,
                    'discount_value': campaign_product.discount_value
                })
        
        # Save QR code to CustomerQR table (sadece mevcut alanları kullan)
        customer_qr = CustomerQRCode(
            customer_id=user_id,
            code=qr_code,
            points_earned=1
        )
        
        # Create campaign usage record (for backward compatibility)
        usage = CampaignUsage(
            campaign_id=campaign_id,
            customer_id=user_id,
            qr_code=qr_code,
            expires_at=expires_at,
            selected_campaign_product_id=selected_product_id,
            selected_product_name=selected_product_name,
            selected_product_details=selected_product_details
        )
        
        db.session.add(customer_qr)
        db.session.add(usage)
        db.session.commit()
        
        print(f"Campaign QR code saved to database: {qr_code} for user {user_id}, campaign {campaign_id}")
        
        return jsonify({
            'success': True,
            'qr_data': qr_code,
            'expires_at': expires_at.isoformat(),
            'campaign_title': campaign.title,
            'message': 'Campaign QR code generated successfully'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/user-qr-codes', methods=['GET'])
def api_user_qr_codes():
    try:
        user_id = request.args.get('user_id')
        if not user_id:
            return jsonify({'success': False, 'error': 'User ID required'}), 400
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 404
        
        # Get user's QR codes (active and recent expired ones)
        qr_codes = CustomerQRCode.query.filter_by(customer_id=user_id).order_by(CustomerQRCode.created_at.desc()).limit(10).all()
        
        qr_codes_data = []
        for qr in qr_codes:
            qr_codes_data.append({
                'id': qr.id,
                'qr_code': qr.code,
                'customer_id': qr.customer_id,
                'points_earned': qr.points_earned,
                'is_used': qr.is_used,
                'is_expired': qr.is_expired(),
                'is_valid': qr.is_valid(),
                'created_at': qr.created_at.isoformat() if qr.created_at else None,
                'used_at': qr.used_at.isoformat() if qr.used_at else None,
                'used_by_branch_id': qr.used_by_branch_id
            })
        
        return jsonify({
            'success': True,
            'qr_codes': qr_codes_data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/scan-qr', methods=['POST'])
def api_scan_qr():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
            
        qr_code = data.get('qr_code')
        branch_id = data.get('branch_id')
        
        if not qr_code:
            return jsonify({'success': False, 'error': 'QR code required'}), 400
        
        # Find campaign usage by QR code
        usage = CampaignUsage.query.filter_by(qr_code=qr_code).first()
        if not usage:
            return jsonify({'success': False, 'error': 'Invalid QR code'}), 404
        
        # Check if QR code can be used
        if not usage.can_be_used(branch_id):
            if usage.is_used:
                return jsonify({'success': False, 'error': 'QR code already used'}), 400
            elif usage.is_expired():
                return jsonify({'success': False, 'error': 'QR code expired'}), 400
            else:
                return jsonify({'success': False, 'error': 'QR code cannot be used at this branch'}), 400
        
        # Mark as used
        usage.is_used = True
        usage.used_at = get_turkey_time()
        usage.used_by_branch_id = branch_id
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'QR code scanned successfully',
            'campaign_title': usage.campaign.title if usage.campaign else 'Unknown Campaign',
            'customer_name': usage.customer.name if usage.customer else 'Unknown Customer',
            'selected_product': usage.selected_product_name
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/profile', methods=['GET', 'PUT'])
def api_profile():
    try:
        if request.method == 'GET':
            user_id = request.args.get('user_id')
            if not user_id:
                return jsonify({'success': False, 'error': 'User ID required'}), 400
                
            user = User.query.get(user_id)
            if not user:
                return jsonify({'success': False, 'error': 'User not found'}), 404
            
            return jsonify({
                'success': True,
                'user': {
                    'id': user.id,
                    'name': user.name,
                    'email': user.email,
                    'phone': user.phone,
                    'points': user.points,
                    'language': user.language,
                    'preferred_branch_id': user.preferred_branch_id
                }
            })
            
        elif request.method == 'PUT':
            data = request.get_json()
            if not data:
                return jsonify({'success': False, 'error': 'No data provided'}), 400
                
            user_id = data.get('user_id')
            if not user_id:
                return jsonify({'success': False, 'error': 'User ID required'}), 400
            
            user = User.query.get(user_id)
            if not user:
                return jsonify({'success': False, 'error': 'User not found'}), 404
            
            # Update user fields
            if 'name' in data:
                user.name = data['name']
            if 'phone' in data:
                user.phone = data['phone']
            if 'language' in data:
                user.language = data['language']
            if 'preferred_branch_id' in data:
                user.preferred_branch_id = data['preferred_branch_id']
            
            user.updated_at = get_turkey_time()
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Profile updated successfully',
                'user': {
                    'id': user.id,
                    'name': user.name,
                    'email': user.email,
                    'phone': user.phone,
                    'points': user.points,
                    'language': user.language,
                    'preferred_branch_id': user.preferred_branch_id
                }
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/redeem-product', methods=['POST'])
def api_redeem_product():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
            
        user_id = data.get('user_id')
        product_id = data.get('product_id')
        
        if not user_id or not product_id:
            return jsonify({'success': False, 'error': 'User ID and Product ID required'}), 400
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 404
            
        product = Product.query.get(product_id)
        if not product:
            return jsonify({'success': False, 'error': 'Product not found'}), 404
        
        # Check if user has enough points
        if user.points < product.points_required:
            return jsonify({'success': False, 'error': 'Insufficient points'}), 400
        
        # Deduct points from user
        user.points -= product.points_required
        
        # Create transaction record
        transaction = Transaction(
            user_id=user.id,
            amount=0.0,  # Set amount to 0 for point redemptions
            transaction_type='redeem',
            points_used=product.points_required,
            description=f'{product.name} ürünü {product.points_required} puan ile alındı',
            timestamp=get_turkey_time()
        )
        
        # Generate 6-digit confirmation code
        import random
        import string
        confirmation_code = ''.join(random.choices(string.digits, k=6))
        
        # Generate QR code for confirmation
        import qrcode
        from io import BytesIO
        import base64
        
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr_data = f"REEV_REDEEM_{confirmation_code}_{user.id}_{product.id}"
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        qr_img = qr.make_image(fill_color="black", back_color="white")
        buffer = BytesIO()
        qr_img.save(buffer, format='PNG')
        qr_base64 = base64.b64encode(buffer.getvalue()).decode()
        
        # Create product redemption record
        redemption = ProductRedemption(
            user_id=user.id,
            product_id=product.id,
            points_used=product.points_required,
            confirmation_code=confirmation_code,
            redeemed_at=get_turkey_time()
        )
        
        db.session.add(transaction)
        db.session.add(redemption)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{product.name} başarıyla satın alındı!',
            'confirmation_code': confirmation_code,
            'qr_code_data': qr_data,
            'qr_code_image': f"data:image/png;base64,{qr_base64}",
            'remaining_points': user.points
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/messages', methods=['GET'])
def api_messages():
    try:
        user_id = request.args.get('user_id')
        print(f"Messages request for user_id: {user_id}")
        
        if not user_id:
            return jsonify({'success': False, 'error': 'User ID required'}), 400
            
        user = User.query.get(user_id)
        if not user:
            print(f"User not found: {user_id}")
            return jsonify({'success': False, 'error': 'User not found'}), 404
        
        print(f"User found: {user.name} ({user.email})")
        
        # Kullanıcının mesajlarını getir (en yeni önce)
        user_messages = Message.query.filter_by(recipient_id=user_id).order_by(
            Message.created_at.desc()
        ).all()
        
        print(f"Message kayıt sayısı: {len(user_messages)}")
        
        messages = []
        for message in user_messages:
            messages.append({
                'id': message.id,
                'title': message.title,
                'content': message.content,
                'is_read': message.is_read,
                'is_admin_message': message.is_admin_message,
                'created_at': message.created_at.isoformat() if message.created_at else None,
                'read_at': message.read_at.isoformat() if message.read_at else None,
                'sender_name': message.sender.name if message.sender else 'Sistem'
            })
        
        print(f"Toplam mesaj sayısı: {len(messages)}")
        
        return jsonify({
            'success': True,
            'messages': messages,
            'total_count': len(messages)
        })
        
    except Exception as e:
        print(f"Messages API error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/request-password-reset', methods=['POST'])
def api_request_password_reset():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        email = data.get('email')
        if not email:
            return jsonify({'success': False, 'error': 'Email required'}), 400
    
        user = User.query.filter_by(email=email).first()
        if not user:
            return jsonify({'success': False, 'error': 'Email not found'}), 404
    
        # Generate reset code (6 digits)
        reset_code = ''.join([str(random.randint(0, 9)) for _ in range(6)])
        
        # Save or update password reset request
        reset_request = PasswordResetRequest.query.filter_by(email=email, is_used=False).first()
        expires_at = get_turkey_time() + timedelta(minutes=15)  # 15 dakika geçerlilik
        
        if reset_request:
            reset_request.token = reset_code
            reset_request.created_at = get_turkey_time()
            reset_request.expires_at = expires_at
        else:
            reset_request = PasswordResetRequest(
                user_id=user.id,
                email=email,
                token=reset_code,
                expires_at=expires_at
            )
            db.session.add(reset_request)
        
        db.session.commit()
        
        # Send password reset email
        reset_link = f"Reset Code: {reset_code}"  # For mobile app, we just need the code
        email_sent = send_password_reset_email(user, reset_link)
        
        return jsonify({
            'success': True,
            'message': 'Password reset code sent to your email',
            'email_sent': email_sent
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reset-password', methods=['POST'])
def api_reset_password():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        email = data.get('email')
        reset_code = data.get('reset_code')
        new_password = data.get('new_password')
        
        if not all([email, reset_code, new_password]):
            return jsonify({'success': False, 'error': 'Email, reset code and new password required'}), 400
        
        # Find valid reset request
        reset_request = PasswordResetRequest.query.filter_by(
            email=email,
            token=reset_code,
            is_used=False
        ).first()
        
        if not reset_request:
            return jsonify({'success': False, 'error': 'Invalid or expired reset code'}), 400
        
        # Check if reset code is not expired
        if reset_request.is_expired():
            return jsonify({'success': False, 'error': 'Reset code has expired'}), 400
        
        # Get user and update password
        user = User.query.get(reset_request.user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 404
        
        user.set_password(new_password)
        reset_request.is_used = True
        reset_request.used_at = get_turkey_time()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Password reset successfully'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Product Request API (creates unconfirmed redemption)
@app.route('/api/request-product', methods=['POST'])
def api_request_product():
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        product_id = data.get('product_id')
        
        if not user_id or not product_id:
            return jsonify({'success': False, 'message': 'Eksik parametreler'}), 400
        
        # Kullanıcı kontrolü
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': 'Kullanıcı bulunamadı'}), 404
        
        # Ürün kontrolü
        product = Product.query.get(product_id)
        if not product or not product.is_active:
            return jsonify({'success': False, 'message': 'Ürün bulunamadı'}), 404
        
        # Yeterli puan kontrolü
        if user.points < product.points_required:
            return jsonify({'success': False, 'message': 'Yetersiz puan'}), 400
        
        # Aynı ürün için birden fazla talep edilebilir - kontrol kaldırıldı
        
        # 6 haneli QR kod oluştur (sadece rakam)
        qr_code = ''.join(random.choices(string.digits, k=6))
        confirmation_code = qr_code  # 6 haneli kod hem QR kod hem onay kodu
        
        print(f"🔍 Generated QR Code: {qr_code}")
        print(f"🔍 Confirmation Code: {confirmation_code}")
        
        # Ürün talebini oluştur (henüz puan düşülmez, is_confirmed=False)
        redemption = ProductRedemption(
            user_id=user_id,
            product_id=product_id,
            points_used=product.points_required,
            confirmation_code=confirmation_code,
            is_confirmed=False  # Henüz onaylanmadı
        )
        
        db.session.add(redemption)
        db.session.commit()
        
        response_data = {
            'success': True,
            'message': f'{product.name} için talebiniz oluşturuldu',
            'qr_code': qr_code,
            'confirmation_code': confirmation_code,
            'redemption_id': redemption.id
        }
        
        print(f"🔍 API Response: {response_data}")
        
        return jsonify(response_data)
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500

# Pending Products API
@app.route('/api/pending-products', methods=['GET'])
def api_pending_products():
    try:
        user_id = request.args.get('user_id')
        
        if not user_id:
            return jsonify({'success': False, 'message': 'User ID gerekli'}), 400
        
        # Kullanıcının onaylanmamış ürün taleplerini getir
        pending_redemptions = db.session.query(ProductRedemption, Product).join(
            Product, ProductRedemption.product_id == Product.id
        ).filter(
            ProductRedemption.user_id == user_id,
            ProductRedemption.is_confirmed == False
        ).order_by(ProductRedemption.redeemed_at.desc()).all()
        
        pending_products = []
        for redemption, product in pending_redemptions:
            # Ürün görsel URL'si oluştur
            image_url = None
            if product.image_filename:
                image_url = f"{request.url_root}static/uploads/{product.image_filename}"
            
            pending_products.append({
                'id': product.id,
                'redemption_id': redemption.id,
                'name': product.name,
                'description': product.description,
                'image_url': image_url,
                'points': redemption.points_used,
                'status': 'pending',
                'request_date': redemption.redeemed_at.strftime('%d.%m.%Y %H:%M'),
                'confirmation_code': redemption.confirmation_code,
                'qr_code': redemption.confirmation_code  # QR kod ve confirmation_code aynı
            })
        
        return jsonify({
            'success': True,
            'pending_products': pending_products
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500

# Approve Product API
@app.route('/api/approve-product', methods=['POST'])
def api_approve_product():
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        product_id = data.get('product_id')
        
        if not user_id or not product_id:
            return jsonify({'success': False, 'message': 'Eksik parametreler'}), 400
        
        # Bekleyen talebi bul
        redemption = ProductRedemption.query.filter_by(
            user_id=user_id,
            product_id=product_id,
            is_confirmed=False
        ).first()
        
        if not redemption:
            return jsonify({'success': False, 'message': 'Bekleyen talep bulunamadı'}), 404
        
        # Kullanıcı ve ürün bilgilerini al
        user = User.query.get(user_id)
        product = Product.query.get(product_id)
        
        if not user or not product:
            return jsonify({'success': False, 'message': 'Kullanıcı veya ürün bulunamadı'}), 404
        
        # Yeterli puan kontrolü (onay anında tekrar kontrol)
        if user.points < redemption.points_used:
            return jsonify({'success': False, 'message': 'Yetersiz puan'}), 400
        
        # Puanları düş ve talebi onayla
        user.points -= redemption.points_used
        redemption.is_confirmed = True
        redemption.confirmed_at = get_turkey_time()
        
        # Transaction kaydı oluştur
        transaction = Transaction(
            user_id=user_id,
            amount=0,  # Puan kullanımı için amount 0
            points_used=redemption.points_used,
            transaction_type='redeem',
            description=f'{product.name} ürünü için puan kullanımı'
        )
        
        db.session.add(transaction)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{product.name} başarıyla onaylandı',
            'points_deducted': redemption.points_used,
            'new_user_points': user.points,
            'confirmation_code': redemption.confirmation_code
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500

# Branch Panel - QR Code ile ürün onaylama
@app.route('/api/branch/approve-by-qr', methods=['POST'])
def api_branch_approve_by_qr():
    try:
        data = request.get_json()
        qr_code = data.get('qr_code')
        branch_id = data.get('branch_id')  # Şube ID'si
        
        if not qr_code:
            return jsonify({'success': False, 'message': 'QR kod gerekli'}), 400
        
        # QR kod (confirmation_code) ile bekleyen talebi bul
        redemption = ProductRedemption.query.filter_by(
            confirmation_code=qr_code,
            is_confirmed=False
        ).first()
        
        if not redemption:
            return jsonify({'success': False, 'message': 'Geçersiz QR kod veya zaten onaylanmış'}), 404
        
        # Kullanıcı ve ürün bilgilerini al
        user = User.query.get(redemption.user_id)
        product = Product.query.get(redemption.product_id)
        
        if not user or not product:
            return jsonify({'success': False, 'message': 'Kullanıcı veya ürün bulunamadı'}), 404
        
        # Yeterli puan kontrolü (onay anında tekrar kontrol)
        if user.points < redemption.points_used:
            return jsonify({'success': False, 'message': 'Kullanıcının yetersiz puanı var'}), 400
        
        # Puanları düş ve talebi onayla
        user.points -= redemption.points_used
        redemption.is_confirmed = True
        redemption.confirmed_at = get_turkey_time()
        redemption.confirmed_by_branch_id = branch_id
        
        # Transaction kaydı oluştur
        transaction = Transaction(
            user_id=redemption.user_id,
            amount=0,  # Puan kullanımı için amount 0
            points_used=redemption.points_used,
            transaction_type='redeem',
            description=f'{product.name} ürünü için puan kullanımı (Şube onayı)'
        )
        
        db.session.add(transaction)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{product.name} başarıyla onaylandı',
            'customer_name': user.name,
            'product_name': product.name,
            'points_deducted': redemption.points_used,
            'customer_remaining_points': user.points,
            'confirmation_code': redemption.confirmation_code
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500

# Branch Panel - Bekleyen talepleri listele
@app.route('/api/branch/pending-redemptions', methods=['GET'])
def api_branch_pending_redemptions():
    try:
        # Tüm bekleyen talepleri getir
        pending_redemptions = db.session.query(ProductRedemption, Product, User).join(
            Product, ProductRedemption.product_id == Product.id
        ).join(
            User, ProductRedemption.user_id == User.id
        ).filter(
            ProductRedemption.is_confirmed == False
        ).order_by(ProductRedemption.redeemed_at.desc()).all()
        
        pending_list = []
        for redemption, product, user in pending_redemptions:
            # Ürün görsel URL'si oluştur
            image_url = None
            if product.image_filename:
                image_url = f"{request.url_root}static/uploads/{product.image_filename}"
            
            pending_list.append({
                'redemption_id': redemption.id,
                'qr_code': redemption.confirmation_code,  # QR kod olarak confirmation_code kullan
                'confirmation_code': redemption.confirmation_code,
                'customer_name': user.name,
                'customer_email': user.email,
                'customer_phone': user.phone,
                'product_name': product.name,
                'product_description': product.description,
                'image_url': image_url,
                'points_required': redemption.points_used,
                'request_date': redemption.redeemed_at.strftime('%d.%m.%Y %H:%M'),
                'customer_current_points': user.points
            })
        
        return jsonify({
            'success': True,
            'pending_redemptions': pending_list,
            'total_count': len(pending_list)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500

# Branches API - Şubeleri listele
@app.route('/api/branches', methods=['GET'])
def api_branches():
    try:
        # Aktif şubeleri getir
        branches = Branch.query.filter_by(is_active=True).all()
        
        branches_list = []
        for branch in branches:
            # Şube görsel URL'si oluştur
            image_url = None
            if branch.image:
                image_url = f"{request.url_root}static/uploads/{branch.image}"
            
            branches_list.append({
                'id': branch.id,
                'name': branch.name,
                'address': branch.address,
                'phone': branch.phone,
                'email': branch.email,
                'image_url': image_url,
                'working_hours': branch.working_hours,
                'is_active': branch.is_active,
                'created_at': branch.created_at.strftime('%d.%m.%Y') if branch.created_at else None
            })
        
        return jsonify({
            'success': True,
            'branches': branches_list,
            'total_count': len(branches_list)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500

# Save Customer QR API - QR kodlarını customerQR tablosuna kaydet
@app.route('/api/save-customer-qr', methods=['POST'])
def api_save_customer_qr():
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'error': 'Veri bulunamadı'}), 400
        
        user_id = data.get('user_id')
        qr_code = data.get('qr_code')
        qr_type = data.get('qr_type')  # 'dashboard', 'campaign', 'product_request'
        campaign_id = data.get('campaign_id')
        product_id = data.get('product_id')
        
        if not user_id or not qr_code or not qr_type:
            return jsonify({'success': False, 'error': 'Gerekli alanlar eksik'}), 400
        
        # Kullanıcıyı kontrol et
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'error': 'Kullanıcı bulunamadı'}), 404
        
        # Aynı QR kod zaten mevcut mu kontrol et
        existing_qr = CustomerQRCode.query.filter_by(code=qr_code).first()
        if existing_qr:
            return jsonify({
                'success': True,
                'message': 'QR kod zaten mevcut',
                'qr_id': existing_qr.id,
                'qr_type': qr_type
            })
        
        # QR kod kaydını oluştur
        customer_qr = CustomerQRCode(
            code=qr_code,
            customer_id=user_id,
            points_earned=1,  # Varsayılan puan
            is_used=False
        )
        
        db.session.add(customer_qr)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'QR kod başarıyla kaydedildi',
            'qr_id': customer_qr.id,
            'qr_type': qr_type
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Hata: {str(e)}'}), 500

# Transaction History API - İşlem geçmişini product_redemption ve customerQR tablolarından al
@app.route('/api/transaction-history', methods=['GET'])
def api_transaction_history():
    try:
        user_id = request.args.get('user_id')
        print(f"Transaction history request for user_id: {user_id}")
        
        if not user_id:
            return jsonify({'success': False, 'error': 'Kullanıcı ID gerekli'}), 400
        
        # Kullanıcıyı kontrol et
        user = User.query.get(user_id)
        if not user:
            print(f"User not found: {user_id}")
            return jsonify({'success': False, 'error': 'Kullanıcı bulunamadı'}), 404
        
        print(f"User found: {user.name} ({user.email})")
        transactions = []
        
        # 1. ProductRedemption tablosundan ürün talep/onay işlemlerini al
        redemptions = db.session.query(ProductRedemption, Product).join(
            Product, ProductRedemption.product_id == Product.id
        ).filter(ProductRedemption.user_id == user_id).order_by(
            ProductRedemption.redeemed_at.desc()
        ).all()
        
        print(f"ProductRedemption kayıt sayısı: {len(redemptions)}")
        
        for redemption, product in redemptions:
            transaction_type = "Ürün Onayı" if redemption.is_confirmed else "Ürün Talebi"
            points = -redemption.points_used if redemption.is_confirmed else 0
            description = f"{product.name}"
            if redemption.is_confirmed:
                description += " (Onaylandı)"
            else:
                description += " (Beklemede)"
            
            # Tarih olarak confirmed_at (onaylandıysa) veya redeemed_at kullan
            transaction_date = redemption.confirmed_at if redemption.is_confirmed and redemption.confirmed_at else redemption.redeemed_at
            
            transactions.append({
                'id': f'redemption_{redemption.id}',
                'transaction_type': transaction_type,
                'description': description,
                'points': points,
                'created_at': transaction_date.isoformat() if transaction_date else None,
                'status': 'confirmed' if redemption.is_confirmed else 'pending'
            })
        
        # 2. CustomerQR tablosundan QR kod okutma işlemlerini al
        qr_codes = CustomerQRCode.query.filter_by(customer_id=user_id).order_by(
            CustomerQRCode.created_at.desc()
        ).all()
        
        print(f"CustomerQRCode kayıt sayısı: {len(qr_codes)}")
        
        for qr_code in qr_codes:
            if qr_code.is_used:
                transactions.append({
                    'id': f'qr_{qr_code.id}',
                    'transaction_type': 'QR Kod Okutma',
                    'description': 'QR kod okutularak puan kazanıldı',
                    'points': qr_code.points_earned or 1,
                    'created_at': qr_code.used_at.isoformat() if qr_code.used_at else qr_code.created_at.isoformat(),
                    'status': 'confirmed'
                })
        
        # 3. CustomerQR (eski tablo) tablosundan da QR kod işlemlerini al
        old_qr_codes = CustomerQR.query.filter_by(customer_id=user_id).order_by(
            CustomerQR.created_at.desc()
        ).all()
        
        print(f"CustomerQR (eski) kayıt sayısı: {len(old_qr_codes)}")
        
        for qr_code in old_qr_codes:
            if qr_code.used_at:  # Kullanılmış QR kodlar
                transactions.append({
                    'id': f'old_qr_{qr_code.id}',
                    'transaction_type': 'QR Kod Okutma',
                    'description': 'QR kod okutularak puan kazanıldı',
                    'points': 1,  # Varsayılan puan
                    'created_at': qr_code.used_at.isoformat(),
                    'status': 'confirmed'
                })
        
        # Tüm işlemleri tarihe göre sırala (en yeni önce)
        transactions.sort(key=lambda x: x['created_at'] or '', reverse=True)
        
        # Son 50 işlemi al
        transactions = transactions[:50]
        
        print(f"Toplam işlem sayısı: {len(transactions)}")
        
        return jsonify({
            'success': True,
            'transactions': transactions,
            'total_count': len(transactions)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Hata: {str(e)}'}), 500

# Splash Image API - Mobile app için splash resmi
@app.route('/api/splash-image', methods=['GET'])
def api_splash_image():
    try:
        # Splash resmini al
        splash_setting = SiteSetting.query.filter_by(key='splash_image').first()
        
        if splash_setting and splash_setting.value:
            splash_url = url_for('static', filename=f'uploads/{splash_setting.value}', _external=True)
            return jsonify({
                'success': True,
                'splash_url': splash_url,
                'has_splash': True
            })
        else:
            return jsonify({
                'success': True,
                'splash_url': None,
                'has_splash': False
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': f'Hata: {str(e)}'}), 500

# Change Password API - Şifre değiştirme
@app.route('/api/change-password', methods=['POST'])
def api_change_password():
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'error': 'Veri bulunamadı'}), 400
        
        user_id = data.get('user_id')
        current_password = data.get('current_password')
        new_password = data.get('new_password')
        
        # Gerekli alanları kontrol et
        if not all([user_id, current_password, new_password]):
            return jsonify({'success': False, 'error': 'Tüm alanlar gereklidir'}), 400
        
        # Kullanıcıyı bul
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'error': 'Kullanıcı bulunamadı'}), 404
        
        # Mevcut şifre kontrolü
        if not user.check_password(current_password):
            return jsonify({'success': False, 'error': 'Mevcut şifre yanlış'}), 400
        
        # Yeni şifre uzunluk kontrolü
        if len(new_password) < 6:
            return jsonify({'success': False, 'error': 'Şifre en az 6 karakter olmalıdır'}), 400
        
        # Şifreyi güncelle
        user.set_password(new_password)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Şifreniz başarıyla değiştirildi'
        })
        
    except Exception as e:
        print(f"Change password API error: {str(e)}")
        return jsonify({'success': False, 'error': f'Hata: {str(e)}'}), 500

# Purchase History API - Kullanıcının satın alma geçmişi (puanlama için)
@app.route('/api/purchase-history', methods=['GET'])
def api_purchase_history():
    try:
        user_id = request.args.get('user_id')
        
        if not user_id:
            return jsonify({'success': False, 'error': 'user_id parametresi gereklidir'}), 400
        
        # Kullanıcının onaylanmış ürün alımlarını getir
        redemptions = db.session.query(ProductRedemption, Product, Category).outerjoin(
            Product, ProductRedemption.product_id == Product.id
        ).outerjoin(
            Category, Product.category_id == Category.id
        ).filter(
            ProductRedemption.user_id == user_id,
            ProductRedemption.is_confirmed == True
        ).order_by(ProductRedemption.created_at.desc()).limit(20).all()
        
        purchases = []
        for redemption, product, category in redemptions:
            # Kullanıcının bu ürün için verdiği puanı kontrol et
            rating = ProductRating.query.filter_by(
                user_id=user_id,
                product_id=redemption.product_id
            ).first()
            
            purchase_data = {
                'id': redemption.product_id,
                'name': product.name if product else 'Unknown Product',
                'category_name': category.name if category else None,
                'points_required': product.points_required if product else 0,
                'purchase_date': redemption.created_at.isoformat(),
                'redemption_id': redemption.id,
                'user_rating': rating.rating if rating else None,
                'user_comment': rating.comment if rating else None,
                'image_url': f"/static/product_images/{product.image_filename}" if product and product.image_filename else None
            }
            purchases.append(purchase_data)
        
        return jsonify({
            'success': True,
            'purchases': purchases,
            'total_count': len(purchases)
        })
        
    except Exception as e:
        print(f"Purchase history API error: {str(e)}")
        return jsonify({'success': False, 'error': f'Hata: {str(e)}'}), 500

# Rate Product API - Ürün puanlama
@app.route('/api/rate-product', methods=['POST'])
def api_rate_product():
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'error': 'Veri bulunamadı'}), 400
        
        user_id = data.get('user_id')
        product_id = data.get('product_id')
        rating = data.get('rating')
        comment = data.get('comment', '')
        
        # Gerekli alanları kontrol et
        if not all([user_id, product_id, rating]):
            return jsonify({'success': False, 'error': 'user_id, product_id ve rating alanları gereklidir'}), 400
        
        # Rating değerini kontrol et (1-5 arası)
        try:
            rating = int(rating)
            if rating < 1 or rating > 5:
                return jsonify({'success': False, 'error': 'Puan 1-5 arasında olmalıdır'}), 400
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'Geçersiz puan değeri'}), 400
        
        # Kullanıcı ve ürün kontrolü
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'error': 'Kullanıcı bulunamadı'}), 404
        
        product = Product.query.get(product_id)
        if not product:
            return jsonify({'success': False, 'error': 'Ürün bulunamadı'}), 404
        
        # Kullanıcının bu ürünü satın alıp almadığını kontrol et
        redemption = ProductRedemption.query.filter_by(
            user_id=user_id,
            product_id=product_id,
            is_confirmed=True
        ).first()
        
        if not redemption:
            return jsonify({'success': False, 'error': 'Bu ürünü satın almadığınız için puanlayamazsınız'}), 400
        
        # Mevcut puanı kontrol et
        existing_rating = ProductRating.query.filter_by(
            user_id=user_id,
            product_id=product_id
        ).first()
        
        if existing_rating:
            # Mevcut puanı güncelle
            existing_rating.rating = rating
            existing_rating.comment = comment
            existing_rating.updated_at = datetime.utcnow()
        else:
            # Yeni puan oluştur
            new_rating = ProductRating(
                user_id=user_id,
                product_id=product_id,
                rating=rating,
                comment=comment,
                created_at=datetime.utcnow()
            )
            db.session.add(new_rating)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Puanınız başarıyla kaydedildi',
            'rating': rating,
            'comment': comment
        })
        
    except Exception as e:
        print(f"Rate product API error: {str(e)}")
        return jsonify({'success': False, 'error': f'Hata: {str(e)}'}), 500

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    
    debug_mode = os.environ.get('FLASK_ENV') != 'production'
    print("HTTP modu aktif - Port 1519")
    app.run(debug=debug_mode, host='0.0.0.0', port=1519)
